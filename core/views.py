import json
import zipfile
import io
import logging
import threading
import zlib
from contextlib import contextmanager
from fractions import Fraction


def _parse_diametro(d):
    """Converte string de diâmetro (ex: '2 1/2"') para float."""
    clean = d.replace('"', '').strip()
    parts = clean.split()
    if len(parts) == 2:
        return int(parts[0]) + float(Fraction(parts[1]))
    return float(Fraction(parts[0]))
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth
from django.db import models as db_models, transaction, connection
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.conf import settings
from django.core.cache import cache
import jwt as pyjwt
import os
import zoneinfo
from functools import wraps
zone = zoneinfo.ZoneInfo('America/Sao_Paulo')


# Vedação corpo/tampa permitida por classe quando Gaveta/Retenção + NBR 15827
_VED_NBR_POR_CLASSE = {
    "150":  ["JUNTA ESPIRALADA"],
    "300":  ["JUNTA ESPIRALADA"],
    "600":  ["JUNTA ESPIRALADA", "RTJ (FJA)", "PRESSURE SEAL"],
    "800":  ["JUNTA ESPIRALADA", "CASTELO SOLDADO"],
    "900":  ["JUNTA ESPIRALADA", "RTJ (FJA)", "PRESSURE SEAL", "CASTELO SOLDADO"],
    "1500": ["JUNTA ESPIRALADA", "RTJ (FJA)", "PRESSURE SEAL", "CASTELO SOLDADO"],
    "2500": ["RTJ (FJA)", "PRESSURE SEAL", "CASTELO SOLDADO"],
}
_VED_NBR_LABELS = {
    "JUNTA ESPIRALADA": "Espiralada",
    "RTJ (FJA)": "FJA",
    "PRESSURE SEAL": "Pressure Seal",
    "CASTELO SOLDADO": "Castelo Soldado",
}


def _junta_categoria(data):
    """Valor de 'Categoria da Junta' (materiais, tipo_material=JUNTA) — pra Gaveta/Globo/
    Retenção/Globo Controle é o corpo/castelo (Junta Espiralada/RTJ/Pressure Seal/Castelo
    Soldado), lido pelas regras de norma no lugar da Vedação Sede/Tampa (campo exclusivo
    de Esfera, ver templates/core/index.html vedacao-container)."""
    for m in data.get("materiais", []):
        if m.get("tipo_material") == "JUNTA":
            return m.get("material") or ""
    return ""


def _validar_vedacao_nbr_classe(tipo_valvula, data):
    """Para Gaveta/Retenção com NBR 15827, restringe a categoria da junta (corpo/castelo)
    conforme a classe. Retorna JsonResponse de erro (400) se inválido, ou None se ok."""
    if tipo_valvula not in ("GAVETA", "RETENCAO") or not data.get("nbr"):
        return None
    classe = data.get("classe")
    permitidas = _VED_NBR_POR_CLASSE.get(classe)
    if not permitidas:
        return None
    tipo_nome = "Gaveta" if tipo_valvula == "GAVETA" else "Retenção"
    ved_val = _junta_categoria(data)
    if ved_val and ved_val not in permitidas:
        opts = ", ".join(_VED_NBR_LABELS.get(v, v) for v in permitidas)
        return JsonResponse({"success": False, "errors": {"materiais": f"Para {tipo_nome} com NBR 15827 e classe {classe}, a categoria da junta deve ser: {opts}"}}, status=400)
    return None


def especial_required(view_func):
    """Bloqueia acesso a quem não for ESPECIAL. Retorna 403 JSON."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.nivel_permissao != "ESPECIAL":
            return JsonResponse({"success": False, "errors": {"__all__": "Sem permissão"}}, status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped

from .models import Valvula, ValvulaMaterial, Vedacao, ComponentesInternos, Material, Tb_Usuario, Projeto, AnexoValvula, TentativaDuplicata, OpcaoFlange, OpcaoPlacaIdentificacao
from . import storage
from .forms import (
    ValvulaForm,
    ValvulaMaterialFormSet,
    VedacaoFormSet,
    ComponentesInternosFormSet,
    MaterialForm,
    PesquisaForm,
)
import uuid
from datetime import timedelta
from django.core.mail import EmailMessage
from django.conf import settings
from django.core.mail import EmailMultiAlternatives

# Try to import openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Mapeamento do tipo de válvula para o prefixo do código (3 letras em inglês)
CODIGO_PREFIXO = {
    "ESFERA": "VES",          # Ball
    "GAVETA": "VGA",          # Gate
    "GLOBO": "VGL",           # Globe
    "RETENCAO": "VRE",        # Check
    "BORBOLETA": "VBO",       # Butterfly
    "GLOBO_CONTROLE": "VGC",  # Globe Control
}

TIPO_LABEL = dict(Valvula.TIPO_VALVULA)

# Acionamentos manuais: sem posição de falha e sem dados elétricos (tensão/fase/frequência)
ACIONAMENTOS_MANUAIS = ("ALAVANCA", "VOLANTE", "VOLANTE COM ENGRENAGEM DE REDUÇÃO")


def gerar_codigo(tipo_valvula):
    """Gera código automático: 3 letras do tipo em inglês + 6 dígitos sequenciais."""
    prefixo = CODIGO_PREFIXO.get(tipo_valvula, "VAL")
    existing = Valvula.objects.filter(codigo__startswith=prefixo)
    max_num = 0
    for v in existing:
        try:
            num_str = v.codigo[len(prefixo):]
            num = int(num_str)
            if num > max_num:
                max_num = num
        except (ValueError, TypeError):
            continue
    return f"{prefixo}{max_num + 1:06d}"


@contextmanager
def _lock_tipo_valvula(tipo_valvula):
    """Trava por tipo de válvula. Serializa dup-check + gerar_codigo + save entre
    requests concorrentes, evitando válvulas duplicadas (TOCTOU) e colisão de
    código (unique). Chave determinística entre workers (crc32). Usar como
    context manager dentro de uma transação:
        with transaction.atomic(), _lock_tipo_valvula(tipo):
            ...

    Postgres: advisory lock transacional (pg_advisory_xact_lock) — libera sozinho
    no commit/rollback, sem precisar de finally aqui.
    MySQL: named lock (GET_LOCK/RELEASE_LOCK) — é por sessão, não por transação,
    então libera explicitamente no finally (cobre commit e rollback/exceção).
    Outros backends (ex.: SQLite nos testes): no-op, sem concorrência real ali."""
    if connection.vendor == "postgresql":
        key = zlib.crc32(f"valvula_criar:{tipo_valvula}".encode())
        with connection.cursor() as cur:
            cur.execute("SELECT pg_advisory_xact_lock(%s)", [key])
        yield
    elif connection.vendor == "mysql":
        nome_lock = f"valvula_criar:{tipo_valvula}"
        with connection.cursor() as cur:
            cur.execute("SELECT GET_LOCK(%s, 10)", [nome_lock])
        try:
            yield
        finally:
            with connection.cursor() as cur:
                cur.execute("SELECT RELEASE_LOCK(%s)", [nome_lock])
    else:
        yield


# ── Página principal (single-page) ──────────────────────────────────────────

def auth_page(request):
    """Renderiza a tela de login/cadastro."""
    return render(request, "core/auth.html")


@require_POST
def login_api(request):
    """Autentica usuário verificando email e senha no banco."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "errors": {"__all__": "Dados inválidos"}}, status=400)

    email = data.get("email", "").strip()
    password = data.get("password", "")

    if not email or not password:
        return JsonResponse({"success": False, "errors": {"__all__": "Preencha todos os campos"}}, status=400)

    user = authenticate(request, username=email, password=password)
    if user is not None and not user.confirmado:
        return JsonResponse({"success": False, "errors": {"__all__": "Conta ainda não foi confirmada. Verifique seu email para ativar sua conta."}}, status=403)

    if user is None:
        return JsonResponse({"success": False, "errors": {"__all__": "Email ou senha incorretos"}}, status=401)

    login(request, user)
    return JsonResponse({"success": True, "nome": user.nome, "nivel_permissao": user.nivel_permissao})


def sso_login(request):
    """Login automatico vindo do sistema maior (TypeScript/React) via JWT.

    Token = JWT HS256 padrao assinado com SSO_SHARED_SECRET (nao a SECRET_KEY do
    Django — segredo proprio, compartilhado so com o sistema que gera o link).
    Claims esperadas: email, nome (opcional), exp (curto, ex. 60s), jti (unico por
    link — o sistema de origem gera, ex. crypto.randomUUID()). `jti` e' marcado no
    cache no primeiro uso p/ um link antigo/reaproveitado nao logar de novo.
    Cria o usuario automaticamente (COMUM, ja confirmado) se o email ainda nao existir.

    Exemplo de emissao em TypeScript (lib "jsonwebtoken"):
        jwt.sign({ email, nome, jti: crypto.randomUUID() }, SSO_SHARED_SECRET,
                 { algorithm: "HS256", expiresIn: "60s" })
    """
    token = request.GET.get("token", "")
    if not token or not settings.SSO_SHARED_SECRET:
        return HttpResponse("Link de acesso invalido.", status=400)

    try:
        payload = pyjwt.decode(token, settings.SSO_SHARED_SECRET, algorithms=["HS256"])
    except pyjwt.ExpiredSignatureError:
        return HttpResponse("Link de acesso expirado. Peca um novo link no sistema de origem.", status=400)
    except pyjwt.InvalidTokenError:
        return HttpResponse("Link de acesso invalido.", status=400)

    jti = payload.get("jti")
    if not jti:
        return HttpResponse("Link de acesso invalido.", status=400)

    cache_key = f"sso_token_usado:{jti}"
    if not cache.add(cache_key, True, timeout=120):
        return HttpResponse("Link de acesso ja foi usado. Peca um novo link no sistema de origem.", status=400)

    email = (payload.get("email") or "").strip().lower()
    if not email or not email.endswith("@imexsolutions.com.br"):
        return HttpResponse("Link de acesso invalido.", status=400)

    nome = (payload.get("nome") or "").strip() or email.split("@")[0]

    usuario = Tb_Usuario.objects.filter(email=email).first()
    if usuario is None:
        usuario = Tb_Usuario.objects.create_user(
            email=email,
            nome=nome,
            password=None,
            confirmado=True,
        )
    elif not usuario.confirmado:
        usuario.confirmado = True
        usuario.save(update_fields=["confirmado"])

    login(request, usuario, backend="django.contrib.auth.backends.ModelBackend")
    return redirect("core:index")


logger = logging.getLogger(__name__)


def _enviar_email_async(email_msg, destinatario):
    """Envia EmailMessage numa thread separada p/ nao bloquear o request (SMTP ~5-11s)."""
    try:
        email_msg.send(fail_silently=False)
    except Exception:
        logger.exception("Falha ao enviar email de verificacao para %s", destinatario)


@csrf_exempt
@require_POST
def cadastro_api(request):
    """Cria novo usuário e envia email de verificação."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "errors": {"__all__": "Dados inválidos"}}, status=400)

    nome = data.get("nome", "").strip()
    email = data.get("email", "").strip()
    password = data.get("password", "")
    password_confirm = data.get("password_confirm", "")

    errors = {}
    if not nome:
        errors["nome"] = "Nome é obrigatório"
    if not email:
        errors["email"] = "Email é obrigatório"
    elif not email.lower().endswith("@imexsolutions.com.br"):
        errors["email"] = "Cadastro permitido só com email @imexsolutions.com.br"
    if not password:
        errors["password"] = "Senha é obrigatória"
    elif len(password) < 8:
        errors["password"] = "Senha deve ter no mínimo 8 caracteres"
    if password != password_confirm:
        errors["password_confirm"] = "As senhas não coincidem"
    if email:
        usuario_existente = Tb_Usuario.objects.filter(email=email).first()
        if usuario_existente and usuario_existente.confirmado:
            errors["email"] = "Este email já está cadastrado"

    if errors:
        return JsonResponse({"success": False, "errors": errors}, status=400)

    token = uuid.uuid4()

    # Se email existe mas não confirmado → atualiza dados e reenviar email
    usuario_existente = Tb_Usuario.objects.filter(email=email, confirmado=False).first()
    if usuario_existente:
        usuario_existente.nome = nome
        usuario_existente.token_verificacao = token
        usuario_existente.set_password(password)
        usuario_existente.save()
        usuario = usuario_existente
        reenvio = True
    else:
        usuario = Tb_Usuario.objects.create_user(
            email=email,
            nome=nome,
            password=password,
            token_verificacao=token,
            confirmado=False,
        )
        reenvio = False

    link = f"{settings.SITE_URL}/verificar-email/{token}/"

    try:
        html = f"""
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
        <meta charset="UTF-8">
        <meta http-equiv="X-UA-Compatible" content="IE=edge">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Confirmação de Cadastro</title>
        </head>

        <body style="margin:0;padding:0;background-color:#f4f6f9;font-family:Arial,Helvetica,sans-serif;">

        <table width="100%" border="0" cellspacing="0" cellpadding="0" bgcolor="#f4f6f9">
        <tr>
        <td align="center" style="padding:30px 15px;">

        <table width="600" border="0" cellspacing="0" cellpadding="0" bgcolor="#ffffff" style="border:1px solid #dddddd;">

            <!-- Header -->
            <tr>
                <td align="center" bgcolor="#0b1629" style="padding:35px 20px;">

                    <h1 style="
                        margin:0;
                        color:#ffffff;
                        font-size:30px;
                        font-family:Arial,Helvetica,sans-serif;
                        font-weight:bold;
                    ">
                        Imex Solutions
                    </h1>

                    <p style="
                        margin:12px 0 0;
                        color:#c8d2e3;
                        font-size:15px;
                    ">
                        Confirmação de Cadastro
                    </p>

                </td>
            </tr>

            <!-- Body -->
            <tr>
                <td style="padding:40px;">

                    <h2 style="
                        margin-top:0;
                        margin-bottom:20px;
                        color:#0b1629;
                        font-size:24px;
                        font-weight:bold;
                    ">
                        Olá, {nome}
                    </h2>

                    <p style="
                        color:#555555;
                        font-size:16px;
                        line-height:26px;
                        margin:0 0 18px;
                    ">
                        Seja bem-vindo(a) à
                        <strong>Imex Solutions</strong>.
                    </p>

                    <p style="
                        color:#555555;
                        font-size:16px;
                        line-height:26px;
                        margin:0 0 18px;
                    ">
                        Recebemos sua solicitação de cadastro e falta apenas um passo
                        para ativar sua conta.
                    </p>

                    <p style="
                        color:#555555;
                        font-size:16px;
                        line-height:26px;
                        margin:0 0 30px;
                    ">
                        Clique no botão abaixo para confirmar seu endereço de e-mail:
                    </p>

                    <!-- Botão -->
                    <table align="center" border="0" cellspacing="0" cellpadding="0">
                        <tr>
                            <td align="center"
                                bgcolor="#2c5aa0"
                                style="padding:16px 34px;">

                                <a href="{link}"
                                target="_blank"
                                style="
                                        color:#ffffff;
                                        text-decoration:none;
                                        font-size:16px;
                                        font-weight:bold;
                                        font-family:Arial,Helvetica,sans-serif;
                                        display:inline-block;
                                ">
                                    Confirmar meu cadastro
                                </a>

                            </td>
                        </tr>
                    </table>
                    <table width="100%" border="0" cellspacing="0" cellpadding="0" bgcolor="#f8f9fb" style="border-left:4px solid #2c5aa0;margin-top:30px;">
                        <tr>
                            <td style="padding:18px;">

                                <strong style="color:#0b1629;font-size:15px;">
                                    Não foi você?
                                </strong>

                                <p style="
                                    margin:10px 0 0;
                                    color:#666666;
                                    font-size:14px;
                                    line-height:22px;
                                ">
                                    Se você não realizou este cadastro, basta ignorar
                                    este e-mail. Nenhuma ação será necessária.
                                </p>

                            </td>
                        </tr>
                    </table>

                    <p style="
                        margin-top:35px;
                        color:#555555;
                        font-size:15px;
                        line-height:24px;
                    ">
                        Atenciosamente,<br>
                        <strong>Equipe Imex Solutions</strong>
                    </p>

                </td>
            </tr>

            <!-- Footer -->
            <tr>
                <td align="center"
                    bgcolor="#f4f6f9"
                    style="
                        padding:20px;
                        color:#888888;
                        font-size:12px;
                    ">

                    © 2026 Imex Solutions<br>
                    Este é um e-mail automático. Não responda esta mensagem.

                </td>
            </tr>

        </table>

        </td>
        </tr>
        </table>

        </body>
        </html>
        """

        email_msg = EmailMessage(
            subject="Confirme seu cadastro — Imex Solutions",
            body=html,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email],
        )

        email_msg.content_subtype = "html"

    except Exception:
        if not reenvio:
            usuario.delete()
        return JsonResponse({
            "success": False,
            "errors": {
                "__all__": "Não foi possível preparar o email de verificação. Tente novamente."
            }
        }, status=500)

    # Envia o email em background: SMTP leva 5-11s e travava o request inteiro.
    # A resposta volta na hora; falha de envio é logada (usuário pode reenviar pelo cadastro).
    threading.Thread(target=_enviar_email_async, args=(email_msg, email), daemon=True).start()

    msg = "Email de verificação reenviado! Verifique sua caixa de entrada." if reenvio else "Cadastro realizado! Verifique seu email para ativar sua conta."
    return JsonResponse({"success": True, "message": msg})



def logout_api(request):
    """Faz logout do usuário."""
    logout(request)
    return redirect("core:auth_page")


def usuarios_page(request):
    """Página de gerenciamento de usuários — só ESPECIAL."""
    if not request.user.is_authenticated or request.user.nivel_permissao != "ESPECIAL":
        return redirect("core:index")
    return render(request, "core/usuarios.html")


def estatisticas_page(request):
    """Página de estatísticas — só ESPECIAL."""
    if not request.user.is_authenticated or request.user.nivel_permissao != "ESPECIAL":
        return redirect("core:index")
    return render(request, "core/estatisticas.html")


def _stats_basicas(qs):
    """Estatísticas derivadas de um queryset de válvulas (reusado p/ comparativo)."""
    total = qs.count()
    por_tipo = list(
        qs.values("tipo_valvula").annotate(total=Count("id_valvula")).order_by("-total")
    )
    por_tipo = [
        {"tipo_key": r["tipo_valvula"], "tipo": TIPO_LABEL.get(r["tipo_valvula"], r["tipo_valvula"]), "total": r["total"]}
        for r in por_tipo
    ]
    nbr_sim = qs.filter(nbr=True).count()
    return {
        "total": total,
        "por_tipo": por_tipo,
        "nbr_sim": nbr_sim,
        "nbr_nao": total - nbr_sim,
        "nbr_pct": round(nbr_sim * 100.0 / total, 1) if total else 0,
    }


def _top(qs, field, limit=10, label_map=None):
    """Top N valores de um campo (ignora vazios/nulos)."""
    rows = (
        qs.exclude(**{f"{field}__in": ["", None]})
        .values(field).annotate(total=Count("id_valvula")).order_by("-total")[:limit]
    )
    out = []
    for r in rows:
        v = r[field]
        out.append({"label": label_map.get(v, v) if label_map else (v or "—"), "total": r["total"]})
    return out


@require_GET
def estatisticas_api(request):
    """Dados agregados para a página de estatísticas — só ESPECIAL.
    Filtros (querystring): projeto, tipo, data_inicio, data_fim, projeto_a, projeto_b."""
    if not request.user.is_authenticated or request.user.nivel_permissao != "ESPECIAL":
        return JsonResponse({"error": "Sem permissão"}, status=403)

    # ── Queryset base + filtros ──────────────────────────────────────────────
    qs = Valvula.objects.all()
    f_projeto = request.GET.get("projeto", "").strip()
    f_tipo = request.GET.get("tipo", "").strip()
    f_ini = request.GET.get("data_inicio", "").strip()
    f_fim = request.GET.get("data_fim", "").strip()
    if f_projeto:
        qs = qs.filter(projetos__pk=f_projeto)
    if f_tipo:
        qs = qs.filter(tipo_valvula=f_tipo)
    if f_ini:
        qs = qs.filter(criado_em__date__gte=f_ini)
    if f_fim:
        qs = qs.filter(criado_em__date__lte=f_fim)

    agora = timezone.now().astimezone(zone)
    inicio_mes = agora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    inicio_ano = agora.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

    # ── Contadores ───────────────────────────────────────────────────────────
    base = _stats_basicas(qs)
    qs_mes = qs.filter(criado_em__gte=inicio_mes)
    qs_ano = qs.filter(criado_em__gte=inicio_ano)
    contadores = {
        "total_valvulas": base["total"],
        "total_materiais": Material.objects.count(),
        "total_projetos": Projeto.objects.count(),
        "projetos_ativos": Projeto.objects.filter(valvulas__isnull=False).distinct().count(),
        "criadas_mes": qs_mes.count(),
        "criadas_ano": qs_ano.count(),
        "criadas_sempre": base["total"],
        "por_tipo": base["por_tipo"],
        "por_tipo_mes": _stats_basicas(qs_mes)["por_tipo"],
        "por_tipo_ano": _stats_basicas(qs_ano)["por_tipo"],
        "nbr_sim": base["nbr_sim"],
        "nbr_nao": base["nbr_nao"],
        "nbr_pct": base["nbr_pct"],
        "duplicatas": TentativaDuplicata.objects.count(),
    }

    # ── Séries temporais (últimos 12 meses) ──────────────────────────────────
    primeiro_mes = (inicio_mes - timedelta(days=365)).replace(day=1)
    meses = []
    cur = primeiro_mes
    for _ in range(13):
        meses.append(cur)
        cur = (cur + timedelta(days=32)).replace(day=1)
    meses_labels = [m.strftime("%m/%Y") for m in meses]
    meses_keys = [m.strftime("%Y-%m") for m in meses]

    qs_12m = qs.filter(criado_em__gte=primeiro_mes)
    # total por mês
    por_mes_raw = {}
    for r in qs_12m.annotate(m=TruncMonth("criado_em")).values("m").annotate(total=Count("id_valvula")):
        if r["m"]:
            por_mes_raw[r["m"].astimezone(zone).strftime("%Y-%m")] = r["total"]
    serie_mensal = [por_mes_raw.get(k, 0) for k in meses_keys]

    # empilhado por tipo
    tipos_keys = [t[0] for t in Valvula.TIPO_VALVULA]
    empilhado_raw = {}
    for r in qs_12m.annotate(m=TruncMonth("criado_em")).values("m", "tipo_valvula").annotate(total=Count("id_valvula")):
        if r["m"]:
            k = r["m"].astimezone(zone).strftime("%Y-%m")
            empilhado_raw[(k, r["tipo_valvula"])] = r["total"]
    empilhado = {
        t: [empilhado_raw.get((k, t), 0) for k in meses_keys] for t in tipos_keys
    }

    # tendência acumulada (sobre todo o período filtrado)
    acumulado = []
    soma = 0
    for v in serie_mensal:
        soma += v
        acumulado.append(soma)

    series = {
        "meses_labels": meses_labels,
        "serie_mensal": serie_mensal,
        "empilhado": empilhado,
        "tipos_labels": {t: TIPO_LABEL.get(t, t) for t in tipos_keys},
        "acumulado": acumulado,
    }

    # ── Por projeto ──────────────────────────────────────────────────────────
    ranking_proj = list(
        qs.filter(projetos__isnull=False)
        .values("projetos__nome").annotate(total=Count("id_valvula")).order_by("-total")[:15]
    )
    ranking_projetos = [{"nome": r["projetos__nome"], "total": r["total"]} for r in ranking_proj]
    orfaos = list(
        Projeto.objects.filter(valvulas__isnull=True).values_list("nome", flat=True)
    )
    sem_projeto = qs.filter(projetos__isnull=True).count()

    # distribuição de tipos dentro de cada projeto (top projetos)
    dist_proj = {}
    for r in qs.filter(projetos__isnull=False).values("projetos__nome", "tipo_valvula").annotate(total=Count("id_valvula")):
        dist_proj.setdefault(r["projetos__nome"], {})[TIPO_LABEL.get(r["tipo_valvula"], r["tipo_valvula"])] = r["total"]

    por_projeto = {
        "ranking": ranking_projetos,
        "orfaos": orfaos,
        "sem_projeto": sem_projeto,
        "distribuicao": dist_proj,
    }

    # ── Por especificação técnica ────────────────────────────────────────────
    materiais_corpo = list(
        ValvulaMaterial.objects.filter(tipo_material="CORPO_TAMPA", valvula__in=qs)
        .values("material__nome").annotate(total=Count("id")).order_by("-total")[:10]
    )
    materiais_corpo = [{"label": r["material__nome"], "total": r["total"]} for r in materiais_corpo]

    combinacoes_raw = (
        qs.exclude(classe="").exclude(diametro="")
        .values("tipo_valvula", "classe", "diametro").annotate(total=Count("id_valvula")).order_by("-total")[:10]
    )
    combinacoes = [
        {
            "label": f"{TIPO_LABEL.get(r['tipo_valvula'], r['tipo_valvula'])} · {r['classe']} · {r['diametro']}",
            "total": r["total"],
        }
        for r in combinacoes_raw
    ]

    por_spec = {
        "diametros": _top(qs, "diametro"),
        "classes": _top(qs, "classe"),
        "extremidades": _top(qs, "tipo_extremidade"),
        "normas": _top(qs, "norma"),
        "materiais_corpo": materiais_corpo,
        "combinacoes": combinacoes,
    }

    # ── Por usuário ──────────────────────────────────────────────────────────
    ranking_user = list(
        qs.values("criado_por__nome").annotate(total=Count("id_valvula")).order_by("-total")
    )
    ranking_usuarios = [
        {"nome": r["criado_por__nome"] or "Desconhecido", "total": r["total"]} for r in ranking_user
    ]
    user_mes = list(
        qs_mes.values("criado_por__nome").annotate(total=Count("id_valvula")).order_by("-total")
    )
    usuarios_mes = [
        {"nome": r["criado_por__nome"] or "Desconhecido", "total": r["total"]} for r in user_mes
    ]
    por_usuario = {"ranking": ranking_usuarios, "mes": usuarios_mes}

    # ── Comparativo entre 2 projetos ─────────────────────────────────────────
    comparativo = None
    pa = request.GET.get("projeto_a", "").strip()
    pb = request.GET.get("projeto_b", "").strip()
    if pa and pb:
        def _bloco(pid):
            p = Projeto.objects.filter(pk=pid).first()
            s = _stats_basicas(Valvula.objects.filter(projetos__pk=pid))
            s["nome"] = p.nome if p else "—"
            return s
        comparativo = {"a": _bloco(pa), "b": _bloco(pb)}

    # ── Opções p/ filtros ────────────────────────────────────────────────────
    opcoes = {
        "projetos": list(Projeto.objects.order_by("nome").values("id_projeto", "nome")),
        "tipos": [{"key": t[0], "label": t[1]} for t in Valvula.TIPO_VALVULA],
    }

    return JsonResponse({
        "contadores": contadores,
        "series": series,
        "por_projeto": por_projeto,
        "por_spec": por_spec,
        "por_usuario": por_usuario,
        "comparativo": comparativo,
        "opcoes": opcoes,
    })


@require_GET
def debug_verification_tokens(request):
    """Endpoint de debug: retorna tokens de verificação dos usuários não confirmados."""
    if not settings.DEBUG:
        return JsonResponse({"error": "Não disponível em produção"}, status=403)
    
    usuarios_pendentes = Tb_Usuario.objects.filter(confirmado=False).values('email', 'nome', 'token_verificacao')
    links = []
    for u in usuarios_pendentes:
        token = u['token_verificacao']
        link = f"{settings.SITE_URL}/verificar-email/{token}/"
        links.append({
            "email": u['email'],
            "nome": u['nome'],
            "token": str(token),
            "verificacao_link": link
        })
    return JsonResponse({"usuarios_pendentes": links})

def verificar_email(request, token):
    """Confirma a conta do usuário via token enviado por email."""
    try:
        usuario = Tb_Usuario.objects.get(token_verificacao=token)
    except Tb_Usuario.DoesNotExist:
        return render(request, "core/verificacao_resultado.html", {
            "sucesso": False,
            "mensagem": "Link de verificação inválido ou já utilizado.",
        })

    if usuario.confirmado:
        return render(request, "core/verificacao_resultado.html", {
            "sucesso": False,
            "mensagem": "Esta conta já foi confirmada anteriormente. Faça login normalmente.",
        })

    # Confirma, ativa a conta e gera novo token (invalida o link usado)
    usuario.confirmado = True
    usuario.is_active = True
    usuario.token_verificacao = uuid.uuid4()
    usuario.save()

    return render(request, "core/verificacao_resultado.html", {
        "sucesso": True,
        "mensagem": f"Email confirmado com sucesso! Bem-vindo(a), {usuario.nome}. Você já pode fazer login.",
    })
@require_GET
def usuario_lista_api(request):
    """Lista paginada de usuários — só ESPECIAL."""
    if not request.user.is_authenticated or request.user.nivel_permissao != "ESPECIAL":
        return JsonResponse({"error": "Sem permissão"}, status=403)

    usuarios = Tb_Usuario.objects.all().order_by("-confirmado", "nome")
    busca = request.GET.get("q", "").strip()
    if busca:
        usuarios = usuarios.filter(nome__icontains=busca)

    page, paginator, page_num, page_size = _paginar_queryset(request, usuarios)
    data = [
        {
            "id": u.id,
            "nome": u.nome,
            "email": u.email,
            "nivel_permissao": u.nivel_permissao,
            "nivel_display": u.get_nivel_permissao_display(),
            "confirmado": u.confirmado,
            "is_self": u.id == request.user.id,
        }
        for u in page.object_list
    ]
    return JsonResponse({
        "meu_id": request.user.id,
        "usuarios": data,
        "total": paginator.count,
        "page": page_num,
        "page_size": page_size,
        "total_pages": paginator.num_pages,
        "has_next": page.has_next(),
        "has_previous": page.has_previous(),
    })


@require_POST
def usuario_alterar_permissao(request, pk):
    """Altera nível de permissão de um usuário — só ESPECIAL."""
    if not request.user.is_authenticated or request.user.nivel_permissao != "ESPECIAL":
        return JsonResponse({"error": "Sem permissão"}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Dados inválidos"}, status=400)

    novo_nivel = data.get("nivel_permissao")
    if novo_nivel not in dict(Tb_Usuario.NIVEL_PERMISSAO):
        return JsonResponse({"error": "Nível inválido"}, status=400)

    usuario = get_object_or_404(Tb_Usuario, pk=pk)
    if usuario.id == request.user.id:
        return JsonResponse({"error": "Não é possível alterar a própria permissão"}, status=400)

    usuario.nivel_permissao = novo_nivel
    usuario.save()
    return JsonResponse({"success": True, "nivel_permissao": novo_nivel, "nivel_display": usuario.get_nivel_permissao_display()})


@require_POST
def usuario_confirmar(request, pk):
    """Confirma/desconfirma um usuário — só ESPECIAL."""
    if not request.user.is_authenticated or request.user.nivel_permissao != "ESPECIAL":
        return JsonResponse({"error": "Sem permissão"}, status=403)

    usuario = get_object_or_404(Tb_Usuario, pk=pk)
    usuario.confirmado = not usuario.confirmado
    usuario.save()
    return JsonResponse({"success": True, "confirmado": usuario.confirmado})


@require_GET
def dashboard_contadores_api(request):
    """Contadores do dashboard (total + por tipo) — usado p/ auto-refresh via JS."""
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Não autenticado"}, status=403)
    total_valvulas = Valvula.objects.count()
    por_tipo = (
        Valvula.objects.values("tipo_valvula")
        .annotate(total=Count("id_valvula"))
        .order_by("-total")
    )
    por_tipo_display = [
        {
            "tipo": TIPO_LABEL.get(item["tipo_valvula"], item["tipo_valvula"]),
            "tipo_key": item["tipo_valvula"],
            "total": item["total"],
        }
        for item in por_tipo
    ]
    return JsonResponse({"total_valvulas": total_valvulas, "por_tipo": por_tipo_display})


def index(request):
    """Renderiza a tela única com cadastro, listagem e visualização."""
    if not request.user.is_authenticated:
        return redirect("core:auth_page")
    total_valvulas = Valvula.objects.count()
    total_materiais = Material.objects.count()
    por_tipo = (
        Valvula.objects.values("tipo_valvula")
        .annotate(total=Count("id_valvula"))
        .order_by("-total")
    )
    por_tipo_display = [
        {
            "tipo": TIPO_LABEL.get(item["tipo_valvula"], item["tipo_valvula"]),
            "tipo_key": item["tipo_valvula"],
            "total": item["total"],
        }
        for item in por_tipo
    ]
    valvulas = Valvula.objects.select_related("criado_por").prefetch_related("projetos").order_by("-criado_em")[:50]
    context = {
        "total_valvulas": total_valvulas,
        "total_materiais": total_materiais,
        "por_tipo": por_tipo_display,
        "valvulas": valvulas,
        "tipos_valvula": Valvula.TIPO_VALVULA,
        "cores_json": mark_safe(json.dumps([c[0] for c in Valvula.COR]).replace("<", "\\u003c")),
        "ip_json": mark_safe(json.dumps([c[0] for c in Valvula.IP]).replace("<", "\\u003c")),
    }
    return render(request, "core/index.html", context)


# ── API Endpoints ────────────────────────────────────────────────────────────

def _paginar_queryset(request, queryset):
    """Pagina um queryset via params ?page= e ?page_size= (default 15, máx 100)."""
    try:
        page_size = min(max(int(request.GET.get("page_size", 15)), 1), 100)
    except (TypeError, ValueError):
        page_size = 15
    try:
        page_num = max(int(request.GET.get("page", 1)), 1)
    except (TypeError, ValueError):
        page_num = 1

    paginator = Paginator(queryset, page_size)
    page_num = min(page_num, paginator.num_pages) if paginator.num_pages else 1
    page = paginator.page(page_num)
    return page, paginator, page_num, page_size


def _serializar_valvula_resumo(v):
    return {
        "id": v.id_valvula,
        "codigo": v.codigo,
        "tipo_valvula": v.tipo_valvula,
        "tipo_label": TIPO_LABEL.get(v.tipo_valvula, v.tipo_valvula),
        "norma": v.norma,
        "diametro": v.diametro,
        "classe": v.classe,
        "projetos": [{"id": p.id_projeto, "nome": p.nome} for p in v.projetos.all()],
        "criado_em": v.criado_em.astimezone(zone).strftime("%d/%m/%Y %H:%M") if v.criado_em else "",
        "criado_por": v.criado_por.nome if v.criado_por_id else "",
    }


@require_GET
def valvula_lista_api(request):
    """Retorna lista paginada de válvulas em JSON para a tabela."""
    queryset = Valvula.objects.select_related("criado_por").prefetch_related("projetos").all()
    codigo = request.GET.get("codigo", "").strip()
    tipo = request.GET.get("tipo_valvula", "").strip()
    projeto = request.GET.get("projeto", "").strip()
    if codigo:
        queryset = queryset.filter(codigo__icontains=codigo)
    if tipo:
        queryset = queryset.filter(tipo_valvula=tipo)
    if projeto:
        queryset = queryset.filter(projetos__pk=projeto)
    queryset = queryset.order_by("-criado_em")

    page, paginator, page_num, page_size = _paginar_queryset(request, queryset)
    data = [_serializar_valvula_resumo(v) for v in page.object_list]
    return JsonResponse({
        "valvulas": data,
        "total": paginator.count,
        "page": page_num,
        "page_size": page_size,
        "total_pages": paginator.num_pages,
        "has_next": page.has_next(),
        "has_previous": page.has_previous(),
    })


@require_GET
def pesquisa_avancada_api(request):
    """Pesquisa avançada: filtra válvulas pelos campos preenchidos (AND lógico)."""
    queryset = Valvula.objects.select_related("criado_por").prefetch_related("projetos").all()

    # Mapeamento campo -> tipo de filtro
    TEXT_FIELDS_EXACT = [
        "tipo_valvula", "fabricante", "norma", "diametro", "classe", "tipo_extremidade",
        "tipo_ranhura", "tipo_montagem", "tipo_passagem", "tipo_acionamento", "marca_atuador", "flange_acoplamento",
        "construcao_corpo", "dib", "tipo_castelo", "juncao_corpo_castelo",
        "tipo_retencao", "configuracao_corpo_retencao", "orientacao_instalacao", "categoria_594",
        "uso_geral", "certificacao_sil", "nace", "revestimento", "categoria_borboleta",
        "face_a_face", "configuracao_disco", "posicionador",
        "ip_posicionador", "ip_solenoide", "ip_chave_fim_curso", "ip_sensor_posicao", "filtro",
        "indicador_posicao", "tubing",
        "chave_fim_curso", "valvula_solenoide", "valvula_lock_up", "sensor_posicao", "valvula_escape_rapido",
    ]
    BOOL_FIELDS = [
        "nbr", "valvula_alivio", "dispositivo_antiestatico",
        "baixa_emissao_fugitiva", "hot_disconnect", "contra_peso",
    ]

    for campo in TEXT_FIELDS_EXACT:
        valor = request.GET.get(campo, "").strip()
        if valor:
            queryset = queryset.filter(**{campo: valor})

    for campo in BOOL_FIELDS:
        valor = request.GET.get(campo, "").strip()
        if valor == "true":
            queryset = queryset.filter(**{campo: True})
        elif valor == "false":
            queryset = queryset.filter(**{campo: False})

    # PMT é texto livre → busca parcial (icontains)
    classe_pmt = request.GET.get("classe_pmt", "").strip()
    if classe_pmt:
        queryset = queryset.filter(classe_pmt__icontains=classe_pmt)

    codigo = request.GET.get("codigo", "").strip()
    if codigo:
        queryset = queryset.filter(codigo__icontains=codigo)

    projeto = request.GET.get("projeto", "").strip()
    if projeto:
        queryset = queryset.filter(projetos__pk=projeto)

    queryset = queryset.order_by("-criado_em").distinct()

    page, paginator, page_num, page_size = _paginar_queryset(request, queryset)
    data = [_serializar_valvula_resumo(v) for v in page.object_list]
    return JsonResponse({
        "valvulas": data,
        "total": paginator.count,
        "page": page_num,
        "page_size": page_size,
        "total_pages": paginator.num_pages,
        "has_next": page.has_next(),
        "has_previous": page.has_previous(),
    })


@require_GET
def valvula_detalhe_api(request, pk):
    """Retorna detalhes completos de uma válvula em JSON."""
    valvula = get_object_or_404(
        Valvula.objects.select_related("criado_por").prefetch_related("projetos"), pk=pk
    )
    materiais = ValvulaMaterial.objects.filter(valvula=valvula).select_related("material")
    vedacoes = Vedacao.objects.filter(valvula=valvula)
    componentes = ComponentesInternos.objects.filter(valvula=valvula)

    tipo_material_label = dict(ValvulaMaterial.TIPO_MATERIAL)

    data = {
        "id": valvula.id_valvula,
        "codigo": valvula.codigo,
        "tipo_valvula": valvula.tipo_valvula,
        "tipo_label": TIPO_LABEL.get(valvula.tipo_valvula, valvula.tipo_valvula),
        "funcao": valvula.funcao or "BLOQUEIO",
        "projetos": [{"id": p.id_projeto, "nome": p.nome} for p in valvula.projetos.all()],
        "fabricante": valvula.fabricante or "",
        "pintura": valvula.pintura or "",
        "cor": valvula.cor or "",
        "norma_pintura": valvula.norma_pintura or "",
        "condicao_pintura": valvula.condicao_pintura or "",
        "pintura_atuador": valvula.pintura_atuador or "",
        "cor_atuador": valvula.cor_atuador or "",
        "norma_pintura_atuador": valvula.norma_pintura_atuador or "",
        "condicao_pintura_atuador": valvula.condicao_pintura_atuador or "",
        "norma": valvula.norma,
        "iogp": valvula.iogp or "",
        "qsl": valvula.qsl or "",
        "nbr": valvula.nbr,
        "diametro": valvula.diametro,
        "classe": valvula.classe,
        "classe_pmt": valvula.classe_pmt or "",
        "tipo_extremidade": valvula.tipo_extremidade or "",
        "tipo_ranhura": valvula.tipo_ranhura or "",
        "tipo_montagem": valvula.tipo_montagem or "",
        "tipo_passagem": valvula.tipo_passagem or "",
        "tipo_acionamento": valvula.tipo_acionamento or "",
        "marca_atuador": valvula.marca_atuador or "",
        "flange_acoplamento": valvula.flange_acoplamento or "",
        "construcao_corpo": valvula.construcao_corpo or "",

        "dib": valvula.dib or "",
        "valvula_alivio": valvula.valvula_alivio,
        "dispositivo_antiestatico": valvula.dispositivo_antiestatico,
        "uso_geral": valvula.uso_geral or "",
        "baixa_emissao_fugitiva": valvula.baixa_emissao_fugitiva,
        "certificacao_sil": valvula.certificacao_sil or "",
        "nace": valvula.nace or "",
        "revestimento": valvula.revestimento or "",
        "tipo_castelo": valvula.tipo_castelo or "",
        "juncao_corpo_castelo": valvula.juncao_corpo_castelo or "",
        "tipo_retencao": valvula.tipo_retencao or "",
        "configuracao_corpo_retencao": valvula.configuracao_corpo_retencao or "",
        "orientacao_instalacao": valvula.orientacao_instalacao or "",
        "categoria_594": valvula.categoria_594 or "",
        "categoria_borboleta": valvula.categoria_borboleta or "",
        "face_a_face": valvula.face_a_face or "",
        "configuracao_disco": valvula.configuracao_disco or "",
        "posicionador": valvula.posicionador or "",
        "ip": valvula.ip or "",
        "ip_posicionador": valvula.ip_posicionador or "",
        "ip_solenoide": valvula.ip_solenoide or "",
        "ip_chave_fim_curso": valvula.ip_chave_fim_curso or "",
        "ip_sensor_posicao": valvula.ip_sensor_posicao or "",
        **{campo: (getattr(valvula, campo) or "") for campo in CAMPOS_CE},
        **{campo: (getattr(valvula, campo) or "") for campo in CAMPOS_ELET},
        "filtro": valvula.filtro or "",
        "indicador_posicao": valvula.indicador_posicao,
        "tubing": valvula.tubing or "",
        "chave_fim_curso": valvula.chave_fim_curso or "",
        "valvula_solenoide": valvula.valvula_solenoide or "",
        "valvula_lock_up": valvula.valvula_lock_up or "",
        "sensor_posicao": valvula.sensor_posicao or "",
        "valvula_escape_rapido": valvula.valvula_escape_rapido or "",
        "caracteristicas": valvula.caracteristicas or "",
        "dreno": valvula.dreno,
        "vent": valvula.vent,
        "alivio_externo": valvula.alivio_externo,
        "hot_disconnect": valvula.hot_disconnect,
        "contra_peso": valvula.contra_peso,
        "placa_identificacao": valvula.placa_identificacao or "",
        "flange": valvula.flange or "",
        "anexo_nbr": valvula.anexo_nbr or "",
        "posicao_falha": valvula.posicao_falha or "",
        "tensao": valvula.tensao or "",
        "fase": valvula.fase or "",
        "frequencia": valvula.frequencia or "",
        "observacao": valvula.observacao or "",
        "materiais": [
            {
                "tipo_material": m.tipo_material,
                "tipo": tipo_material_label.get(m.tipo_material, m.tipo_material),
                "material": m.material.nome if m.material else "",
            }
            for m in materiais
            
        ],
        "vedacoes": [{"vedacao_corpo_tampa": v.vedacao_corpo_tampa, "vedacao_junta": getattr(v, "vedacao_junta", "")} for v in vedacoes],
        "componentes": [{"inserto_rede": c.inserto_rede} for c in componentes],
        "anexos": _listar_anexos_seguro(valvula),
        "pode_excluir_anexo": request.user.is_authenticated and request.user.nivel_permissao == "ESPECIAL",
        "criado_em": valvula.criado_em.astimezone(zone).strftime("%d/%m/%Y %H:%M") if valvula.criado_em else "",
        "criado_por": valvula.criado_por.nome if valvula.criado_por_id else "",
        "atualizado_em": valvula.atualizado_em.astimezone(zone).strftime("%d/%m/%Y %H:%M") if valvula.atualizado_em else "",
    }
    return JsonResponse(data)


# ── Anexos de Válvula ────────────────────────────────────────────────────────
ANEXO_EXTENSOES = {"pdf", "png", "jpg", "jpeg"}
ANEXO_TAMANHO_MAX = 10 * 1024 * 1024  # 10 MB


def _listar_anexos_seguro(valvula):
    """Lista anexos sem derrubar a página caso a tabela não exista (migration
    pendente) ou outro erro de banco. Anexos são opcionais."""
    try:
        return [_serialize_anexo(a) for a in valvula.anexos.all()]
    except Exception as exc:
        logger.warning("Falha ao listar anexos da válvula %s: %s", valvula.pk, exc)
        return []


def _serialize_anexo(anexo):
    return {
        "id": anexo.id,
        "nome_original": anexo.nome_original,
        "content_type": anexo.content_type,
        "tamanho": anexo.tamanho,
        "url_download": f"/api/anexos/{anexo.id}/download/",
        "enviado_em": anexo.enviado_em.astimezone(zone).strftime("%d/%m/%Y %H:%M") if anexo.enviado_em else "",
    }


@require_POST
def anexo_upload(request, pk):
    """Recebe 1+ arquivos (multipart) e os anexa à válvula. Qualquer usuário autenticado."""
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "Sem permissão"}, status=403)
    valvula = get_object_or_404(Valvula, pk=pk)

    arquivos = request.FILES.getlist("arquivos") or request.FILES.getlist("arquivo")
    if not arquivos:
        return JsonResponse({"success": False, "error": "Nenhum arquivo enviado"}, status=400)

    existing_count = AnexoValvula.objects.filter(valvula=valvula).count()
    if existing_count + len(arquivos) > 5:
        return JsonResponse({"success": False, "error": f"Limite de 5 anexos por válvula. Já existem {existing_count}."}, status=400)

    criados = []
    for f in arquivos:
        ext = f.name.rsplit(".", 1)[-1].lower() if "." in f.name else ""
        if ext not in ANEXO_EXTENSOES:
            return JsonResponse({"success": False, "error": f"Tipo não permitido: {f.name}. Aceitos: PDF, PNG, JPG, JPEG"}, status=400)
        if f.size > ANEXO_TAMANHO_MAX:
            return JsonResponse({"success": False, "error": f"Arquivo {f.name} excede 10 MB"}, status=400)

        storage_key = f"valvulas/{valvula.pk}/{uuid.uuid4().hex}.{ext}"
        content_type = f.content_type or "application/octet-stream"
        try:
            storage.upload(storage_key, f.read(), content_type)
        except storage.StorageError as exc:
            return JsonResponse({"success": False, "error": str(exc)}, status=502)

        anexo = AnexoValvula.objects.create(
            valvula=valvula,
            storage_key=storage_key,
            nome_original=f.name,
            content_type=content_type,
            tamanho=f.size,
        )
        criados.append(_serialize_anexo(anexo))

    return JsonResponse({"success": True, "anexos": criados})


@require_GET
def anexo_download(request, anexo_id):
    """Devolve o conteúdo do anexo (stream). Funciona com Supabase ou disco local."""
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "Sem permissão"}, status=403)
    anexo = get_object_or_404(AnexoValvula, pk=anexo_id)
    try:
        conteudo = storage.download(anexo.storage_key)
    except storage.StorageError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=502)
    resp = HttpResponse(conteudo, content_type=anexo.content_type or "application/octet-stream")
    resp["Content-Disposition"] = f'inline; filename="{anexo.nome_original}"'
    return resp


@require_POST
@especial_required
def anexo_excluir(request, anexo_id):
    """Exclui anexo do storage e do banco. Apenas ESPECIAL."""
    anexo = get_object_or_404(AnexoValvula, pk=anexo_id)
    try:
        storage.delete(anexo.storage_key)
    except storage.StorageError as exc:
        logger.warning("Falha ao excluir anexo %s do storage: %s", anexo.storage_key, exc)
    anexo.delete()
    return JsonResponse({"success": True})


@require_GET
def projeto_lista_api(request):
    """Lista projetos para autocomplete/filtro. ?q filtra por nome; ?status filtra por status."""
    q = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "").strip()
    qs = Projeto.objects.all()
    if q:
        qs = qs.filter(nome__icontains=q)
    if status_filter:
        qs = qs.filter(status=status_filter)
    qs = qs.order_by("-criado_em")[:50]
    data = [
        {
            "id": p.id_projeto,
            "nome": p.nome,
            "status": p.status,
            "criado_em": p.criado_em.astimezone(zone).strftime("%d/%m/%Y %H:%M") if p.criado_em else "",
            "total_valvulas": p.valvulas.count(),
        }
        for p in qs
    ]
    return JsonResponse({"projetos": data})


@require_POST
@especial_required
def projeto_criar(request):
    """Cria um projeto (ou retorna o existente de mesmo nome)."""
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Dados inválidos"}, status=400)
    nome = (body.get("nome") or "").strip()
    if not nome:
        return JsonResponse({"success": False, "error": "Nome do projeto é obrigatório"}, status=400)
    projeto, _ = Projeto.objects.get_or_create(nome=nome)
    return JsonResponse({"success": True, "projeto": {"id": projeto.id_projeto, "nome": projeto.nome, "status": projeto.status}})


@require_POST
@especial_required
def projeto_alterar_status(request, pk):
    """Altera status do projeto (ATIVO/CONCLUIDO)."""
    projeto = get_object_or_404(Projeto, pk=pk)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Dados inválidos"}, status=400)
    novo_status = body.get("status", "").strip()
    if novo_status not in ("ATIVO", "CONCLUIDO"):
        return JsonResponse({"success": False, "error": "Status inválido"}, status=400)
    projeto.status = novo_status
    projeto.save(update_fields=["status"])
    return JsonResponse({"success": True, "projeto": {"id": projeto.id_projeto, "nome": projeto.nome, "status": projeto.status}})


@require_POST
@especial_required
def projeto_excluir(request, pk):
    """Exclui projeto. As válvulas continuam existindo; apenas a associação (M2M) é removida."""
    projeto = get_object_or_404(Projeto, pk=pk)
    nome = projeto.nome
    total_valvulas = projeto.valvulas.count()
    projeto.delete()
    return JsonResponse({"success": True, "nome": nome, "valvulas_desassociadas": total_valvulas})


@require_POST
@especial_required
def valvula_atribuir_projeto(request):
    """Atribui uma ou várias válvulas (existentes) a um ou mais projetos via M2M.
    Body: {valvula_ids: [...], projeto_ids: [...], projeto_nomes: [...]}.
    Aceita também legado: projeto_id (int) ou projeto_nome (str) no singular.
    Restrito a ESPECIAL: gerência de projeto (criar/associar/excluir) é toda ESPECIAL.
    Sem esse gate, um COMUM criaria projetos via get_or_create, contornando projeto_criar."""
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Dados inválidos"}, status=400)
    ids = body.get("valvula_ids") or []
    if not ids:
        return JsonResponse({"success": False, "error": "Nenhuma válvula selecionada"}, status=400)

    projeto_ids = list(body.get("projeto_ids") or [])
    projeto_nomes = list(body.get("projeto_nomes") or [])
    # legado: campos singulares
    if not projeto_ids and not projeto_nomes:
        pid = body.get("projeto_id")
        pnome = (body.get("projeto_nome") or "").strip()
        if pid:
            projeto_ids = [pid]
        elif pnome:
            projeto_nomes = [pnome]

    if not projeto_ids and not projeto_nomes:
        return JsonResponse({"success": False, "error": "Informe ao menos um projeto"}, status=400)

    projetos = []
    for pid in projeto_ids:
        projetos.append(get_object_or_404(Projeto, pk=pid))
    for nome in projeto_nomes:
        nome = nome.strip()
        if nome:
            p, _ = Projeto.objects.get_or_create(nome=nome)
            projetos.append(p)

    if not projetos:
        return JsonResponse({"success": False, "error": "Informe ao menos um projeto"}, status=400)

    valvulas = Valvula.objects.filter(pk__in=ids)
    for projeto in projetos:
        projeto.valvulas.add(*valvulas)
    return JsonResponse({
        "success": True,
        "total": valvulas.count(),
        "projetos": [{"id": p.id_projeto, "nome": p.nome} for p in projetos],
        "projeto": {"id": projetos[0].id_projeto, "nome": projetos[0].nome},  # legado
    })


@require_POST
@especial_required
def valvula_desatribuir_projeto(request):
    """Remove a associação de uma ou várias válvulas a um projeto (M2M). Não exclui válvulas.
    Body: {valvula_ids: [...], projeto_id}. Restrito a ESPECIAL (espelha atribuir/projeto_*)."""
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Dados inválidos"}, status=400)
    ids = body.get("valvula_ids") or []
    projeto_id = body.get("projeto_id")
    if not ids or not projeto_id:
        return JsonResponse({"success": False, "error": "Válvulas e projeto são obrigatórios"}, status=400)
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    valvulas = Valvula.objects.filter(pk__in=ids)
    projeto.valvulas.remove(*valvulas)
    return JsonResponse({
        "success": True,
        "total": valvulas.count(),
        "projeto": {"id": projeto.id_projeto, "nome": projeto.nome},
    })


def projetos_page(request):
    """Página de gestão de projetos — só ESPECIAL."""
    if not request.user.is_authenticated or request.user.nivel_permissao != "ESPECIAL":
        return redirect("core:index")
    return render(request, "core/projetos.html")


def _opcoes_folha(model, fallback):
    """Opções (valor, valor) vindas do admin. Se tabela vazia, usa a lista
    hardcoded (fallback) — nada quebra se ninguém cadastrou nada ainda."""
    valores = list(model.objects.filter(ativo=True).values_list("valor", flat=True))
    if not valores:
        valores = [v[0] for v in fallback]
    return [(v, v) for v in valores]


@require_GET
def opcoes_por_tipo(request):
    """Retorna opções de campos dinâmicos por tipo de válvula."""
    tipo = request.GET.get("tipo", "")
    if not tipo:
        return JsonResponse({"error": "Parâmetro 'tipo' é obrigatório"}, status=400)

    normas = Valvula.NORMAS_POR_TIPO.get(tipo, [])
    diametros = Valvula.DIAMETROS_POR_TIPO.get(tipo, [])
    classes = Valvula.CLASSES_POR_TIPO.get(tipo, [])
    extremidades = Valvula.TIPO_EXTREMIDADE_POR_TIPO.get(tipo, [])
    acionamentos = Valvula.TIPO_ACIONAMENTO_POR_TIPO.get(tipo, [])
    passagens = Valvula.TIPO_PASSAGEM_POR_TIPO.get(tipo, [])
    castelos = Valvula.TIPO_CASTELO_POR_TIPO.get(tipo, [])
    juncoes = Valvula.JUNCAO_CORPO_CASTELO_POR_TIPO.get(tipo, [])
    campos_visiveis = Valvula.CAMPOS_POR_TIPO.get(tipo, [])
    tipos_material = Valvula.TIPOS_MATERIAL_POR_TIPO.get(tipo, [])
    vedacoes = Valvula.VEDACAO_POR_TIPO.get(tipo, [])
    uso_geral = Valvula.USO_GERAL_POR_TIPO.get(tipo, [])

    ranhuras = Valvula.TIPO_RANHURA
    if tipo == "BORBOLETA":
        ranhuras = Valvula.TIPO_RANHURA

    data = {
        "normas": list(normas) if normas and isinstance(normas[0], (list, tuple)) else [(n, n) for n in normas],
        "diametros": list(diametros) if diametros and isinstance(diametros[0], (list, tuple)) else [(d, d) for d in diametros],
        "classes": list(classes) if classes and isinstance(classes[0], (list, tuple)) else [(c, c) for c in classes],
        "extremidades": list(extremidades) if extremidades and isinstance(extremidades[0], (list, tuple)) else [(e, e) for e in extremidades],
        "ranhuras": list(ranhuras),
        "acionamentos": list(acionamentos) if acionamentos and isinstance(acionamentos[0], (list, tuple)) else [(a, a) for a in acionamentos],
        "passagens": list(passagens) if passagens and isinstance(passagens[0], (list, tuple)) else [(p, p) for p in passagens],
        "castelos": list(castelos) if castelos and isinstance(castelos[0], (list, tuple)) else [(c, c) for c in castelos],
        "juncoes": list(juncoes) if juncoes and isinstance(juncoes[0], (list, tuple)) else [(j, j) for j in juncoes],
        "campos_visiveis": campos_visiveis,
        "iogp": list(Valvula.IOGP),
        "fabricante": list(Valvula.FABRICANTE),
        "pintura": list(Valvula.PINTURA),
        "condicao_pintura_por_norma": {k: list(v) for k, v in Valvula.CONDICAO_POR_NORMA_PINTURA.items()},
        "tipos_material": tipos_material,
        "vedacoes": list(vedacoes) if vedacoes and isinstance(vedacoes[0], (list, tuple)) else [(v, v) for v in vedacoes],
        "uso_geral": list(uso_geral) if uso_geral and isinstance(uso_geral[0], (list, tuple)) else [(u, u) for u in uso_geral],
        "tipo_montagem": list(Valvula.TIPO_MONTAGEM_ESFERA) if tipo == "ESFERA" else [],
        "construcao_corpo": list(Valvula.CONSTRUCAO_CORPO_ESFERA) if tipo == "ESFERA" else [],
        "dib": list(Valvula.DIB),
        "certificacao_sil": list(Valvula.CERTIFICACAO_SIL),
        "nace": list(Valvula.NACE),
        "revestimento": list(Valvula.REVESTIMENTO),
        "categoria_borboleta": list(Valvula.CATEGORIA_BORBOLETA) if tipo == "BORBOLETA" else [],
        "face_a_face": list(Valvula.FACE_A_FACE) if tipo == "BORBOLETA" else [],
        "configuracao_disco": list(Valvula.CONFIGURACAO_DISCO) if tipo == "BORBOLETA" else [],
        "tipo_retencao": list(Valvula.TIPO_RETENCAO) if tipo == "RETENCAO" else [],
        "configuracao_corpo_retencao": list(Valvula.CONFIGURACAO_CORPO_RETENCAO) if tipo == "RETENCAO" else [],
        "orientacao_instalacao": list(Valvula.ORIENTACAO_INSTALACAO_RETENCAO) if tipo == "RETENCAO" else [],
        "categoria_594": list(Valvula.CATEGORIA_594) if tipo == "RETENCAO" else [],
        "posicionador": list(Valvula.POSICIONADOR),
        "filtro": list(Valvula.FILTRO) if tipo == "GLOBO_CONTROLE" else [],
        "tubing": list(Valvula.TUBING) if tipo == "GLOBO_CONTROLE" else [],
        "chave_fim_curso": list(Valvula.CHAVE_FIM_CURSO),
        "valvula_solenoide": list(Valvula.VALVULA_SOLENOIDE),
        "valvula_lock_up": list(Valvula.VALVULA_LOCK_UP),
        "sensor_posicao": list(Valvula.SENSOR_POSICAO),
        "valvula_escape_rapido": list(Valvula.VALVULA_ESCAPE_RAPIDO),
        "inserto_rede": list(Valvula.INSERTO_SEDE) if tipo in ("ESFERA", "RETENCAO") else [],
        "caracteristicas": list(Valvula.CARACTERISTICAS),
        "placa_identificacao": _opcoes_folha(OpcaoPlacaIdentificacao, Valvula.PLACA_IDENTIFICACAO),
        "flange": _opcoes_folha(OpcaoFlange, Valvula.FLANGE),
        "marca_atuador": list(Valvula.MARCA_ATUADOR),
        "flange_acoplamento": list(Valvula.FLANGE_ACOPLAMENTO_ISO5211),
        "posicao_falha": list(Valvula.POSICAO_FALHA),
    }
    return JsonResponse(data)


@require_GET
def materiais_por_tipo(request):
    """Retorna opções de materiais por tipo de válvula e tipo de material."""
    tipo_valvula = request.GET.get("tipo_valvula", "")
    tipo_material = request.GET.get("tipo_material", "")
    if not tipo_valvula or not tipo_material:
        return JsonResponse({"error": "Parâmetros obrigatórios: tipo_valvula, tipo_material"}, status=400)

    opcoes = Valvula.MATERIAIS_POR_TIPO.get(tipo_valvula, {}).get(tipo_material, [])
    if opcoes and isinstance(opcoes[0], (list, tuple)):
        data = {"materiais": list(opcoes)}
    else:
        data = {"materiais": [(m, m) for m in opcoes]}
    return JsonResponse(data)


# ── CRUD Válvulas ────────────────────────────────────────────────────────────

CAMPOS_BOOL_VALVULA = ["nbr", "valvula_alivio", "dispositivo_antiestatico", "baixa_emissao_fugitiva", "indicador_posicao", "dreno", "vent", "alivio_externo", "hot_disconnect", "contra_peso"]

# Subcategorias de instrumentação: (sufixo do campo, rótulo, campo "principal").
SUBCAT_INSTRUMENTACAO = [
    ("posicionador", "Posicionador", "posicionador"),
    ("solenoide", "Solenoid", "valvula_solenoide"),
    ("chave_fim_curso", "Chave Fim de Curso", "chave_fim_curso"),
    ("sensor_posicao", "Sensor de Posição", "sensor_posicao"),
]
# Partes das Características Elétricas (prefixo do campo): ex_<sub>, protecao_<sub>, ...
CE_PARTES = ["ex", "protecao", "grupo", "temp", "epl"]
# Todos os 20 campos de Características Elétricas
CAMPOS_CE = [f"{p}_{suf}" for suf, _, _ in SUBCAT_INSTRUMENTACAO for p in CE_PARTES]
# Elétrica por subcategoria: tensão / corrente / potência (12 campos)
ELET_PARTES = [("tensao", "Tensão", "Voltage"), ("corrente", "Corrente", "Current"), ("potencia", "Potência", "Power")]
CAMPOS_ELET = [f"{p}_{suf}" for suf, _, _ in SUBCAT_INSTRUMENTACAO for p, _, _ in ELET_PARTES]

def _campos_ce_elet_suf(suf):
    return [f"{p}_{suf}" for p in CE_PARTES] + [f"{p}_{suf}" for p, _, _ in ELET_PARTES]

# Campos da seção Instrumentação exclusivos do Posicionador (+ gerais) — exibidos/
# guardados só quando Função = Controle (posicionador só faz sentido controlando).
CAMPOS_INSTRUMENTACAO = ["posicionador", "ip_posicionador", "filtro", "tubing", "indicador_posicao",
                         "valvula_lock_up", "valvula_escape_rapido"] + _campos_ce_elet_suf("posicionador")

# Solenoide/Chave Fim de Curso/Sensor de Posição: opcionais em qualquer Função
# (também fazem sentido numa válvula de Bloqueio, sem precisar de posicionador).
CAMPOS_INSTRUMENTACAO_OPCIONAL = (
    ["valvula_solenoide", "ip_solenoide"] + _campos_ce_elet_suf("solenoide")
    + ["chave_fim_curso", "ip_chave_fim_curso"] + _campos_ce_elet_suf("chave_fim_curso")
    + ["sensor_posicao", "ip_sensor_posicao"] + _campos_ce_elet_suf("sensor_posicao")
)

# Campos texto comparados no dup-check
CAMPOS_TEXTO_DUP = [
    "fabricante", "pintura", "cor", "norma_pintura", "condicao_pintura", "norma", "iogp", "diametro", "classe", "tipo_extremidade", "tipo_ranhura",
    "tipo_passagem", "tipo_acionamento", "marca_atuador", "flange_acoplamento", "tipo_montagem", "construcao_corpo",
    "pintura_atuador", "cor_atuador", "norma_pintura_atuador", "condicao_pintura_atuador",
    "tipo_castelo", "juncao_corpo_castelo", "dib", "uso_geral",
    "tipo_retencao", "configuracao_corpo_retencao", "orientacao_instalacao", "categoria_594",
    "certificacao_sil", "nace", "revestimento", "categoria_borboleta", "face_a_face",
    "configuracao_disco", "posicionador", "ip", "ip_posicionador", "ip_solenoide", "ip_chave_fim_curso", "ip_sensor_posicao", "filtro", "tubing",
    "chave_fim_curso", "valvula_solenoide", "valvula_lock_up", "sensor_posicao", "valvula_escape_rapido",
    "caracteristicas", "placa_identificacao", "flange", "anexo_nbr",
    "posicao_falha", "tensao", "fase", "frequencia",
] + CAMPOS_CE + CAMPOS_ELET


def _aplicar_regras_automaticas(tipo_valvula, data):
    """Força valores derivados de outros campos. Aplicado na criação e na edição
    (mesmo com bypass de validação — consistência do dado)."""
    # Regra: Retenção → junção corpo/castelo sempre Aparafusado
    if tipo_valvula == "RETENCAO":
        data["juncao_corpo_castelo"] = "APARAFUSADO"

    # Regra: Retenção tipo Pistão → orientação de instalação só Horizontal (única
    # opção válida pra esse tipo construtivo — força, sem input manual, mesmo padrão
    # da junção corpo/castelo acima).
    if tipo_valvula == "RETENCAO" and data.get("tipo_retencao") == "PISTAO":
        data["orientacao_instalacao"] = "HORIZONTAL"

    # Regra: Gaveta + norma API 600 ou ISO 10434 → junção corpo/castelo sempre Aparafusado
    # (única opção dentro do escopo de ambas as normas — "bolted bonnet", Seção 1)
    if tipo_valvula == "GAVETA" and data.get("norma") in ("API 600", "ISO 10434"):
        data["juncao_corpo_castelo"] = "APARAFUSADO"

    # Regra: Globo + norma BS 1873 ou API 623 → junção corpo/castelo sempre Aparafusado
    # (BS 1873 9.3: "the body to bonnet connection shall be flanged"; API 623 Seção 1:
    # "bolted bonnet" é característica definidora do escopo — nenhuma das duas descreve
    # bonnet soldado/roscado em qualquer cláusula, só flange aparafusado)
    if tipo_valvula == "GLOBO" and data.get("norma") in ("BS 1873", "API 623"):
        data["juncao_corpo_castelo"] = "APARAFUSADO"

    # Regra: Função = Bloqueio → sem instrumentação (limpa todos os campos) e
    # característica obrigatoriamente On - Off. Só Função = Controle exibe/guarda
    # instrumentação e permite outras características. Independe do posicionador.
    if data.get("funcao") != "CONTROLE":
        for _campo in CAMPOS_INSTRUMENTACAO:
            data[_campo] = False if _campo == "indicador_posicao" else ""
        data["caracteristicas"] = "On - Off"
    else:
        # Controle: Solenoide/Chave Fim de Curso/Sensor de Posição (Sim/Não) são
        # obrigatórios → força "SIM".
        for _campo in ("valvula_solenoide", "chave_fim_curso", "sensor_posicao"):
            if _campo in Valvula.CAMPOS_POR_TIPO.get(tipo_valvula, []):
                data[_campo] = "SIM"

    # Regra: acionamento manual (Alavanca/Volante/Volante c/ Caixa de Redução) →
    # sem posição de falha (trava N/A) e sem dados elétricos
    if data.get("tipo_acionamento") in ACIONAMENTOS_MANUAIS:
        data["posicao_falha"] = "N/A"
        data["marca_atuador"] = "PADRÃO FABRICANTE"
        data["flange_acoplamento"] = ""
        data["tensao"] = ""
        data["fase"] = ""
        data["frequencia"] = ""
        data["hot_disconnect"] = False

    # Regra: acabamento da face do flange (tipo_ranhura) só existe em conexão
    # Flange/Wafer/Lug — qualquer outra (Butt-Welding, Socket-Welding, Rosca,
    # Niple, Gray Loc Hub) força N/A. Só se aplica a tipos com campo "Conexão"
    # (Borboleta não tem tipo_extremidade).
    if "tipo_extremidade" in Valvula.CAMPOS_POR_TIPO.get(tipo_valvula, []):
        _ext = (data.get("tipo_extremidade") or "").upper()
        if _ext and not (_ext.startswith("FLANGE") or _ext in ("WAFER", "LUG")):
            data["tipo_ranhura"] = "N/A"

    # Regra: norma de conexão (campo "Flange") derivada do tipo de extremidade.
    # Butt-Welding → ASME B16.25; Socket-Welding → ASME B16.11; Rosca NPT →
    # ASME B16.20; Niple → ASME B36.10 (corpo carbono) ou B36.19 (corpo inox
    # austenítico) — mesmos conjuntos de corpo usados nas regras de parafuso/
    # porca por corpo NBR (CORPOS_NBR_B7 / CORPOS_NBR_INOX). Flange/Wafer/Lug/
    # Gray Loc Hub não têm norma derivada — campo continua livre.
    _FLANGE_CORPOS_CARBONO = ("ASTM A105", "ASTM A105N", "ASTM A181", "ASTM A216 WCB", "ASTM A216 GR WCB")
    _FLANGE_CORPOS_INOX = ("ASTM A182 F304", "ASTM A351 CF8", "ASTM A182 F316", "ASTM A351 CF8M",
                            "ASTM A182 F317", "ASTM A351 CG8M", "ASTM A182 F347", "ASTM A351 CF8C")
    _ext_flange = (data.get("tipo_extremidade") or "").upper()
    if _ext_flange.startswith("BUTT-WELDING"):
        data["flange"] = "ASME B16.25"
    elif _ext_flange == "SOCKET-WELDING":
        data["flange"] = "ASME B16.11"
    elif _ext_flange == "ROSCA NPT":
        data["flange"] = "ASME B16.20"
    elif _ext_flange.startswith("NIPLE"):
        _materiais_flange = data.get("materiais", [])
        _corpo_flange = next((m.get("material") for m in _materiais_flange if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if _corpo_bate_base(_corpo_flange, _FLANGE_CORPOS_CARBONO):
            data["flange"] = "ASME B36.10"
        elif _corpo_bate_base(_corpo_flange, _FLANGE_CORPOS_INOX):
            data["flange"] = "ASME B36.19"

    # Regra: pintura (que agora já é a norma) define cor e condição de pintura.
    # Padrão Fabricante força "PADRÃO DO FABRICANTE"; Sem pintura força "N/A".
    # Uma norma (N-442, N-1735, ...) deixa cor/condição livres (validadas depois).
    _pintura_val = data.get("pintura") or ""
    data["norma_pintura"] = ""  # campo aposentado
    if _pintura_val == "PADRÃO FABRICANTE":
        data["cor"] = "PADRÃO DO FABRICANTE"
        data["condicao_pintura"] = "PADRÃO FABRICANTE"
    elif _pintura_val == "SEM PINTURA":
        data["cor"] = "N/A"
        data["condicao_pintura"] = "N/A"

    # Regra: pintura do atuador (espelha a regra acima)
    _pintura_atuador_val = data.get("pintura_atuador") or ""
    data["norma_pintura_atuador"] = ""  # campo aposentado
    if _pintura_atuador_val == "PADRÃO FABRICANTE":
        data["cor_atuador"] = "PADRÃO DO FABRICANTE"
        data["condicao_pintura_atuador"] = "PADRÃO FABRICANTE"
    elif _pintura_atuador_val == "SEM PINTURA":
        data["cor_atuador"] = "N/A"
        data["condicao_pintura_atuador"] = "N/A"

    # Regra: Esfera + Trunnion + NBR + DIB-1 → válvula de alívio obrigatória (força, sem input manual)
    if (tipo_valvula == "ESFERA" and data.get("nbr")
            and data.get("tipo_montagem") == "TRUNNION" and data.get("dib") == "DIB-1"):
        data["valvula_alivio"] = True

    # Regra: Esfera + diâmetro ≤2" → construção da sede trava em DBB (independe de NBR:
    # diâmetro pequeno não comporta DIB-1/DIB-2).
    if tipo_valvula == "ESFERA":
        _diam_dib = data.get("diametro") or ""
        if _diam_dib and _parse_diametro(_diam_dib) <= 2:
            data["dib"] = "DBB"


def _corpo_bate_base(material, bases):
    """True se `material` é exatamente uma das `bases` ou uma variante dela
    (ex.: "ASTM A105N revestimento interno de INCONEL" bate a base "ASTM A105N").
    Exige fronteira de espaço após a base para não confundir grau diferente
    com prefixo igual (ex.: "ASTM A182 F5" não deve bater "ASTM A182 F55").
    """
    if not material:
        return False
    return any(material == base or material.startswith(base + " ") for base in bases)


def _validar_regras_valvula(tipo_valvula, data):
    """Regras de negócio (NBR 15827 etc.) compartilhadas entre criação e edição.
    Assume _aplicar_regras_automaticas já executado.
    Retorna JsonResponse de erro (400) se inválido, ou None se ok."""
    # Regra: Função = Controle → característica não pode ser On - Off (esta é exclusiva de Bloqueio)
    if data.get("funcao") == "CONTROLE" and (data.get("caracteristicas") or "") == "On - Off":
        return JsonResponse({"success": False, "errors": {"caracteristicas": "Para Função = Controle, a característica não pode ser On - Off"}}, status=400)

    # Regra: Retenção — tipo construtivo (Pistão/Esfera/Disco) restringe configuração do
    # corpo e orientação de instalação:
    #   Pistão -> configuração Angular/Reto; orientação só Horizontal
    #   Esfera -> configuração Angular/Reto; orientação Horizontal/Vertical
    #   Disco  -> configuração só Reto; orientação Horizontal/Vertical
    _RETENCAO_CONFIG_POR_TIPO = {
        "PISTAO": {"ANGULAR", "RETO"},
        "ESFERA": {"ANGULAR", "RETO"},
        "DISCO": {"RETO"},
    }
    _RETENCAO_ORIENTACAO_POR_TIPO = {
        "PISTAO": {"HORIZONTAL"},
        "ESFERA": {"HORIZONTAL", "VERTICAL"},
        "DISCO": {"HORIZONTAL", "VERTICAL"},
    }
    if tipo_valvula == "RETENCAO":
        _tipo_ret = data.get("tipo_retencao") or ""
        _config_ret = data.get("configuracao_corpo_retencao") or ""
        _orient_ret = data.get("orientacao_instalacao") or ""
        if _tipo_ret in _RETENCAO_CONFIG_POR_TIPO and _config_ret and _config_ret not in _RETENCAO_CONFIG_POR_TIPO[_tipo_ret]:
            _opts_config = "/".join(v.capitalize() for v in _RETENCAO_CONFIG_POR_TIPO[_tipo_ret])
            return JsonResponse({"success": False, "errors": {"configuracao_corpo_retencao": f"Para Retenção tipo {_tipo_ret.capitalize()}, a configuração do corpo deve ser {_opts_config}"}}, status=400)
        if _tipo_ret in _RETENCAO_ORIENTACAO_POR_TIPO and _orient_ret and _orient_ret not in _RETENCAO_ORIENTACAO_POR_TIPO[_tipo_ret]:
            _opts_orient = "/".join(v.capitalize() for v in _RETENCAO_ORIENTACAO_POR_TIPO[_tipo_ret])
            return JsonResponse({"success": False, "errors": {"orientacao_instalacao": f"Para Retenção tipo {_tipo_ret.capitalize()}, a orientação de instalação deve ser {_opts_orient}"}}, status=400)

    # Regra: pintura com norma (não Padrão Fabricante/Sem Pintura) exige cor específica.
    _PINTURA_MODOS = ("", "PADRÃO FABRICANTE", "SEM PINTURA")
    if (data.get("pintura") or "") not in _PINTURA_MODOS and (data.get("cor") or "") in ("", "PADRÃO DO FABRICANTE", "N/A"):
        return JsonResponse({"success": False, "errors": {"cor": "Para pintura com norma, selecione uma cor específica (não pode ser Padrão do Fabricante nem N/A)"}}, status=400)

    # Regra: pintura do atuador com norma exige cor específica (espelha a regra acima)
    if (data.get("pintura_atuador") or "") not in _PINTURA_MODOS and (data.get("cor_atuador") or "") in ("", "PADRÃO DO FABRICANTE", "N/A"):
        return JsonResponse({"success": False, "errors": {"cor_atuador": "Para pintura do atuador com norma, selecione uma cor específica (não pode ser Padrão do Fabricante nem N/A)"}}, status=400)

    # Regra: norma API 6D → QSL obrigatório
    if data.get("norma") == "API 6D" and not (data.get("qsl") or "").strip():
        return JsonResponse({"success": False, "errors": {"qsl": "Para norma API 6D, o QSL é obrigatório"}}, status=400)

    # Regra: norma API 6D → classe não pode ser PN, 400, 800 ou 4500
    # (400 só existe no modelo pra Retenção/BS 1868 — a API 6D não tem essa designação)
    _classe_6d_val = data.get("classe") or ""
    if data.get("norma") == "API 6D" and (_classe_6d_val.startswith("PN") or _classe_6d_val in ("400", "800", "4500")):
        return JsonResponse({"success": False, "errors": {"classe": "Para norma API 6D, as classes 400, 800 e 4500 não são permitidas e a classe não pode ser PN. Use 150, 300, 600, 900, 1500 ou 2500"}}, status=400)

    # Regra: norma API 6D → diâmetro máximo por classe
    _API6D_MAX_DN = {"900": 48, "1500": 36, "2500": 20}
    if data.get("norma") == "API 6D":
        _classe_6d = data.get("classe", "")
        _max_dn_6d = _API6D_MAX_DN.get(_classe_6d)
        if _max_dn_6d:
            _diam_6d = data.get("diametro")
            _dn_6d = _parse_diametro(_diam_6d) if _diam_6d else None
            if _dn_6d is not None and _dn_6d > _max_dn_6d:
                return JsonResponse({"success": False, "errors": {"diametro": f'Para norma API 6D com classe {_classe_6d}, o diâmetro máximo é {_max_dn_6d}"'}}, status=400)

    # Regra: Gaveta + NBR 15827 → gaxeta obrigatoriamente com inibidor
    GAXETA_NBR_OBRIGATORIA = "GRAFITE FLEXÍVEL + FIO DE INCONEL C/ INIBIDOR (molibdato de bário e/ou fios de zinco)"
    if tipo_valvula == "GAVETA" and data.get("nbr"):
        materiais = data.get("materiais", [])
        gaxeta_selecionada = next((m.get("material") for m in materiais if m.get("tipo_material") == "GAXETA"), None)
        if gaxeta_selecionada and gaxeta_selecionada != GAXETA_NBR_OBRIGATORIA:
            return JsonResponse({"success": False, "errors": {"materiais": "Para Gaveta com NBR 15827, a gaxeta deve ser obrigatoriamente Grafite Flex. + Fio Inconel c/ Inibidor (molibdato bário/zinco)"}}, status=400)

    # Regra: Gaveta/Retenção + NBR 15827 → vedação corpo/tampa restrita por classe
    _erro_ved = _validar_vedacao_nbr_classe(tipo_valvula, data)
    if _erro_ved:
        return _erro_ved

    # Regra: Retenção + NBR + Flange/Butt-Welding →
    # diâmetro 2"-36"; classe 150-2500; norma BS 1868/ASME B16.34/ISO 14313/API 6D.
    # Sem 800: a Tabela 2 põe 800 só na coluna "Encaixe para solda" (forjado). 800 é
    # designação de forjado (API 602/ISO 15761); não existe flange classe 800.
    #
    # A célula "Flange ou solda de topo" da Tabela 2 tem DUAS alternativas de padrão
    # construtivo, e só uma delas exige passagem plena:
    #   - "BS 1868, ASME B16.34 e Anexo B"              → passagem livre
    #   - "ISO 14313 (API 6D) e Anexo B (Passagem Plena)" → passagem plena obrigatória
    _RET_FB_NORMAS = {"BS 1868", "ASME B16.34", "ISO 14313", "API 6D"}
    _RET_FB_CLASSES = {"150", "300", "600", "900", "1500", "2500"}
    _RET_FB_NORMAS_PLENA = {"ISO 14313", "API 6D"}
    _ext_ret = data.get("tipo_extremidade", "")
    if (tipo_valvula == "RETENCAO" and data.get("nbr")
            and (_ext_ret.startswith("FLANGE") or _ext_ret == "BUTT-WELDING")):
        _diam = data.get("diametro")
        _classe = data.get("classe")
        _norma = data.get("norma")
        _dn = _parse_diametro(_diam) if _diam else None
        if _dn is not None and not (2 <= _dn <= 36):
            return JsonResponse({"success": False, "errors": {"diametro": "Para Retenção com NBR 15827 e extremidade Flange/Butt-Welding, o diâmetro deve ser de 2\" a 36\""}}, status=400)
        if _classe and _classe not in _RET_FB_CLASSES:
            return JsonResponse({"success": False, "errors": {"classe": "Para Retenção com NBR 15827 e extremidade Flange/Butt-Welding, a classe deve ser de 150 a 2500"}}, status=400)
        if _norma and _norma not in _RET_FB_NORMAS:
            return JsonResponse({"success": False, "errors": {"norma": "Para Retenção com NBR 15827 e extremidade Flange/Butt-Welding, a norma deve ser BS 1868, ASME B16.34, ISO 14313 ou API 6D"}}, status=400)
        if _norma in _RET_FB_NORMAS_PLENA and (data.get("tipo_passagem") or "") == "REDUZIDA":
            return JsonResponse({"success": False, "errors": {"tipo_passagem": "Para Retenção com NBR 15827, Flange/Butt-Welding e norma ISO 14313 ou API 6D, a passagem deve ser Plena"}}, status=400)

    # Regra: Retenção + NBR + Wafer →
    # diâmetro 2"-42"; classe 150-2500; norma API 594. Sem 800 (ver regra acima).
    _RET_WAFER_CLASSES = {"150", "300", "600", "900", "1500", "2500"}
    if (tipo_valvula == "RETENCAO" and data.get("nbr")
            and data.get("tipo_extremidade") == "Wafer"):
        _diam = data.get("diametro")
        _classe = data.get("classe")
        _norma = data.get("norma")
        _dn = _parse_diametro(_diam) if _diam else None
        if _dn is not None and not (2 <= _dn <= 42):
            return JsonResponse({"success": False, "errors": {"diametro": "Para Retenção com NBR 15827 e extremidade Wafer, o diâmetro deve ser de 2\" a 42\""}}, status=400)
        if _classe and _classe not in _RET_WAFER_CLASSES:
            return JsonResponse({"success": False, "errors": {"classe": "Para Retenção com NBR 15827 e extremidade Wafer, a classe deve ser de 150 a 2500"}}, status=400)
        if _norma and _norma != "API 594":
            return JsonResponse({"success": False, "errors": {"norma": "Para Retenção com NBR 15827 e extremidade Wafer, a norma deve ser API 594"}}, status=400)

    # Regra: Retenção + NBR + Socket-Welding + corpo (A105/A182/A350/B564/B865) →
    # diâmetro 1/2" a 1 1/2"; classe 800/1500/2500;
    # classe 800/1500 → norma ISO 15761 ou API 602; classe 2500 → norma ASME B16.34.
    _RET_NBR_SW_CORPOS = ("ASTM A105", "ASTM A182", "ASTM A350", "ASTM B564", "ASTM B865")
    _DIAMETROS_RET_SW = {'1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"'}
    _CLASSES_RET_SW = {"800", "1500", "2500"}
    _NORMAS_RET_SW_800_1500 = {"ISO 15761", "API 602"}
    if (tipo_valvula == "RETENCAO" and data.get("nbr")
            and data.get("tipo_extremidade") == "SOCKET-WELDING"):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if mat_corpo and mat_corpo.startswith(_RET_NBR_SW_CORPOS):
            diametro = data.get("diametro")
            classe = data.get("classe")
            norma = data.get("norma")
            if diametro and diametro not in _DIAMETROS_RET_SW:
                return JsonResponse({"success": False, "errors": {"diametro": "Para Retenção com NBR 15827, Socket-Welding e corpo A105/A182/A350/B564/B865, o diâmetro deve ser de 1/2\" a 1 1/2\""}}, status=400)
            if classe and classe not in _CLASSES_RET_SW:
                return JsonResponse({"success": False, "errors": {"classe": "Para Retenção com NBR 15827, Socket-Welding e corpo A105/A182/A350/B564/B865, a classe deve ser 800, 1500 ou 2500"}}, status=400)
            if classe == "2500" and norma and norma != "ASME B16.34":
                return JsonResponse({"success": False, "errors": {"norma": "Para Retenção com NBR 15827, Socket-Welding, corpo A105/A182/A350/B564/B865 e classe 2500, a norma deve ser ASME B16.34"}}, status=400)
            if classe in {"800", "1500"} and norma and norma not in _NORMAS_RET_SW_800_1500:
                return JsonResponse({"success": False, "errors": {"norma": "Para Retenção com NBR 15827, Socket-Welding, corpo A105/A182/A350/B564/B865 e classe 800/1500, a norma deve ser ISO 15761 ou API 602"}}, status=400)

    # Regra: Gaveta + NBR 15827 + Socket-Welding + corpo (A105/A182/A350/B564/B865) →
    # diâmetro 1/2" a 1 1/2"; classe 800/1500/2500; norma conforme classe
    # (classe 800/1500 → ISO 15761/API 602/ASME B16.34; classe 2500 → só ASME B16.34).
    _GAVETA_NBR_SW_CORPOS = ("ASTM A105", "ASTM A182", "ASTM A350", "ASTM B564", "ASTM B865")
    _DIAMETROS_GAVETA_SW = {'1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"'}
    _CLASSES_GAVETA_SW = {"800", "1500", "2500"}
    # A célula da Tabela 1 cobre "800 e 1500" com 3 normas juntas, sem dizer qual vale para
    # qual classe. A B16.34 (2.1.1) não tem Class 800 → sobra p/ 800 só ISO 15761/API 602.
    _NORMAS_GAVETA_SW_800 = {"ISO 15761", "API 602"}
    _NORMAS_GAVETA_SW_1500 = {"ISO 15761", "API 602", "ASME B16.34"}
    if (tipo_valvula == "GAVETA" and data.get("nbr")
            and data.get("tipo_extremidade") == "SOCKET-WELDING"):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if mat_corpo and mat_corpo.startswith(_GAVETA_NBR_SW_CORPOS):
            diametro = data.get("diametro")
            classe = data.get("classe")
            norma = data.get("norma")
            if diametro and diametro not in _DIAMETROS_GAVETA_SW:
                return JsonResponse({"success": False, "errors": {"diametro": "Para Gaveta com NBR 15827, Socket-Welding e corpo A105/A182/A350/B564/B865, o diâmetro deve ser de 1/2\" a 1 1/2\""}}, status=400)
            if classe and classe not in _CLASSES_GAVETA_SW:
                return JsonResponse({"success": False, "errors": {"classe": "Para Gaveta com NBR 15827, Socket-Welding e corpo A105/A182/A350/B564/B865, a classe deve ser 800, 1500 ou 2500"}}, status=400)
            if classe == "2500" and norma and norma != "ASME B16.34":
                return JsonResponse({"success": False, "errors": {"norma": "Para Gaveta com NBR 15827, Socket-Welding, corpo A105/A182/A350/B564/B865 e classe 2500, a norma deve ser ASME B16.34"}}, status=400)
            if classe == "800" and norma and norma not in _NORMAS_GAVETA_SW_800:
                return JsonResponse({"success": False, "errors": {"norma": "Para Gaveta com NBR 15827, Socket-Welding, corpo A105/A182/A350/B564/B865 e classe 800, a norma deve ser ISO 15761 ou API 602 (a ASME B16.34 não tem classe 800)"}}, status=400)
            if classe == "1500" and norma and norma not in _NORMAS_GAVETA_SW_1500:
                return JsonResponse({"success": False, "errors": {"norma": "Para Gaveta com NBR 15827, Socket-Welding, corpo A105/A182/A350/B564/B865 e classe 1500, a norma deve ser ISO 15761, API 602 ou ASME B16.34"}}, status=400)

    # Regra: Globo + NBR 15827 + Socket-Welding + corpo (A105/A182/A350/B564/B865) →
    # diâmetro 1/2" a 1 1/2"; classe 800/1500; norma ISO 15761/API 602/ASME B16.34.
    _GLOBO_SW_DIAM = {'1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"'}
    _GLOBO_SW_CLASSES = {"800", "1500"}
    # Mesma leitura da Gaveta: a célula "800 e 1500" da Tabela 4 lista as normas juntas;
    # a B16.34 não tem Class 800, então ela só cabe no 1500.
    _GLOBO_SW_NORMAS_800 = {"ISO 15761", "API 602"}
    _GLOBO_SW_NORMAS_1500 = {"ISO 15761", "API 602", "ASME B16.34"}
    if (tipo_valvula == "GLOBO" and data.get("nbr")
            and data.get("tipo_extremidade") == "SOCKET-WELDING"):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if mat_corpo and mat_corpo.startswith(_GAVETA_NBR_SW_CORPOS):
            diametro = data.get("diametro")
            classe = data.get("classe")
            norma = data.get("norma")
            if diametro and diametro not in _GLOBO_SW_DIAM:
                return JsonResponse({"success": False, "errors": {"diametro": "Para Globo com NBR 15827, Socket-Welding e corpo A105/A182/A350/B564/B865, o diâmetro deve ser de 1/2\" a 1 1/2\""}}, status=400)
            if classe and classe not in _GLOBO_SW_CLASSES:
                return JsonResponse({"success": False, "errors": {"classe": "Para Globo com NBR 15827, Socket-Welding e corpo A105/A182/A350/B564/B865, a classe deve ser 800 ou 1500"}}, status=400)
            if classe == "800" and norma and norma not in _GLOBO_SW_NORMAS_800:
                return JsonResponse({"success": False, "errors": {"norma": "Para Globo com NBR 15827, Socket-Welding, corpo A105/A182/A350/B564/B865 e classe 800, a norma deve ser ISO 15761 ou API 602 (a ASME B16.34 não tem classe 800)"}}, status=400)
            if classe == "1500" and norma and norma not in _GLOBO_SW_NORMAS_1500:
                return JsonResponse({"success": False, "errors": {"norma": "Para Globo com NBR 15827, Socket-Welding, corpo A105/A182/A350/B564/B865 e classe 1500, a norma deve ser ISO 15761, API 602 ou ASME B16.34"}}, status=400)

    # Regra: Globo + NBR 15827 + Butt-Welding →
    # corpo A105/A182/A350/B564/B865 → classe 1500 ou 2500; demais corpos → só 1500.
    # classe 2500 → diâmetro 1" a 1 1/2", norma ASME B16.34.
    # classe 1500 → diâmetro 2" a 16", norma BS 1873/ASME B16.34.
    _GLOBO_BW_DIAM_2500 = {'1"', '1 1/4"', '1 1/2"'}
    if (tipo_valvula == "GLOBO" and data.get("nbr")
            and data.get("tipo_extremidade") == "BUTT-WELDING"):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        _corpo_ok = bool(mat_corpo and mat_corpo.startswith(_GAVETA_NBR_SW_CORPOS))
        diametro = data.get("diametro")
        classe = data.get("classe")
        norma = data.get("norma")
        _classes_ok = {"1500", "2500"} if _corpo_ok else {"1500"}
        if classe and classe not in _classes_ok:
            _msg = "1500 ou 2500" if _corpo_ok else "1500 (classe 2500 só p/ corpo A105/A182/A350/B564/B865)"
            return JsonResponse({"success": False, "errors": {"classe": "Para Globo com NBR 15827 e Butt-Welding, a classe deve ser " + _msg}}, status=400)
        if classe == "2500":
            if diametro and diametro not in _GLOBO_BW_DIAM_2500:
                return JsonResponse({"success": False, "errors": {"diametro": "Para Globo com NBR 15827, Butt-Welding e classe 2500, o diâmetro deve ser de 1\" a 1 1/2\""}}, status=400)
            if norma and norma != "ASME B16.34":
                return JsonResponse({"success": False, "errors": {"norma": "Para Globo com NBR 15827, Butt-Welding e classe 2500, a norma deve ser ASME B16.34"}}, status=400)
        elif classe == "1500":
            _dn = _parse_diametro(diametro) if diametro else None
            if _dn is not None and not (2 <= _dn <= 16):
                return JsonResponse({"success": False, "errors": {"diametro": "Para Globo com NBR 15827, Butt-Welding e classe 1500, o diâmetro deve ser de 2\" a 16\""}}, status=400)
            if norma and norma not in {"BS 1873", "ASME B16.34"}:
                return JsonResponse({"success": False, "errors": {"norma": "Para Globo com NBR 15827, Butt-Welding e classe 1500, a norma deve ser BS 1873 ou ASME B16.34"}}, status=400)

    # Regra: Globo + NBR 15827 + Flange (qualquer) →
    # diâmetro 2"-12"; classe 150-2500; norma BS 1873 ou ASME B16.34.
    # Sem 800: a Tabela 4 põe 800 só na coluna "Encaixe para solda" (forjado).
    _GLOBO_FLANGE_CLASSES = {"150", "300", "600", "900", "1500", "2500"}
    _GLOBO_FLANGE_NORMAS = {"BS 1873", "ASME B16.34"}
    _ext_globo = data.get("tipo_extremidade", "")
    if tipo_valvula == "GLOBO" and data.get("nbr") and _ext_globo.startswith("FLANGE"):
        _diam = data.get("diametro")
        _classe = data.get("classe")
        _norma = data.get("norma")
        _dn = _parse_diametro(_diam) if _diam else None
        if _dn is not None and not (2 <= _dn <= 12):
            return JsonResponse({"success": False, "errors": {"diametro": "Para Globo com NBR 15827 e extremidade Flange, o diâmetro deve ser de 2\" a 12\""}}, status=400)
        if _classe and _classe not in _GLOBO_FLANGE_CLASSES:
            return JsonResponse({"success": False, "errors": {"classe": "Para Globo com NBR 15827 e extremidade Flange, a classe deve ser de 150 a 2500"}}, status=400)
        if _norma and _norma not in _GLOBO_FLANGE_NORMAS:
            return JsonResponse({"success": False, "errors": {"norma": "Para Globo com NBR 15827 e extremidade Flange, a norma deve ser BS 1873 ou ASME B16.34"}}, status=400)

    # Regra: Borboleta + Categoria A → configuração do disco obrigatoriamente Concêntrica.
    if tipo_valvula == "BORBOLETA" and data.get("categoria_borboleta") == "CATEGORIA A":
        _disco_catA = data.get("configuracao_disco", "")
        if _disco_catA and _disco_catA != "CONCÊNTRICA":
            return JsonResponse({"success": False, "errors": {"configuracao_disco": "Para Borboleta Categoria A, a configuração do disco deve ser Concêntrica"}}, status=400)

    # Regra: Borboleta + diâmetro > 2" → classe não pode ser 800.
    if tipo_valvula == "BORBOLETA":
        _diam = data.get("diametro")
        _dn = _parse_diametro(_diam) if _diam else None
        if _dn is not None and _dn >= 2 and data.get("classe") == "800":
            return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta com diâmetro maior ou igual a 2\", a classe não pode ser 800"}}, status=400)

    # Regra: Borboleta + Categoria A + norma API 609 → disco Concêntrica; diâmetro 2"-48"; classe 125/150/PMT;
    # face a face Lug ou Wafer. PMT exige valor textual em classe_pmt. Só aplica quando a norma é API 609.
    if (tipo_valvula == "BORBOLETA" and data.get("categoria_borboleta") == "CATEGORIA A"
            and data.get("norma") == "API 609"):
        _diam = data.get("diametro")
        _classe = data.get("classe")
        _dn = _parse_diametro(_diam) if _diam else None
        if _dn is not None and not (2 <= _dn <= 48):
            return JsonResponse({"success": False, "errors": {"diametro": "Para Borboleta Categoria A, o diâmetro deve ser de 2\" a 48\""}}, status=400)
        if _classe and _classe not in {"125", "150", "PMT"}:
            return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta Categoria A, a classe deve ser 125, 150 ou PMT"}}, status=400)
        _disco_a = data.get("configuracao_disco", "")
        if _disco_a and _disco_a != "CONCÊNTRICA":
            return JsonResponse({"success": False, "errors": {"configuracao_disco": "Para Borboleta Categoria A, a configuração do disco deve ser Concêntrica"}}, status=400)
        # API 609 1.3 a): Categoria A é coberta só em lug e wafer — as tabelas de duplo
        # flange (3B/3C) são exclusivas de Categoria B.
        _faf_a = data.get("face_a_face", "")
        if _faf_a and _faf_a not in {"LUG", "WAFER"}:
            return JsonResponse({"success": False, "errors": {"face_a_face": "Para Borboleta Categoria A com norma API 609, o face a face deve ser Lug ou Wafer"}}, status=400)
        if _classe == "PMT" and not (data.get("classe_pmt") or "").strip():
            return JsonResponse({"success": False, "errors": {"classe_pmt": "Informe a PMT"}}, status=400)

    # Regra: Borboleta + Categoria B + norma API 609 → classe 150/300/600; diâmetro por (face a face, classe).
    #   LUG/WAFER: 3"-48". PADRÃO LONGO (150/300/600): 3"-36".
    #   PADRÃO CURTO (150/300): 3"-48"; (600): 3"-24".
    if (tipo_valvula == "BORBOLETA" and data.get("categoria_borboleta") == "CATEGORIA B"
            and data.get("norma") == "API 609"):
        _classe = data.get("classe")
        if _classe and _classe not in {"150", "300", "600"}:
            return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta Categoria B com norma API 609, a classe deve ser 150, 300 ou 600"}}, status=400)
        _faf = data.get("face_a_face", "")
        _diam = data.get("diametro")
        _dn = _parse_diametro(_diam) if _diam else None
        # 1.3 b): toda Categoria B vai de 3" a 48"; o face a face só aperta o teto.
        _max = 48
        if _faf == "FLANGEADA PADRÃO LONGO":
            _max = 36
        elif _faf == "FLANGEADA PADRÃO CURTO" and _classe == "600":
            _max = 24
        if _dn is not None and not (3 <= _dn <= _max):
            return JsonResponse({"success": False, "errors": {"diametro": f"Para Borboleta Categoria B (API 609, {_faf}, classe {_classe}), o diâmetro deve ser de 3\" a {_max}\""}}, status=400)

    # Regra: Borboleta + norma API 609 → classe dentro do escopo da norma (Classes 125-600,
    # + PMT/CWP da Categoria A). Roda depois das regras de Categoria A/B, que são mais
    # específicas; esta pega o caso de categoria em branco, que aquelas não cobrem.
    _CLASSES_API609 = {"125", "150", "300", "600", "PMT"}
    if tipo_valvula == "BORBOLETA" and data.get("norma") == "API 609":
        _classe = data.get("classe")
        if _classe and _classe not in _CLASSES_API609:
            return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta com norma API 609, a classe deve ser 125, 150, 300, 600 ou PMT"}}, status=400)

    # Regra: Borboleta + NBR + (Wafer/Lug) + disco Concêntrica + corpo ASTM A536 65-45-12
    # → diâmetro 2"-48"; classe PMT (abre campo texto); norma API 609.
    if (tipo_valvula == "BORBOLETA" and data.get("nbr")
            and data.get("face_a_face") in {"LUG", "WAFER"}
            and data.get("configuracao_disco") == "CONCÊNTRICA"):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if mat_corpo == "ASTM A536 65-45-12":
            _diam = data.get("diametro")
            _dn = _parse_diametro(_diam) if _diam else None
            if _dn is not None and not (2 <= _dn <= 48):
                return JsonResponse({"success": False, "errors": {"diametro": "Para Borboleta com NBR, Wafer/Lug, Concêntrica e corpo ASTM A536 65-45-12, o diâmetro deve ser de 2\" a 48\""}}, status=400)
            if data.get("classe") and data.get("classe") != "PMT":
                return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta com NBR, Wafer/Lug, Concêntrica e corpo ASTM A536 65-45-12, a classe deve ser PMT"}}, status=400)
            if data.get("norma") and data.get("norma") != "API 609":
                return JsonResponse({"success": False, "errors": {"norma": "Para Borboleta com NBR, Wafer/Lug, Concêntrica e corpo ASTM A536 65-45-12, a norma deve ser API 609"}}, status=400)
            if data.get("classe") == "PMT" and not (data.get("classe_pmt") or "").strip():
                return JsonResponse({"success": False, "errors": {"classe_pmt": "Informe a PMT"}}, status=400)

    # Regra: Borboleta + NBR + (Wafer/Lug/Flangeada) + corpo ASTM A216 (qualquer) ou ASTM A105 →
    # Bi-Excêntrica:  diâmetro 2"-24"; classe 150/300/600; norma API 609.
    # Tri-Excêntrica: diâmetro 2"-48"; classe 150-1500; norma API 609/ASME B16.34.
    _faf_a216 = data.get("face_a_face", "")
    _disco_a216 = data.get("configuracao_disco", "")
    if (tipo_valvula == "BORBOLETA" and data.get("nbr")
            and _disco_a216 in {"BI-EXCÊNTRICA", "TRI-EXCÊNTRICA"}
            and (_faf_a216 in {"LUG", "WAFER"} or _faf_a216.startswith("FLANGEAD"))):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if mat_corpo and mat_corpo.startswith(("ASTM A216", "ASTM A105")):
            _diam = data.get("diametro")
            _classe = data.get("classe")
            _norma = data.get("norma")
            _dn = _parse_diametro(_diam) if _diam else None
            if _disco_a216 == "BI-EXCÊNTRICA":
                if _dn is not None and not (2 <= _dn <= 24):
                    return JsonResponse({"success": False, "errors": {"diametro": "Para Borboleta com NBR 15827, Bi-Excêntrica (Wafer/Lug/Flangeada) e corpo ASTM A216 ou A105, o diâmetro deve ser de 2\" a 24\""}}, status=400)
                if _classe and _classe not in {"150", "300", "600"}:
                    return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta com NBR 15827, Bi-Excêntrica (Wafer/Lug/Flangeada) e corpo ASTM A216 ou A105, a classe deve ser de 150 a 600"}}, status=400)
                if _norma and _norma != "API 609":
                    return JsonResponse({"success": False, "errors": {"norma": "Para Borboleta com NBR 15827, Bi-Excêntrica (Wafer/Lug/Flangeada) e corpo ASTM A216 ou A105, a norma deve ser API 609"}}, status=400)
            else:  # TRI-EXCÊNTRICA
                if _dn is not None and not (2 <= _dn <= 48):
                    return JsonResponse({"success": False, "errors": {"diametro": "Para Borboleta com NBR 15827, Tri-Excêntrica (Wafer/Lug/Flangeada) e corpo ASTM A216 ou A105, o diâmetro deve ser de 2\" a 48\""}}, status=400)
                if _norma and _norma not in {"API 609", "ASME B16.34"}:
                    return JsonResponse({"success": False, "errors": {"norma": "Para Borboleta com NBR 15827, Tri-Excêntrica (Wafer/Lug/Flangeada) e corpo ASTM A216 ou A105, a norma deve ser API 609 ou ASME B16.34"}}, status=400)
                # API 609 só cobre classes 150/300/600; ASME B16.34 cobre até 1500.
                if _norma == "API 609":
                    if _classe and _classe not in {"150", "300", "600"}:
                        return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta com NBR 15827, Tri-Excêntrica (Wafer/Lug/Flangeada), corpo ASTM A216 ou A105 e norma API 609, a classe deve ser de 150 a 600"}}, status=400)
                elif _classe and _classe not in {"150", "300", "600", "900", "1500"}:
                    return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta com NBR 15827, Tri-Excêntrica (Wafer/Lug/Flangeada) e corpo ASTM A216 ou A105, a classe deve ser de 150 a 1500"}}, status=400)

    # Regra: Borboleta + NBR + (Wafer/Lug/Flangeada) + corpo NÃO A105/A182/A350/B564/B865 →
    # Bi-Excêntrica:  diâmetro 2"-24"; classe 150/300/600; norma API 609.
    # Tri-Excêntrica: diâmetro 2"-48"; classe 150-1500; norma API 609/ASME B16.34.
    _faf_borb = data.get("face_a_face", "")
    _disco_borb = data.get("configuracao_disco", "")
    if (tipo_valvula == "BORBOLETA" and data.get("nbr")
            and _disco_borb in {"BI-EXCÊNTRICA", "TRI-EXCÊNTRICA"}
            and (_faf_borb in {"LUG", "WAFER"} or _faf_borb.startswith("FLANGEAD"))):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if mat_corpo and not mat_corpo.startswith(_GAVETA_NBR_SW_CORPOS):
            _diam = data.get("diametro")
            _classe = data.get("classe")
            _norma = data.get("norma")
            _dn = _parse_diametro(_diam) if _diam else None
            if _disco_borb == "BI-EXCÊNTRICA":
                if _dn is not None and not (2 <= _dn <= 24):
                    return JsonResponse({"success": False, "errors": {"diametro": "Para Borboleta com NBR 15827, Bi-Excêntrica (Wafer/Lug/Flangeada) e corpo fora de A105/A182/A350/B564/B865, o diâmetro deve ser de 2\" a 24\""}}, status=400)
                if _classe and _classe not in {"150", "300", "600"}:
                    return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta com NBR 15827, Bi-Excêntrica (Wafer/Lug/Flangeada) e corpo fora de A105/A182/A350/B564/B865, a classe deve ser de 150 a 600"}}, status=400)
                if _norma and _norma != "API 609":
                    return JsonResponse({"success": False, "errors": {"norma": "Para Borboleta com NBR 15827, Bi-Excêntrica (Wafer/Lug/Flangeada) e corpo fora de A105/A182/A350/B564/B865, a norma deve ser API 609"}}, status=400)
            else:  # TRI-EXCÊNTRICA
                if _dn is not None and not (2 <= _dn <= 48):
                    return JsonResponse({"success": False, "errors": {"diametro": "Para Borboleta com NBR 15827, Tri-Excêntrica (Wafer/Lug/Flangeada) e corpo fora de A105/A182/A350/B564/B865, o diâmetro deve ser de 2\" a 48\""}}, status=400)
                if _classe and _classe not in {"150", "300", "600", "900", "1500"}:
                    return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta com NBR 15827, Tri-Excêntrica (Wafer/Lug/Flangeada) e corpo fora de A105/A182/A350/B564/B865, a classe deve ser de 150 a 1500"}}, status=400)
                if _norma and _norma not in {"API 609", "ASME B16.34"}:
                    return JsonResponse({"success": False, "errors": {"norma": "Para Borboleta com NBR 15827, Tri-Excêntrica (Wafer/Lug/Flangeada) e corpo fora de A105/A182/A350/B564/B865, a norma deve ser API 609 ou ASME B16.34"}}, status=400)

    # Regra: Borboleta + norma MSS SP-67 → diâmetro em lista fechada (1½" a 72", com
    # buracos — 5", 64", 66", 72" não existem em DIAMETROS/DIAMETROS_POR_TIPO, únicos
    # pra essa norma); classe 125/150/PMT (Seções 3.1-3.3 só cobrem flange até Classe
    # 150 — ASME B16.1 Cl 25/125, B16.5 Cl 150, B16.24/B16.42 Cl 150, B16.47 Cl 150
    # Série A; 4.3 diz que rating de flange fora dessa lista está fora do escopo. PMT
    # admitido pelo teste de prova alternativo de 4.1.1.3/4.1.3.3/4.1.4.3, mesmo caso
    # do CWP de API 609 Categoria A). Roda depois das regras de NBR Bi/Tri-Excêntrica
    # (que restringem norma por corpo e devem falar primeiro).
    _CLASSES_MSS_SP67 = {"125", "150", "PMT"}
    _DIAMETROS_MSS_SP67 = {
        '1 1/2"', '2"', '2 1/2"', '3"', '4"', '5"', '6"', '8"', '10"', '12"',
        '14"', '16"', '18"', '20"', '24"', '30"', '36"', '42"', '48"', '54"',
        '60"', '64"', '66"', '72"',
    }
    if tipo_valvula == "BORBOLETA" and data.get("norma") == "MSS SP67":
        _diam = data.get("diametro")
        if _diam and _diam not in _DIAMETROS_MSS_SP67:
            return JsonResponse({"success": False, "errors": {"diametro": "Para Borboleta com norma MSS SP-67, o diâmetro deve ser 1 1/2\", 2\", 2 1/2\", 3\", 4\", 5\", 6\", 8\", 10\", 12\", 14\", 16\", 18\", 20\", 24\", 30\", 36\", 42\", 48\", 54\", 60\", 64\", 66\" ou 72\""}}, status=400)
        _classe = data.get("classe")
        if _classe and _classe not in _CLASSES_MSS_SP67:
            return JsonResponse({"success": False, "errors": {"classe": "Para Borboleta com norma MSS SP-67, a classe deve ser 125, 150 ou PMT"}}, status=400)
        if _classe == "PMT" and not (data.get("classe_pmt") or "").strip():
            return JsonResponse({"success": False, "errors": {"classe_pmt": "Informe a PMT"}}, status=400)

    # Regra: Gaveta + NBR + Flange/Butt-Welding → diâmetro/classe/norma por faixa.
    #   2"-24"  classe 150/300/600/900 → norma ISO 10434/API 600/ASME B16.34
    #   2"-16"  classe 1500              → norma ISO 10434/API 600/ASME B16.34
    #   2"-12"  classe 2500              → norma ISO 10434/API 600/ASME B16.34
    #   26"-42" classe 150-600           → norma ASME B16.34
    _GAVETA_FLANGE_NORMAS_A = {"ISO 10434", "API 600", "ASME B16.34"}
    _ext = data.get("tipo_extremidade", "")
    if (tipo_valvula == "GAVETA" and data.get("nbr")
            and (_ext.startswith("FLANGE") or _ext == "BUTT-WELDING")):
        _diam = data.get("diametro")
        _classe = data.get("classe")
        _norma = data.get("norma")
        _dn = _parse_diametro(_diam) if _diam else None
        if _dn is not None and not (2 <= _dn <= 42):
            return JsonResponse({"success": False, "errors": {"diametro": "Para Gaveta com NBR 15827 e extremidade Flange/Butt-Welding, o diâmetro deve ser de 2\" a 42\""}}, status=400)
        if _dn is not None and _classe and not _classe.startswith("PN"):
            _cl = int(_classe) if _classe.isdigit() else None
            if 2 <= _dn <= 24:
                # Sem 800: a Tabela 1 dá "150 a 900" para flange/solda de topo, e 800 não
                # entra nessa sequência (é designação de forjado, coluna "Encaixe p/ solda").
                if _cl in {150, 300, 600, 900}:
                    if _norma and _norma not in _GAVETA_FLANGE_NORMAS_A:
                        return JsonResponse({"success": False, "errors": {"norma": "Para Gaveta com NBR 15827, Flange/Butt-Welding, diâmetro 2\"-24\" e classe 150-900, a norma deve ser ISO 10434, API 600 ou ASME B16.34"}}, status=400)
                elif _cl == 1500 and _dn <= 16:
                    if _norma and _norma not in _GAVETA_FLANGE_NORMAS_A:
                        return JsonResponse({"success": False, "errors": {"norma": "Para Gaveta com NBR 15827, Flange/Butt-Welding, diâmetro 2\"-16\" e classe 1500, a norma deve ser ISO 10434, API 600 ou ASME B16.34"}}, status=400)
                elif _cl == 2500 and _dn <= 12:
                    if _norma and _norma not in _GAVETA_FLANGE_NORMAS_A:
                        return JsonResponse({"success": False, "errors": {"norma": "Para Gaveta com NBR 15827, Flange/Butt-Welding, diâmetro 2\"-12\" e classe 2500, a norma deve ser ISO 10434, API 600 ou ASME B16.34"}}, status=400)
                else:
                    return JsonResponse({"success": False, "errors": {"classe": "Para Gaveta com NBR 15827, Flange/Butt-Welding e diâmetro 2\"-24\", a classe deve ser 150, 300, 600 ou 900 (1500 só até 16\", 2500 só até 12\"). A classe 800 só existe em extremidade de encaixe para solda"}}, status=400)
            elif 26 <= _dn <= 42:
                if _cl in {150, 300, 600}:
                    if _norma and _norma != "ASME B16.34":
                        return JsonResponse({"success": False, "errors": {"norma": "Para Gaveta com NBR 15827, Flange/Butt-Welding e diâmetro 26\"-42\", a norma deve ser ASME B16.34"}}, status=400)
                else:
                    return JsonResponse({"success": False, "errors": {"classe": "Para Gaveta com NBR 15827, Flange/Butt-Welding e diâmetro 26\"-42\", a classe deve ser de 150 a 600"}}, status=400)

    # Regra: Gaveta/Retenção + NBR 15827 → junta não pode ser N/A
    if tipo_valvula in ("GAVETA", "RETENCAO") and data.get("nbr"):
        materiais = data.get("materiais", [])
        mat_junta = next((m.get("material") for m in materiais if m.get("tipo_material") == "MATERIAL_JUNTA"), None)
        if mat_junta == "N/A":
            return JsonResponse({"success": False, "errors": {"materiais": "Para Gaveta/Retenção com NBR 15827, a junta não pode ser N/A"}}, status=400)

    # Regra: NBR 15827 + corpo carbono (A105/A181/A216 WCB, inclui variantes
    # revestidas: A105N revestido, A216 WCB revestido) → parafusos B7 e porcas 2H
    CORPOS_NBR_B7 = ("ASTM A105", "ASTM A105N", "ASTM A181", "ASTM A216 WCB", "ASTM A216 GR WCB")
    if data.get("nbr"):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if _corpo_bate_base(mat_corpo, CORPOS_NBR_B7):
            mat_paraf = next((m.get("material") for m in materiais if m.get("tipo_material") == "PARAFUSOS"), None)
            mat_porca = next((m.get("material") for m in materiais if m.get("tipo_material") == "PORCAS"), None)
            if mat_paraf and mat_paraf != "ASTM A193 B7":
                return JsonResponse({"success": False, "errors": {"materiais": "Para NBR 15827 com corpo ASTM A105/A181/A216 WCB, os parafusos devem ser ASTM A193 Gr B7"}}, status=400)
            if mat_porca and mat_porca != "ASTM A194 2H":
                return JsonResponse({"success": False, "errors": {"materiais": "Para NBR 15827 com corpo ASTM A105/A181/A216 WCB, as porcas devem ser ASTM A194 Gr 2H"}}, status=400)

    # Regra: NBR 15827 + corpo liga/baixa temp (A350 LF2/LF3, A352 LCB, A217 C5) → parafusos/porcas restritos
    CORPOS_NBR_LIGA = ("ASTM A350 LF2", "ASTM A350 LF3", "ASTM A352 LCB", "ASTM A352 LC3")
    PARAFUSOS_NBR_LIGA = {"ASTM A320 Gr L7", "ASTM A193 Gr B8M", "ASTM A193 Gr B8M CL2"}
    PORCAS_NBR_LIGA = {"ASTM A194 Gr 8M", "ASTM A194 Gr 4L", "ASTM A194 Gr 7L"}
    if data.get("nbr"):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if _corpo_bate_base(mat_corpo, CORPOS_NBR_LIGA):
            mat_paraf = next((m.get("material") for m in materiais if m.get("tipo_material") == "PARAFUSOS"), None)
            mat_porca = next((m.get("material") for m in materiais if m.get("tipo_material") == "PORCAS"), None)
            if mat_paraf and mat_paraf not in PARAFUSOS_NBR_LIGA:
                return JsonResponse({"success": False, "errors": {"materiais": "Para NBR 15827 com corpo ASTM A350 LF2/LF3, A352 LCB ou A352 LC3, os parafusos devem ser ASTM A320 Gr L7, A193 Gr B8M ou A193 Gr B8M CL2"}}, status=400)
            if mat_porca and mat_porca not in PORCAS_NBR_LIGA:
                return JsonResponse({"success": False, "errors": {"materiais": "Para NBR 15827 com corpo ASTM A350 LF2/LF3, A352 LCB ou A352 LC3, as porcas devem ser ASTM A194 Gr 8M, Gr 4L ou Gr 7L"}}, status=400)

    # Regra: NBR 15827 + corpo cromo-molib (A182 F11 CL2/F5, A217 WC6/C5) → parafusos B16 e porcas Gr 7
    CORPOS_NBR_B16 = ("ASTM A182 F11 CL2", "ASTM A182 F5", "ASTM A217 WC6", "ASTM A217 C5")
    if data.get("nbr"):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if _corpo_bate_base(mat_corpo, CORPOS_NBR_B16):
            mat_paraf = next((m.get("material") for m in materiais if m.get("tipo_material") == "PARAFUSOS"), None)
            mat_porca = next((m.get("material") for m in materiais if m.get("tipo_material") == "PORCAS"), None)
            if mat_paraf and mat_paraf != "ASTM A193 Gr B16":
                return JsonResponse({"success": False, "errors": {"materiais": "Para NBR 15827 com corpo ASTM A182 F11 CL2/F5 ou A217 WC6/C5, os parafusos devem ser ASTM A193 Gr B16"}}, status=400)
            if mat_porca and mat_porca != "ASTM A194 Gr 7":
                return JsonResponse({"success": False, "errors": {"materiais": "Para NBR 15827 com corpo ASTM A182 F11 CL2/F5 ou A217 WC6/C5, as porcas devem ser ASTM A194 Gr 7"}}, status=400)

    # Regra: NBR 15827 + corpo inox austenítico (F304/F316/F317/F347, CF8/CF8M/CG8M/CF8C) → B8M e porca 8M, revestimento N/A
    CORPOS_NBR_INOX = ("ASTM A182 F304", "ASTM A351 CF8", "ASTM A182 F316", "ASTM A351 CF8M",
                       "ASTM A182 F317", "ASTM A351 CG8M", "ASTM A182 F347", "ASTM A351 CF8C")
    PARAFUSOS_NBR_INOX = {"ASTM A193 Gr B8M", "ASTM A193 Gr B8M CL2"}
    if data.get("nbr"):
        materiais = data.get("materiais", [])
        mat_corpo = next((m.get("material") for m in materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if _corpo_bate_base(mat_corpo, CORPOS_NBR_INOX):
            mat_paraf = next((m.get("material") for m in materiais if m.get("tipo_material") == "PARAFUSOS"), None)
            mat_porca = next((m.get("material") for m in materiais if m.get("tipo_material") == "PORCAS"), None)
            if mat_paraf and mat_paraf not in PARAFUSOS_NBR_INOX:
                return JsonResponse({"success": False, "errors": {"materiais": "Para NBR 15827 com corpo inox austenítico, os parafusos devem ser ASTM A193 Gr B8M ou B8M CL2"}}, status=400)
            if mat_porca and mat_porca != "ASTM A194 Gr 8M":
                return JsonResponse({"success": False, "errors": {"materiais": "Para NBR 15827 com corpo inox austenítico, as porcas devem ser ASTM A194 Gr 8M"}}, status=400)
            if data.get("revestimento") and data.get("revestimento") != "N/A":
                return JsonResponse({"success": False, "errors": {"revestimento": "Para NBR 15827 com corpo inox austenítico, o revestimento deve ser N/A"}}, status=400)

    # Regra: sem NBR 15827 → parafuso/porca ainda seguem par compatível (fora da NBR o
    # corpo não dita o par, mas parafuso e porca continuam tendo que combinar entre si).
    # Com NBR ativa, as regras de corpo acima já restringem o par.
    PARAFUSO_PORCA_COMPAT = {
        "ASTM A193 B7": {"ASTM A194 2H"},
        "ASTM A193 B8": {"ASTM A194 Gr 8", "ASTM A194 8A"},
        "ASTM A193 Gr B8M": {"ASTM A194 Gr 8M", "ASTM A194 Grade 8MA"},
        "ASTM A193 Gr B8M CL2": {"ASTM A194 Gr 8M", "ASTM A194 Grade 8MA"},
        "ASTM A193 Gr B16": {"ASTM A194 Gr 4", "ASTM A194 Gr 7"},
        "ASTM A320 Gr L7": {"ASTM A194 Gr 7", "ASTM A194 Gr 7L", "ASTM A194 Gr 4", "ASTM A194 Gr 4L"},
        "ASTM A193 Grade B7M": {"ASTM A194 Grade 2HM", "ASTM A194 Grade 7M"},
        "ASTM A193 Grade B8MA, Class 1A": {"ASTM A194 Grade 8MA", "ASTM A194 Gr 8M"},
        "ASTM A320 Grade L7M": {"ASTM A194 Grade 7M", "ASTM A194 Grade 2HM"},
        "ASTM A193 B8A": {"ASTM A194 8A", "ASTM A194 Gr 8"},
        "ASTM A193 B8T": {"ASTM A194 8T"},
        "ZERON 100 FG": {"ZERON 100 FG"},
        "UNS S32760": {"UNS S32760"},
        "UNS S32550": {"UNS S32550"},
        "PADRÃO FABRICANTE": {"PADRÃO FABRICANTE", "N/A"},
        "N/A": {"N/A"},
    }
    if not data.get("nbr"):
        materiais = data.get("materiais", [])
        mat_paraf = next((m.get("material") for m in materiais if m.get("tipo_material") == "PARAFUSOS"), None)
        mat_porca = next((m.get("material") for m in materiais if m.get("tipo_material") == "PORCAS"), None)
        _porcas_ok = PARAFUSO_PORCA_COMPAT.get(mat_paraf)
        if mat_paraf and mat_porca and _porcas_ok and mat_porca not in _porcas_ok:
            _opts = " / ".join(sorted(_porcas_ok))
            return JsonResponse({"success": False, "errors": {"materiais": f"Para parafuso {mat_paraf}, a porca deve ser {_opts}"}}, status=400)

    # Regra: Esfera + NBR 15827 → haste deve ter mesmo material do obturador
    if tipo_valvula == "ESFERA" and data.get("nbr"):
        materiais = data.get("materiais", [])
        mat_obturador = next((m.get("material") for m in materiais if m.get("tipo_material") == "OBTURADOR"), None)
        mat_haste = next((m.get("material") for m in materiais if m.get("tipo_material") == "HASTE"), None)
        if mat_obturador and mat_haste and mat_haste != mat_obturador:
            return JsonResponse({"success": False, "errors": {"materiais": "Para Esfera com NBR 15827, o material da haste deve ser igual ao do obturador"}}, status=400)

    # Regra: Esfera + NBR 15827 → dispositivo antiestático obrigatório
    if tipo_valvula == "ESFERA" and data.get("nbr") and not data.get("dispositivo_antiestatico"):
        return JsonResponse({"success": False, "errors": {"dispositivo_antiestatico": "Para Esfera com NBR 15827, o dispositivo antiestático é obrigatório"}}, status=400)

    # Regra PRIORITÁRIA: Esfera + NBR + Socket-Welding + classe 800 → norma ISO 17292
    # Tem prioridade sobre a regra de extremidade Niple (que força Niple p/ classe 800).
    _esfera_socket_800 = (
        tipo_valvula == "ESFERA" and data.get("nbr")
        and data.get("tipo_extremidade") == "SOCKET-WELDING"
        and data.get("classe") == "800"
    )
    if _esfera_socket_800 and data.get("norma") != "ISO 17292":
        return JsonResponse({"success": False, "errors": {"norma": "Para Esfera com NBR 15827, extremidade Socket-Welding e classe 800, a norma deve ser ISO 17292"}}, status=400)

    # Regra PRIORITÁRIA: Esfera + NBR + classe 1500/2500 + Socket-Welding → norma ASME B16.34
    _esfera_b1634_socket = (
        tipo_valvula == "ESFERA" and data.get("nbr")
        and data.get("classe") in {"1500", "2500"}
        and data.get("tipo_extremidade") == "SOCKET-WELDING"
    )
    if _esfera_b1634_socket and data.get("norma") != "ASME B16.34":
        return JsonResponse({"success": False, "errors": {"norma": "Para Esfera com NBR 15827, classes 1500/2500 e extremidade Socket-Welding, a norma deve ser ASME B16.34"}}, status=400)

    # Regra: Esfera + NBR + Flange/Butt-Welding →
    # norma/uso geral ISO 14313/API 6D/ISO 10497/API 607; classe 150/300/600/900/1500/2500.
    # Diâmetro por classe: 150-600 → 2"-36"; 900 → 2"-24"; 1500 → 2"-16"; 2500 → 2"-12".
    _ESF_FB_NORMAS = {"ISO 14313", "API 6D", "ISO 10497", "API 607"}
    _ESF_FB_USO = {"ISO 14313", "API 6D", "ISO 10497", "API 607"}
    _ESF_FB_CLASSES = {"150", "300", "600", "900", "1500", "2500"}
    _ext_esf = data.get("tipo_extremidade", "")
    if (tipo_valvula == "ESFERA" and data.get("nbr")
            and (_ext_esf.startswith("FLANGE") or _ext_esf == "BUTT-WELDING")):
        _diam = data.get("diametro")
        _classe = data.get("classe")
        _norma = data.get("norma")
        _uso = data.get("uso_geral")
        _dn = _parse_diametro(_diam) if _diam else None
        _max_dn = 24 if _classe == "900" else 16 if _classe == "1500" else 12 if _classe == "2500" else 36
        if _dn is not None and not (2 <= _dn <= _max_dn):
            return JsonResponse({"success": False, "errors": {"diametro": f"Para Esfera com NBR 15827, extremidade Flange/Butt-Welding e classe {_classe}, o diâmetro deve ser de 2\" a {_max_dn}\""}}, status=400)
        if _classe and _classe not in _ESF_FB_CLASSES:
            return JsonResponse({"success": False, "errors": {"classe": "Para Esfera com NBR 15827 e extremidade Flange/Butt-Welding, a classe deve ser 150, 300, 600, 900, 1500 ou 2500"}}, status=400)
        if _norma and _norma not in _ESF_FB_NORMAS:
            return JsonResponse({"success": False, "errors": {"norma": "Para Esfera com NBR 15827 e extremidade Flange/Butt-Welding, a norma deve ser ISO 14313, API 6D, ISO 10497 ou API 607"}}, status=400)
        if _uso and _uso not in _ESF_FB_USO:
            return JsonResponse({"success": False, "errors": {"uso_geral": "Para Esfera com NBR 15827 e extremidade Flange/Butt-Welding, o uso geral deve ser ISO 14313, API 6D, ISO 10497 ou API 607"}}, status=400)

    # Regra: Esfera + NBR + Rosca + corpo (A105/A182/A350/B564/B865) →
    # diâmetro 1/2"-1 1/2". A NORMA CONSTRUTIVA escolhida é o parâmetro que dita
    # classe e uso geral (não o contrário): a Tabela 3 amarra cada padrão construtivo
    # a uma célula, e a leitura por classe era ambígua no OCR (ver CLAUDE.md —
    # "Suspeita não confirmada: Esfera + Rosca + classe 800").
    #   BS ISO 7121 → classe 150; uso N/A
    #   ISO 17292   → classe 800; uso ISO 17292/ISO 10497/API 607
    _ESF_ROSCA_CORPOS = ("ASTM A105", "ASTM A182", "ASTM A350", "ASTM B564", "ASTM B865")
    _ESF_ROSCA_DIAM = {'1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"'}
    _ESF_ROSCA_USO_800 = {"ISO 17292", "ISO 10497", "API 607"}
    _ESF_ROSCA_NORMAS = {"BS ISO 7121", "ISO 17292"}
    _ext_rosca = data.get("tipo_extremidade", "")
    if tipo_valvula == "ESFERA" and data.get("nbr") and _ext_rosca.startswith("ROSCA"):
        _materiais = data.get("materiais", [])
        _corpo = next((m.get("material") for m in _materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if _corpo and _corpo.startswith(_ESF_ROSCA_CORPOS):
            _diam = data.get("diametro")
            _classe = data.get("classe")
            _norma = data.get("norma")
            _uso = data.get("uso_geral")
            if _diam and _diam not in _ESF_ROSCA_DIAM:
                return JsonResponse({"success": False, "errors": {"diametro": "Para Esfera com NBR 15827, Rosca e corpo A105/A182/A350/B564/B865, o diâmetro deve ser de 1/2\" a 1 1/2\""}}, status=400)
            if _classe and _classe not in {"150", "800"}:
                return JsonResponse({"success": False, "errors": {"classe": "Para Esfera com NBR 15827, Rosca e corpo A105/A182/A350/B564/B865, a classe deve ser 150 ou 800"}}, status=400)
            if _norma and _norma not in _ESF_ROSCA_NORMAS:
                return JsonResponse({"success": False, "errors": {"norma": "Para Esfera com NBR 15827, Rosca e corpo A105/A182/A350/B564/B865, a norma deve ser BS ISO 7121 ou ISO 17292"}}, status=400)
            if _norma == "BS ISO 7121":
                if _classe and _classe != "150":
                    return JsonResponse({"success": False, "errors": {"classe": "Para Esfera com NBR 15827, Rosca e norma BS ISO 7121, a classe deve ser 150"}}, status=400)
                if _uso and _uso != "N/A":
                    return JsonResponse({"success": False, "errors": {"uso_geral": "Para Esfera com NBR 15827, Rosca e norma BS ISO 7121, o uso geral deve ser N/A"}}, status=400)
            elif _norma == "ISO 17292":
                if _classe and _classe != "800":
                    return JsonResponse({"success": False, "errors": {"classe": "Para Esfera com NBR 15827, Rosca e norma ISO 17292, a classe deve ser 800"}}, status=400)
                if _uso and _uso not in _ESF_ROSCA_USO_800:
                    return JsonResponse({"success": False, "errors": {"uso_geral": "Para Esfera com NBR 15827, Rosca e norma ISO 17292, o uso geral deve ser ISO 17292, ISO 10497 ou API 607"}}, status=400)

    # Regra: Esfera + NBR + Socket-Welding + corpo (A105/A182/A350/B564/B865) →
    # diâmetro 1/2"-1 1/2"; classe 800/1500/2500;
    # 800 → norma ISO 17292 + uso ISO 17292/ISO 10497/API 607;
    # 1500/2500 → norma ASME B16.34 + uso ASME B16.34/ISO 10497/API 607.
    # (a norma já é validada pelas regras _esfera_socket_800/_esfera_b1634_socket acima.)
    _ESF_SW_CORPOS = ("ASTM A105", "ASTM A182", "ASTM A350", "ASTM B564", "ASTM B865")
    _ESF_SW_DIAM = {'1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"'}
    _ESF_SW_USO_800 = {"ISO 17292", "ISO 10497", "API 607"}
    _ESF_SW_USO_1500 = {"ASME B16.34", "ISO 10497", "API 607"}
    if (tipo_valvula == "ESFERA" and data.get("nbr")
            and data.get("tipo_extremidade") == "SOCKET-WELDING"):
        _materiais = data.get("materiais", [])
        _corpo = next((m.get("material") for m in _materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if _corpo and _corpo.startswith(_ESF_SW_CORPOS):
            _diam = data.get("diametro")
            _classe = data.get("classe")
            _uso = data.get("uso_geral")
            if _diam and _diam not in _ESF_SW_DIAM:
                return JsonResponse({"success": False, "errors": {"diametro": "Para Esfera com NBR 15827, Socket-Welding e corpo A105/A182/A350/B564/B865, o diâmetro deve ser de 1/2\" a 1 1/2\""}}, status=400)
            if _classe and _classe not in {"800", "1500", "2500"}:
                return JsonResponse({"success": False, "errors": {"classe": "Para Esfera com NBR 15827, Socket-Welding e corpo A105/A182/A350/B564/B865, a classe deve ser 800, 1500 ou 2500"}}, status=400)
            if _classe == "800" and _uso and _uso not in _ESF_SW_USO_800:
                return JsonResponse({"success": False, "errors": {"uso_geral": "Para Esfera com NBR 15827, Socket-Welding e classe 800, o uso geral deve ser ISO 17292, ISO 10497 ou API 607"}}, status=400)
            if _classe in {"1500", "2500"} and _uso and _uso not in _ESF_SW_USO_1500:
                return JsonResponse({"success": False, "errors": {"uso_geral": "Para Esfera com NBR 15827, Socket-Welding e classe 1500/2500, o uso geral deve ser ASME B16.34, ISO 10497 ou API 607"}}, status=400)

    # Regra: diâmetro > 3" → extremidade Niple não é oferecida (peça de pequeno porte;
    # acima de 3" usa-se conexão direta — Flange/Butt-Welding/Socket-Welding/Rosca).
    # Vale pra qualquer tipo de válvula que ofereça Niple (Esfera/Gaveta/Globo/Retenção).
    _diam_niple = data.get("diametro")
    if data.get("tipo_extremidade", "").startswith("NIPLE") and _diam_niple:
        if _parse_diametro(_diam_niple) > 3:
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": 'Para diâmetro acima de 3", a extremidade não pode ser Niple'}}, status=400)

    # Regras Esfera: extremidade Niple forçada por corpo + classe (Tabela C.2 da NBR 15827)
    _CLASSES_NIPLE_NBR = {'150', '300', '600', '800'}
    _CLASSES_NIPLE_SCH160 = {'900', '1500'}
    _MATERIAIS_SCH160 = {
        'ASTM A105', 'ASTM A182 F11 CL1', 'ASTM A182 F11 CL2',
        'ASTM A216 WCB', 'ASTM A216 WCC',
        'ASTM A217 WC5', 'ASTM A217 WC6', 'ASTM A217 C5', 'ASTM A217 C12',
        'ASTM A350 LF1', 'ASTM A350 LF2', 'ASTM A350 LF3',
        'ASTM A352 LCA', 'ASTM A352 LCB', 'ASTM A352 LCC',
    }
    _MATERIAIS_SCH80 = {
        'ASTM A182 F6a', 'ASTM A182 F304', 'ASTM A182 F304 L',
        'ASTM A182 F316', 'ASTM A182 F316 L', 'ASTM A182 F317',
        'ASTM A182 F321', 'ASTM A182 F347',
        'ASTM A182 F51', 'ASTM A182 F53', 'ASTM A182 F54', 'ASTM A182 F55',
        'ASTM A351 CF3', 'ASTM A351 CF8', 'ASTM A351 CF3M',
        'ASTM A351 CF8M', 'ASTM A351 CF8C', 'ASTM A351 CN7M',
        'ASTM A352 CA6NM',
        'ASTM A995 3A', 'ASTM A995 4A', 'ASTM A995 5A', 'ASTM A995 6A',
    }
    if (tipo_valvula == "ESFERA"
            and data.get("tipo_extremidade", "").startswith("NIPLE")):
        _materiais = data.get("materiais", [])
        _corpo = next((m.get("material") for m in _materiais if m.get("tipo_material") == "CORPO_TAMPA"), None)
        _classe = data.get("classe", "")
        _extremidade = data.get("tipo_extremidade", "")
        _corpo_nbr = _corpo in _MATERIAIS_SCH160 or _corpo in _MATERIAIS_SCH80
        # Regra NBR classe 2500: XXS para carbono, liga e inox/duplex.
        if data.get("nbr") and _classe == "2500":
            if _corpo_nbr and _extremidade != 'NIPLE 4" COMP. SCH XXS':
                return JsonResponse({"success": False, "errors": {"tipo_extremidade": 'Para Esfera com NBR 15827 e classe 2500, a extremidade deve ser Niple 4" Comp. SCH XXS'}}, status=400)
        # Regra NBR classe 900/1500: SCH 160 para carbono, liga e inox/duplex.
        # (socket nessas classes já é tratado antes pela regra de norma ASME B16.34).
        elif data.get("nbr") and _classe in _CLASSES_NIPLE_SCH160:
            if _corpo_nbr and _extremidade != 'NIPLE 4" COMP. SCH 160':
                return JsonResponse({"success": False, "errors": {"tipo_extremidade": 'Para Esfera com NBR 15827 e classes 900/1500, a extremidade deve ser Niple 4" Comp. SCH 160'}}, status=400)
        # Regras NBR classe 150/300/600/800: extremidade Niple pelo material do corpo
        elif data.get("nbr") and _classe in _CLASSES_NIPLE_NBR:
            if _corpo in _MATERIAIS_SCH160 and _extremidade != 'NIPLE 4" COMP. SCH 160':
                return JsonResponse({"success": False, "errors": {"tipo_extremidade": 'Para Esfera com NBR 15827, classes 150/300/600/800 e corpo carbono/liga, a extremidade deve ser Niple 4" Comp. SCH 160'}}, status=400)
            if _corpo in _MATERIAIS_SCH80 and _extremidade != 'NIPLE 4" COMP. SCH 80':
                return JsonResponse({"success": False, "errors": {"tipo_extremidade": 'Para Esfera com NBR 15827, classes 150/300/600/800 e corpo inox/duplex, a extremidade deve ser Niple 4" Comp. SCH 80'}}, status=400)

    # Regras Esfera: tipo_montagem forçado por diâmetro + classe
    DIAMETROS_PEQUENOS_NBR = {'1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"'}
    DIAMETROS_MEDIOS_NBR = {'2"', '2 1/2"', '3"', '4"'}
    CLASSES_FLUTUANTE_NBR = {'600', '800', '900'}
    CLASSES_FLUTUANTE_MEDIO_NBR = {'150', '300'}
    CLASSES_TRUNNION_PEQUENO_NBR = {'1500', '2500', '4500'}
    CLASSES_TRUNNION_MEDIO_NBR = {'600', '800', '900', '1500', '2500'}
    if tipo_valvula == "ESFERA":
        diametro = data.get("diametro", "")
        classe = data.get("classe", "")
        montagem = data.get("tipo_montagem", "")
        # Regra geral: diâmetro ≥6" + classe sem PN → Trunnion
        if diametro and classe and not classe.startswith("PN") and _parse_diametro(diametro) >= 6:
            if montagem != "TRUNNION":
                return JsonResponse({"success": False, "errors": {"tipo_montagem": "Para Esfera com diâmetro ≥6\" e classe ANSI/API (sem PN), o tipo de montagem deve ser Trunnion"}}, status=400)
        # Regras NBR 15827 para diâmetros ≤4"
        elif data.get("nbr"):
            if diametro in DIAMETROS_PEQUENOS_NBR:
                if classe in CLASSES_FLUTUANTE_NBR and montagem != "FLUTUANTE":
                    return JsonResponse({"success": False, "errors": {"tipo_montagem": "Para Esfera com NBR 15827 nos diâmetros 1/2\" a 1 1/2\" e classes 600/800/900, o tipo de montagem deve ser Flutuante"}}, status=400)
                if classe in CLASSES_TRUNNION_PEQUENO_NBR and montagem != "TRUNNION":
                    return JsonResponse({"success": False, "errors": {"tipo_montagem": "Para Esfera com NBR 15827 nos diâmetros 1/2\" a 1 1/2\" e classes 1500/2500/4500, o tipo de montagem deve ser Trunnion"}}, status=400)
            elif diametro in DIAMETROS_MEDIOS_NBR:
                if classe in CLASSES_FLUTUANTE_MEDIO_NBR and montagem != "FLUTUANTE":
                    return JsonResponse({"success": False, "errors": {"tipo_montagem": "Para Esfera com NBR 15827 nos diâmetros 2\" a 4\" e classes 150/300, o tipo de montagem deve ser Flutuante"}}, status=400)
                if classe in CLASSES_TRUNNION_MEDIO_NBR and montagem != "TRUNNION":
                    return JsonResponse({"success": False, "errors": {"tipo_montagem": "Para Esfera com NBR 15827 nos diâmetros 2\" a 4\" e classes 600/800/900/1500/2500, o tipo de montagem deve ser Trunnion"}}, status=400)

    # ── ISO 14313 (7.2) ────────────────────────────────────────────────────
    # "Valves covered by this International Standard shall be furnished in one of the
    # following classes: PN 20 (class 150); PN 50 (class 300); PN 64 (class 400);
    # PN 100 (class 600); PN 150 (class 900); PN 250 (class 1500); PN 420 (class 2500)".
    #
    # A ISO 14313 é a harmonização da ISO 14313:1999 com a API Spec 6D-2002, mas as duas
    # NÃO são intercambiáveis aqui: a ISO aceita a série PN como designação primária (a
    # API 6D é só Class) e não tem QSL (é conceito da API 6D, Anexo I). Por isso ela tem
    # regra própria em vez de reusar a da API 6D.
    # Class 400 / PN 64 não existe no CLASSES do modelo — fora da lista por isso.
    _CLASSES_ISO14313 = {"150", "300", "600", "900", "1500", "2500",
                         "PN 20", "PN 50", "PN 100", "PN 150", "PN 250", "PN 420"}
    if data.get("norma") == "ISO 14313":
        _classe_iso = data.get("classe") or ""
        if _classe_iso and _classe_iso not in _CLASSES_ISO14313:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma ISO 14313, a classe deve ser 150, 300, 600, 900, 1500, 2500 ou PN 20/50/100/150/250/420. As classes 800, 4500 e PN 10/16/25/40 não são cobertas pela norma"}}, status=400)

    # ── ISO 17292 (Seção 1 + 5.2.7) ────────────────────────────────────────
    # Metal ball valves (só Esfera oferece a norma). Escopo da Seção 1:
    #   - classes: Class 150/300/600/800 e PN 16/25/40/63/100 — Class 800 "applies
    #     only for valves with threaded and socket welding end". PN 63 não existe
    #     no CLASSES do modelo, por isso está fora da lista.
    #   - extremidades: só flange, butt-welding, socket-welding e rosca
    #     (Wafer/Lug/Gray Loc Hub não existem na norma). Niple conta como
    #     socket-welding (NBR C.1.4.1: válvula SW fornecida com niple soldado).
    #   - diâmetro por extremidade: flange/butt-welding 1/2"-24" (DN 15-600);
    #     socket-welding/rosca 1/4"-2" (DN 8-50). O modelo começa em 1/2",
    #     então só o teto precisa de validação.
    # 5.2.7: "Valves shall incorporate an anti-static feature" → dispositivo
    # antiestático obrigatório (vale com ou sem NBR).
    _CLASSES_ISO17292 = {"150", "300", "600", "800", "PN 16", "PN 25", "PN 40", "PN 100"}
    if data.get("norma") == "ISO 17292":
        _classe_17292 = data.get("classe") or ""
        _ext_17292 = data.get("tipo_extremidade") or ""
        _sw_rosca_17292 = (_ext_17292 == "SOCKET-WELDING" or _ext_17292.startswith("ROSCA")
                           or _ext_17292.startswith("NIPLE"))
        _fb_17292 = _ext_17292.startswith("FLANGE") or _ext_17292.startswith("BUTT-WELDING")
        if _classe_17292 and _classe_17292 not in _CLASSES_ISO17292:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma ISO 17292, a classe deve ser 150, 300, 600, 800 ou PN 16/25/40/100"}}, status=400)
        if _ext_17292 and not (_sw_rosca_17292 or _fb_17292):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma ISO 17292, a extremidade deve ser Flange, Butt-Welding, Socket-Welding, Rosca ou Niple (a norma não cobre Wafer/Lug/Gray Loc Hub)"}}, status=400)
        if _classe_17292 == "800" and _fb_17292:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma ISO 17292, a classe 800 só existe em extremidade Rosca ou Socket-Welding"}}, status=400)
        _diam_17292 = data.get("diametro")
        _dn_17292 = _parse_diametro(_diam_17292) if _diam_17292 else None
        if _dn_17292 is not None:
            if _fb_17292 and _dn_17292 > 24:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma ISO 17292 com extremidade Flange/Butt-Welding, o diâmetro máximo é 24"'}}, status=400)
            if _sw_rosca_17292 and _dn_17292 > 2:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma ISO 17292 com extremidade Socket-Welding/Rosca, o diâmetro máximo é 2"'}}, status=400)
        if tipo_valvula == "ESFERA" and not data.get("dispositivo_antiestatico"):
            return JsonResponse({"success": False, "errors": {"dispositivo_antiestatico": "Para norma ISO 17292 (5.2.7), o dispositivo antiestático é obrigatório"}}, status=400)

    # ── BS ISO 7121 (Cláusula 1 + 5.1/Tabela 2) ─────────────────────────────
    # Steel ball valves for general-purpose industrial applications (só Esfera oferece
    # a norma). Escopo da Cláusula 1:
    #   - classes: Class 150/300/600/900 e PN 10/16/25/40/63/100 — Class 900 "only
    #     valves having reduced port are within the scope" (5.1/Tabela 2, nota de
    #     rodapé). PN 63 não existe no CLASSES do modelo (mesmo caso do PN 63 da
    #     ISO 17292 e do PN 64 da ISO 14313), por isso está fora da lista.
    #   - extremidades: só flange, butt-welding, socket-welding e rosca (Wafer/Lug/
    #     Gray Loc Hub não existem na norma). Niple conta como socket-welding
    #     (NBR C.1.4.1: válvula SW fornecida com niple soldado).
    #   - diâmetro por extremidade: flange/butt-welding 1/2"-20" (DN 15-500);
    #     socket-welding 1/4"-4" (DN 8-100); rosca 1/4"-2" (DN 8-50) — diferente da
    #     ISO 17292, aqui socket-welding vai mais longe que rosca.
    # 5.2.7: o antiestático é "when specified in the purchase order" — opcional, NÃO
    # trava (diferente da ISO 17292, onde é mandatório).
    _CLASSES_BS7121 = {"150", "300", "600", "900", "PN 10", "PN 16", "PN 25", "PN 40", "PN 100"}
    if data.get("norma") == "BS ISO 7121":
        _classe_7121 = data.get("classe") or ""
        _ext_7121 = data.get("tipo_extremidade") or ""
        _rosca_7121 = _ext_7121.startswith("ROSCA")
        _sw_7121 = _ext_7121 == "SOCKET-WELDING" or _ext_7121.startswith("NIPLE")
        _fb_7121 = _ext_7121.startswith("FLANGE") or _ext_7121.startswith("BUTT-WELDING")
        if _classe_7121 and _classe_7121 not in _CLASSES_BS7121:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma BS ISO 7121, a classe deve ser 150, 300, 600, 900 ou PN 10/16/25/40/100"}}, status=400)
        if _ext_7121 and not (_rosca_7121 or _sw_7121 or _fb_7121):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma BS ISO 7121, a extremidade deve ser Flange, Butt-Welding, Socket-Welding, Rosca ou Niple (a norma não cobre Wafer/Lug/Gray Loc Hub)"}}, status=400)
        if _classe_7121 == "900" and (data.get("tipo_passagem") or "") == "PLENA":
            return JsonResponse({"success": False, "errors": {"tipo_passagem": "Para norma BS ISO 7121 e classe 900, a passagem deve ser Reduzida (a norma só cobre passagem reduzida nessa classe)"}}, status=400)
        _diam_7121 = data.get("diametro")
        _dn_7121 = _parse_diametro(_diam_7121) if _diam_7121 else None
        if _dn_7121 is not None:
            if _fb_7121 and _dn_7121 > 20:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma BS ISO 7121 com extremidade Flange/Butt-Welding, o diâmetro máximo é 20"'}}, status=400)
            if _sw_7121 and _dn_7121 > 4:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma BS ISO 7121 com extremidade Socket-Welding, o diâmetro máximo é 4"'}}, status=400)
            if _rosca_7121 and _dn_7121 > 2:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma BS ISO 7121 com extremidade Rosca, o diâmetro máximo é 2"'}}, status=400)

    # ── API 608 (Cláusula 1) ─────────────────────────────────────────────────
    # Metal Ball Valves—Flanged, Threaded, and Welding Ends (só Esfera oferece a norma).
    # 1.2: cobre requisitos adicionais aos de ASME B16.34 Standard Class (a regra
    # genérica da B16.34 já roda depois, no fim de _validar_regras_valvula). Escopo
    # da Cláusula 1:
    #   - 1.1: "butt-welding or flanged ends for ... NPS 1/2 through NPS 12 and threaded
    #     or socket-welding ends for sizes NPS 1/2 through NPS 2" → só Flange/Butt-Welding/
    #     Socket-Welding/Rosca (Niple conta como SW, NBR C.1.4.1) — a norma não cobre
    #     Wafer/Lug/Gray Loc Hub. Diâmetro Flange/BW até 12"; Socket-Welding/Rosca até 2"
    #   - 1.3: "flanged and butt-welding end valves in Standard Classes 150 and 300 and
    #     threaded and socket-welding end valves in Standard Classes 150, 300, and 600"
    #     → Flange/BW só 150 ou 300 (sem 600); Socket-Welding/Rosca 150, 300 ou 600
    # 4.4 (continuidade elétrica/antiestático) é "when specified in the purchase order"
    # — opcional, não trava (mesmo caso da BS ISO 7121, diferente da ISO 17292).
    _CLASSES_API608_FB = {"150", "300"}
    _CLASSES_API608_SW = {"150", "300", "600"}
    if tipo_valvula == "ESFERA" and data.get("norma") == "API 608":
        _ext_608 = data.get("tipo_extremidade") or ""
        _sw_rosca_608 = (_ext_608 == "SOCKET-WELDING" or _ext_608.startswith("ROSCA")
                         or _ext_608.startswith("NIPLE"))
        _fb_608 = _ext_608.startswith("FLANGE") or _ext_608.startswith("BUTT-WELDING")
        if _ext_608 and not (_sw_rosca_608 or _fb_608):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma API 608, a extremidade deve ser Flange, Butt-Welding, Socket-Welding ou Rosca (a norma não cobre Wafer/Lug/Gray Loc Hub)"}}, status=400)
        _classe_608 = data.get("classe") or ""
        if _classe_608 and _fb_608 and _classe_608 not in _CLASSES_API608_FB:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma API 608 com extremidade Flange/Butt-Welding, a classe deve ser 150 ou 300 (a norma não cobre 600 nessa extremidade)"}}, status=400)
        if _classe_608 and _sw_rosca_608 and _classe_608 not in _CLASSES_API608_SW:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma API 608 com extremidade Socket-Welding/Rosca, a classe deve ser 150, 300 ou 600"}}, status=400)
        _diam_608 = data.get("diametro")
        _dn_608 = _parse_diametro(_diam_608) if _diam_608 else None
        if _dn_608 is not None:
            if _fb_608 and _dn_608 > 12:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma API 608 com extremidade Flange/Butt-Welding, o diâmetro máximo é 12"'}}, status=400)
            if _sw_rosca_608 and _dn_608 > 2:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma API 608 com extremidade Socket-Welding/Rosca, o diâmetro máximo é 2"'}}, status=400)

    # ── MSS SP-72 (Seções 1, 2, 3) ────────────────────────────────────────────
    # Ball Valves with Flanged or Butt-Welding Ends for General Service (só Esfera
    # oferece a norma, valor do choice é "MSS SP72"). Escopo da Seção 1:
    #   - 1.1: "flanged or butt-weld end ball valves" → extremidade só Flange ou
    #     Butt-Welding (a norma não cobre Socket-Welding/Rosca/Niple/Wafer/Lug/
    #     Gray Loc Hub — diferente da ISO 17292/BS ISO 7121/API 608, que cobrem SW/Rosca)
    #   - 1.3: "NPS 1/2 (DN 15) through NPS 36 (DN 900)" → diâmetro máximo 36" (o
    #     modelo já começa em 1/2", então só o teto precisa de validação)
    # 2.1 (rating por material: carbon/alloy/stainless steel → ASME B16.5/B16.34; ductile
    # iron → B16.42; gray iron → B16.1; copper alloy → B16.24) não vira trava adicional
    # de classe: os materiais de corpo do modelo (MATERIAIS_CORPO_TAMPA) são só aço
    # carbono/liga/inox, cobertos pela 2.1.1-2.1.3 (mesma designação Class da B16.34,
    # já validada de forma genérica no fim desta função) — não existe opção de ferro
    # fundido/dútil ou liga de cobre no modelo, então 2.1.4-2.1.6 ficam fora de alcance.
    if tipo_valvula == "ESFERA" and data.get("norma") == "MSS SP72":
        _ext_sp72 = data.get("tipo_extremidade") or ""
        _fb_sp72 = _ext_sp72.startswith("FLANGE") or _ext_sp72.startswith("BUTT-WELDING")
        if _ext_sp72 and not _fb_sp72:
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma MSS SP72, a extremidade deve ser Flange ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/Niple/Wafer/Lug/Gray Loc Hub)"}}, status=400)
        _diam_sp72 = data.get("diametro")
        _dn_sp72 = _parse_diametro(_diam_sp72) if _diam_sp72 else None
        if _dn_sp72 is not None and _dn_sp72 > 36:
            return JsonResponse({"success": False, "errors": {"diametro": 'Para norma MSS SP72, o diâmetro máximo é 36" (NPS 36 / DN 900)'}}, status=400)

    # ── NBR 14788 (Objetivo/Seções 1, 5, 6) ───────────────────────────────────
    # Válvulas de esfera - Requisitos (baseada na ISO 7121:1986 e API 6D:1994; só Esfera
    # oferece a norma). Objetivo: "extremidades roscadas, flangeadas ou soldadas ...
    # diâmetros nominais DN 10 a DN 500, nas pressões nominais ISO PN 10 a ISO PN 100,
    # como definido nas seções 5 e 6". "Soldadas" aqui é solda de topo (8.1.3.1 cita a
    # ASME/ANSI B16.25 — Buttwelding ends — para o encaixe de solda; a Tabela 7 do
    # documento, "extremidades de solda de topo", confirma) — não há Socket-Welding em
    # nenhum lugar do texto (diferente da ISO 17292/BS ISO 7121/API 608, que o cobrem).
    #   - classe (Seções 5/6): só PN 10/16/20/25/40/50/100 — a norma usa só a série ISO
    #     PN como designação de pressão, sem Class ASME
    #   - extremidade: só Flange, Butt-Welding ou Rosca (a norma não cobre
    #     Socket-Welding/Niple/Wafer/Lug/Gray Loc Hub)
    #   - diâmetro (Seção 5, DN 10-500): DN 500 = NPS 20" é o teto; DN 10 não tem
    #     equivalente em NPS no modelo (mínimo já é 1/2"/DN 15), então só o teto importa
    #   - PN 20 só em Rosca: as Tabelas 6 e 7 (dimensão face a face) agrupam as colunas
    #     de pressão em "10 e 16" / "25 a 50" / "100" — PN 20 não tem coluna em nenhuma
    #     das duas (confirmado pela contagem de valores por linha, 5 colunas batendo
    #     exatamente com os 3 grupos, sem espaço para um 4º). Como 8.1.3.1 torna a
    #     dimensão face a face OBRIGATÓRIA para Flange/Butt-Welding (só Rosca e
    #     "encaixes para solda" — aqui sem uso, ver acima — são isentos), PN 20 fica sem
    #     dimensão válida nessas duas extremidades — só sobra Rosca, que não tem
    #     requisito de face a face.
    # 8.6 (antiestático) e 8.1.3.6 (dreno) são "quando especificado pelo cliente" —
    # opcionais, não travam (mesmo caso da BS ISO 7121/API 608).
    _CLASSES_NBR14788 = {"PN 10", "PN 16", "PN 20", "PN 25", "PN 40", "PN 50", "PN 100"}
    if tipo_valvula == "ESFERA" and data.get("norma") == "NBR 14788":
        _ext_14788 = data.get("tipo_extremidade") or ""
        if _ext_14788 and not (_ext_14788.startswith("FLANGE") or _ext_14788.startswith("BUTT-WELDING") or _ext_14788.startswith("ROSCA")):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma NBR 14788, a extremidade deve ser Flange, Butt-Welding ou Rosca (a norma não cobre Socket-Welding/Niple/Wafer/Lug/Gray Loc Hub)"}}, status=400)
        _classe_14788 = data.get("classe") or ""
        if _classe_14788 and _classe_14788 not in _CLASSES_NBR14788:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma NBR 14788, a classe deve ser PN 10, 16, 20, 25, 40, 50 ou 100 (a norma usa só a série ISO PN, sem Class ASME)"}}, status=400)
        if _classe_14788 == "PN 20" and (_ext_14788.startswith("FLANGE") or _ext_14788.startswith("BUTT-WELDING")):
            return JsonResponse({"success": False, "errors": {"classe": "Para norma NBR 14788 com extremidade Flange/Butt-Welding, a classe não pode ser PN 20 (as Tabelas 6/7 de dimensão face a face não cobrem PN 20 — só Rosca, que é isenta desse requisito)"}}, status=400)
        _diam_14788 = data.get("diametro")
        _dn_14788 = _parse_diametro(_diam_14788) if _diam_14788 else None
        if _dn_14788 is not None and _dn_14788 > 20:
            return JsonResponse({"success": False, "errors": {"diametro": 'Para norma NBR 14788, o diâmetro máximo é 20" (DN 500)'}}, status=400)

    # ── ISO 15761 (Cláusula 1) ───────────────────────────────────────────────
    # Steel gate, globe and check valves for sizes DN 100 and smaller, petroleum/gas
    # industry — a Introdução diz que o "general construction parallels that specified
    # by API 602 and BS 5352" (é a versão ISO da API 602). Só Gaveta/Globo/Retenção
    # oferecem a norma. Escopo da Cláusula 1:
    #   - classes: Class 150/300/600/800/1500 — sem 900 (nota da 4.1: "Class 900 is not
    #     specifically referenced... because seldom used for the compact valves described
    #     herein"), sem 2500/4500 nem PN
    #   - extremidades: só flange, butt-welding, socket-welding e rosca (Wafer/Lug/Gray
    #     Loc Hub não existem na norma). Niple conta como socket-welding (NBR C.1.4.1)
    #   - "socket welding or threaded ends, in sizes 8≤DN≤65 and pressure designations of
    #     Class 800 and Class 1500" → SW/Rosca só em classe 800/1500 (não 150/300/600),
    #     diâmetro até 2 1/2"
    #   - "flanged or butt-welding ends, in sizes 15≤DN≤100 and 150≤Class≤1500, excluding
    #     flanged end Class 800" → Flange/BW até 4"; Flange não cobre classe 800 (BW cobre)
    #   - 5.5.1: bonnet fixado por 1 de 4 métodos (bolting/welding/threaded com seal weld/
    #     threaded union nut). "threaded union nut, provided it is of Class ⩽ 800" →
    #     junção Roscado (union nut) só até classe 800. Pressure Seal não é um dos 4
    #     métodos (mesmo caso da API 602, texto idêntico) → vedação não pode ser Pressure Seal
    _CLASSES_ISO15761 = {"150", "300", "600", "800", "1500"}
    if data.get("norma") == "ISO 15761":
        _classe_15761 = data.get("classe") or ""
        _ext_15761 = data.get("tipo_extremidade") or ""
        _sw_rosca_15761 = (_ext_15761 == "SOCKET-WELDING" or _ext_15761.startswith("ROSCA")
                           or _ext_15761.startswith("NIPLE"))
        _flange_15761 = _ext_15761.startswith("FLANGE")
        _bw_15761 = _ext_15761.startswith("BUTT-WELDING")
        if _classe_15761 and _classe_15761 not in _CLASSES_ISO15761:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma ISO 15761, a classe deve ser 150, 300, 600, 800 ou 1500"}}, status=400)
        if _ext_15761 and not (_sw_rosca_15761 or _flange_15761 or _bw_15761):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma ISO 15761, a extremidade deve ser Flange, Butt-Welding, Socket-Welding, Rosca ou Niple (a norma não cobre Wafer/Lug/Gray Loc Hub)"}}, status=400)
        if _classe_15761 == "800" and _flange_15761:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma ISO 15761 com extremidade Flange, a classe não pode ser 800 (a norma não cobre flangeado nessa classe — só Butt-Welding, Socket-Welding ou Rosca)"}}, status=400)
        if _classe_15761 in ("150", "300", "600") and _sw_rosca_15761:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma ISO 15761 com extremidade Socket-Welding/Rosca/Niple, a classe deve ser 800 ou 1500"}}, status=400)
        _diam_15761 = data.get("diametro")
        _dn_15761 = _parse_diametro(_diam_15761) if _diam_15761 else None
        if _dn_15761 is not None:
            if (_flange_15761 or _bw_15761) and _dn_15761 > 4:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma ISO 15761 com extremidade Flange/Butt-Welding, o diâmetro máximo é 4"'}}, status=400)
            if _sw_rosca_15761 and _dn_15761 > 2.5:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma ISO 15761 com extremidade Socket-Welding/Rosca, o diâmetro máximo é 2 1/2"'}}, status=400)
        if data.get("juncao_corpo_castelo") == "ROSCADO" and _classe_15761 == "1500":
            return JsonResponse({"success": False, "errors": {"juncao_corpo_castelo": "Para norma ISO 15761, junção Roscado (union nut) só é permitida até classe 800 (não cobre classe 1500)"}}, status=400)
        if _junta_categoria(data) == "PRESSURE SEAL":
            return JsonResponse({"success": False, "errors": {"materiais": "Para norma ISO 15761, a categoria da junta não pode ser Pressure Seal (5.5.1 lista só 4 métodos de junção corpo/castelo — bolting, welding, threaded com seal weld, threaded union nut — pressure seal não é um deles)"}}, status=400)

    # ── BS 1873 (Cláusulas 1, 4, 6, 9.3, Apêndice A) ──────────────────────────
    # Steel globe and globe stop and check valves (flanged/butt-welding ends) — só Globo
    # oferece a norma. Escopo:
    #   - classe (Cláusula 4): 150/300/400/600/900/1500/2500 — sem 800 nem PN. Class 400
    #     não existe no CLASSES do modelo (mesmo caso do PN 64/Class 400 da ISO 14313),
    #     então na prática sobra {150,300,600,900,1500,2500}
    #   - extremidade (Cláusula 1): só Flange ou Butt-Welding — a norma não cobre
    #     Socket-Welding/Rosca/Niple/Wafer/Lug/Gray Loc Hub
    #   - diâmetro por classe (Apêndice A, Tabelas 3-9): cada classe tem faixa própria de
    #     DN — 150: 1/2"-16"; 300/600/2500: 1/2"-12"; 1500: 1/2"-14"; 900: 3"-14" (única
    #     com PISO, não só teto — a Tabela 7 começa em DN 80/NPS 3, não tem tamanho menor)
    #   - 9.3: "the body to bonnet connection shall be flanged" — bonnet sempre flange
    #     aparafusado (forçado em _aplicar_regras_automaticas, igual ao padrão API 600/
    #     ISO 10434); os 3 tipos de facing (male-and-female/tongue-and-groove/ring joint,
    #     +flat face só p/ classe 150) não incluem bonnet soldado nem pressure-seal →
    #     vedação (junta) não pode ser Castelo Soldado nem Pressure Seal (mesma restrição
    #     da ISO 10434) — só Junta Espiralada ou RTJ (FJA)
    _CLASSES_BS1873 = {"150", "300", "600", "900", "1500", "2500"}
    # (min_num, max_num, min_display, max_display) — display evita "0.5" no erro
    _DIAM_POR_CLASSE_BS1873 = {
        "150": (0.5, 16, '1/2"', '16"'),
        "300": (0.5, 12, '1/2"', '12"'),
        "600": (0.5, 12, '1/2"', '12"'),
        "900": (3, 14, '3"', '14"'),
        "1500": (0.5, 14, '1/2"', '14"'),
        "2500": (0.5, 12, '1/2"', '12"'),
    }
    if data.get("norma") == "BS 1873":
        _classe_1873 = data.get("classe") or ""
        _ext_1873 = data.get("tipo_extremidade") or ""
        if _classe_1873 and _classe_1873 not in _CLASSES_BS1873:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma BS 1873, a classe deve ser 150, 300, 600, 900, 1500 ou 2500"}}, status=400)
        if _ext_1873 and not (_ext_1873.startswith("FLANGE") or _ext_1873.startswith("BUTT-WELDING")):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma BS 1873, a extremidade deve ser Flange ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/Niple/Wafer/Lug/Gray Loc Hub)"}}, status=400)
        _diam_1873 = data.get("diametro")
        _dn_1873 = _parse_diametro(_diam_1873) if _diam_1873 else None
        if _dn_1873 is not None and _classe_1873 in _DIAM_POR_CLASSE_BS1873:
            _min_1873, _max_1873, _min_disp_1873, _max_disp_1873 = _DIAM_POR_CLASSE_BS1873[_classe_1873]
            if _dn_1873 < _min_1873 or _dn_1873 > _max_1873:
                return JsonResponse({"success": False, "errors": {"diametro": f'Para norma BS 1873 e classe {_classe_1873}, o diâmetro deve ser de {_min_disp_1873} a {_max_disp_1873}'}}, status=400)
        # juncao_corpo_castelo já vem forçado p/ APARAFUSADO por _aplicar_regras_automaticas
        # (única opção dentro do escopo — 9.3 "the body to bonnet connection shall be
        # flanged"); nada a validar aqui.
        if _junta_categoria(data) in ("CASTELO SOLDADO", "PRESSURE SEAL"):
            return JsonResponse({"success": False, "errors": {"materiais": "Para norma BS 1873, a categoria da junta deve ser Junta Espiralada ou RTJ (FJA) (9.3 restringe o bonnet-to-body joint a facings de flange — male-and-female/tongue-and-groove/ring joint, ou flat face só na classe 150 — sem opção soldada nem pressure-seal)"}}, status=400)

    # ── API 623 (Seção 1) ────────────────────────────────────────────────────
    # Steel Globe Valves—Flanged and Butt-Welding Ends, Bolted Bonnets (só Globo oferece
    # a norma). Escopo da Seção 1 ("This standard sets forth the requirements for the
    # following globe valve features: bolted bonnet, ... flanged or butt-welding ends"):
    #   - bonnet aparafusado: "bolted bonnet" é característica definidora do escopo
    #     (forçado em _aplicar_regras_automaticas, mesmo padrão API 600/ISO 10434/BS 1873)
    #   - extremidade só Flange ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/
    #     Niple/Wafer/Lug/Gray Loc Hub — 5.3.3/5.3.4 citam SW/rosca como referência
    #     normativa genérica, mas o Escopo da Seção 1 só oferece "flanged or
    #     butt-welding ends")
    #   - classe em {150, 300, 600, 900, 1500, 2500} — sem 800 nem PN (mesma designação
    #     da ASME B16.34; Tabelas 1/2/4/5 usam essas 6 classes)
    #   - diâmetro: NPS 2, 2 1/2, 3, 4, 6, 8, 10, 12, 14, 16, 18, 20, 24 (DN 50-600) —
    #     sem 1/2"-1 1/2" (norma começa em NPS 2, diferente da API 600/ISO 10434, que
    #     começam em NPS 1), sem 22" (mesmo buraco da API 600/ISO 10434/BS 1868) nem
    #     acima de 24"
    #   - 5.5.1/5.5.2: bonnet-to-body joint "shall be a flange and gasket type" — raised
    #     face/tongue-and-groove/spigot-and-recess/ring-joint — nenhum soldado nem
    #     pressure-seal → vedação (junta) não pode ser Castelo Soldado nem Pressure Seal
    #     (mesma restrição da ISO 10434, nenhuma das duas cabe no escopo)
    #   - classe 2500 só até NPS 12: a Tabela 1 (espessura mínima de parede, mandatória
    #     por 5.1) não tem valor pra Classe 2500 em NPS 14-24 (célula "—") — acima de
    #     NPS 12 a norma simplesmente não define espessura pra essa classe, mesmo padrão
    #     do buraco PN 20/face-a-face da NBR 14788
    _DIAMETROS_API623 = {
        '2"', '2 1/2"', '3"', '4"', '6"', '8"', '10"',
        '12"', '14"', '16"', '18"', '20"', '24"',
    }
    _CLASSES_API623 = {"150", "300", "600", "900", "1500", "2500"}
    if tipo_valvula == "GLOBO" and data.get("norma") == "API 623":
        # juncao_corpo_castelo já vem forçado p/ APARAFUSADO por _aplicar_regras_automaticas
        # (única opção dentro do escopo — "bolted bonnet"); nada a validar aqui.
        _ext_623 = data.get("tipo_extremidade") or ""
        if _ext_623 and not (_ext_623.startswith("FLANGE") or _ext_623.startswith("BUTT-WELDING")):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma API 623, a extremidade deve ser Flange ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/Niple/Wafer/Lug/Gray Loc Hub)"}}, status=400)
        _classe_623 = data.get("classe") or ""
        if _classe_623 and _classe_623 not in _CLASSES_API623:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma API 623, a classe deve ser 150, 300, 600, 900, 1500 ou 2500"}}, status=400)
        _diam_623 = data.get("diametro") or ""
        if _diam_623 and _diam_623 not in _DIAMETROS_API623:
            return JsonResponse({"success": False, "errors": {"diametro": 'Para norma API 623, o diâmetro deve ser de 2" a 24" (a norma não cobre 1/2" a 1 1/2", 22" nem acima de 24")'}}, status=400)
        if _classe_623 == "2500" and _diam_623 and _parse_diametro(_diam_623) > 12:
            return JsonResponse({"success": False, "errors": {"diametro": 'Para norma API 623 e classe 2500, o diâmetro máximo é 12" (a Tabela 1 de espessura mínima de parede não cobre Classe 2500 acima de NPS 12)'}}, status=400)
        if _junta_categoria(data) in ("CASTELO SOLDADO", "PRESSURE SEAL"):
            return JsonResponse({"success": False, "errors": {"materiais": "Para norma API 623, a categoria da junta deve ser Junta Espiralada ou RTJ (FJA) (5.5.1/5.5.2 restringem o bonnet-to-body joint a facings de flange — raised face/tongue-and-groove/spigot-and-recess/ring-joint — sem opção soldada nem pressure-seal)"}}, status=400)

    # ── API 600 (Seção 1) ────────────────────────────────────────────────────
    # Steel Gate Valves—Flanged and Butt-welding Ends, Bolted Bonnets (só Gaveta oferece
    # a norma). Escopo da Seção 1 ("This standard sets forth the requirements for the
    # following gate valve features: bolted bonnet, ... flanged or butt-welding ends"):
    #   - bonnet aparafusado: "bolted bonnet" é característica definidora do título/escopo
    #     (não "solda"/"rosca" — essas ficam pra outras normas de gaveta, ex. API 602/
    #     ISO 15761 pra classes/tamanhos pequenos com bonnet integral)
    #   - extremidade só Flange ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/
    #     Niple/Wafer/Lug/Gray Loc Hub)
    #   - classe em {150, 300, 600, 900, 1500, 2500} — sem 800 nem PN (mesma designação
    #     de classe da ASME B16.34, citada em 4.1 pras tabelas P-T)
    #   - diâmetro DN 25-1050 (NPS 1-42), com buraco: a norma não cobre NPS 22 (a lista
    #     pula de 20" pra 24"), nem abaixo de 1" (sem 1/2"/3/4"), nem acima de 42"
    _DIAMETROS_API600 = {
        '1"', '1 1/4"', '1 1/2"', '2"', '2 1/2"', '3"', '4"', '6"', '8"', '10"',
        '12"', '14"', '16"', '18"', '20"', '24"', '26"', '28"', '30"', '32"',
        '34"', '36"', '38"', '40"', '42"',
    }
    _CLASSES_API600 = {"150", "300", "600", "900", "1500", "2500"}
    if tipo_valvula == "GAVETA" and data.get("norma") == "API 600":
        # juncao_corpo_castelo já vem forçado p/ APARAFUSADO por _aplicar_regras_automaticas
        # (única opção dentro do escopo — "bolted bonnet"); nada a validar aqui.
        _ext_600 = data.get("tipo_extremidade") or ""
        if _ext_600 and not (_ext_600.startswith("FLANGE") or _ext_600.startswith("BUTT-WELDING")):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma API 600, a extremidade deve ser Flange ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/Niple/Wafer/Lug/Gray Loc Hub)"}}, status=400)
        _classe_600 = data.get("classe") or ""
        if _classe_600 and _classe_600 not in _CLASSES_API600:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma API 600, a classe deve ser 150, 300, 600, 900, 1500 ou 2500"}}, status=400)
        _diam_600 = data.get("diametro") or ""
        if _diam_600 and _diam_600 not in _DIAMETROS_API600:
            return JsonResponse({"success": False, "errors": {"diametro": 'Para norma API 600, o diâmetro deve ser de 1" a 42" (a norma não cobre 1/2", 3/4", 22" nem acima de 42")'}}, status=400)
        # 5.5.11 cobre bonnet pressure-seal como opção dentro do escopo, mas "Castelo
        # Soldado" na categoria da junta é bonnet SOLDADO — contradiz o bonnet aparafusado
        # da Seção 1. Lê a categoria da junta (materiais, tipo_material=JUNTA).
        if _junta_categoria(data) == "CASTELO SOLDADO":
            return JsonResponse({"success": False, "errors": {"materiais": 'Para norma API 600, a categoria da junta não pode ser Castelo Soldado (a norma cobre só "bolted bonnet")'}}, status=400)

    # ── ISO 10434 (Seção 1) ──────────────────────────────────────────────────
    # Bolted Bonnet Steel Gate Valves (só Gaveta oferece a norma). É a versão ISO/EN da
    # API 600 (a própria Introdução diz que "parallel those given in API Standard 600") —
    # mesmo escopo de tipo/extremidade/classe, mas faixa de diâmetro menor (DN 25-600,
    # NPS 1-24, contra NPS 1-42 da API 600) e SEM opção de bonnet pressure-seal (a API 600
    # tem 5.5.11 permitindo; a ISO 10434 não tem cláusula equivalente — 5.5.1 restringe o
    # bonnet-to-body joint a "flange and gasket type", sem exceção).
    #   - bonnet aparafusado: mesma característica definidora do escopo (Seção 1)
    #   - extremidade só Flange ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/
    #     Niple/Wafer/Lug/Gray Loc Hub)
    #   - classe em {150, 300, 600, 900, 1500, 2500} — sem 800 nem PN
    #   - diâmetro DN 25-600 (NPS 1-24), com o mesmo buraco no NPS 22 (a lista pula de
    #     20" pra 24"), nem abaixo de 1" (sem 1/2"/3/4"), nem acima de 24"
    #   - vedação (junta) não pode ser Castelo Soldado NEM Pressure Seal (diferente da
    #     API 600, aqui nenhuma das duas cabe no escopo) — só Junta Espiralada ou RTJ (FJA)
    _DIAMETROS_ISO10434 = {
        '1"', '1 1/4"', '1 1/2"', '2"', '2 1/2"', '3"', '4"', '6"', '8"', '10"',
        '12"', '14"', '16"', '18"', '20"', '24"',
    }
    _CLASSES_ISO10434 = {"150", "300", "600", "900", "1500", "2500"}
    if tipo_valvula == "GAVETA" and data.get("norma") == "ISO 10434":
        # juncao_corpo_castelo já vem forçado p/ APARAFUSADO por _aplicar_regras_automaticas
        # (única opção dentro do escopo — "bolted bonnet"); nada a validar aqui.
        _ext_10434 = data.get("tipo_extremidade") or ""
        if _ext_10434 and not (_ext_10434.startswith("FLANGE") or _ext_10434.startswith("BUTT-WELDING")):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma ISO 10434, a extremidade deve ser Flange ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/Niple/Wafer/Lug/Gray Loc Hub)"}}, status=400)
        _classe_10434 = data.get("classe") or ""
        if _classe_10434 and _classe_10434 not in _CLASSES_ISO10434:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma ISO 10434, a classe deve ser 150, 300, 600, 900, 1500 ou 2500"}}, status=400)
        _diam_10434 = data.get("diametro") or ""
        if _diam_10434 and _diam_10434 not in _DIAMETROS_ISO10434:
            return JsonResponse({"success": False, "errors": {"diametro": 'Para norma ISO 10434, o diâmetro deve ser de 1" a 24" (a norma não cobre 1/2", 3/4", 22" nem acima de 24")'}}, status=400)
        if _junta_categoria(data) in ("CASTELO SOLDADO", "PRESSURE SEAL"):
            return JsonResponse({"success": False, "errors": {"materiais": "Para norma ISO 10434, a categoria da junta deve ser Junta Espiralada ou RTJ (FJA) (a norma restringe o bonnet-to-body joint a flange e junta, sem opção soldada nem pressure-seal)"}}, status=400)

    # ── API 602 (Seção 1) ────────────────────────────────────────────────────
    # Gate, Globe, and Check Valves for Sizes DN 100 (NPS 4) and Smaller (Gaveta, Globo
    # e Retenção oferecem a norma). Série compacta/forjada. Escopo da Seção 1:
    #   - diâmetro sempre <= 4" (DN 100) — a norma não cobre nada acima disso
    #   - classe em {150, 300, 600, 800, 1500} — sem 900, 2500, 4500 nem PN (diferente
    #     da API 600/ISO 10434: aqui o 800 ENTRA — API 602 é a norma-base do 800)
    #   - Flange -> classe != 800 ("flanged or butt-welding ends ... excluding flanged
    #     end class 800"); diâmetro 1/2"-4"
    #   - Butt-Welding -> diâmetro 1/2"-4", qualquer classe do escopo (800 inclusive —
    #     só o Flange exclui 800, não o Butt-Welding)
    #   - Socket-Welding/Rosca -> diâmetro 1/4"-2 1/2" (modelo não desce de 1/2", então
    #     efetivamente <= 2 1/2"); classe só 800 ou 1500 ("socket welding or threaded
    #     ends ... pressure designations of class 800 and class 1500"). Niple conta
    #     como SW (NBR C.1.4.1); Wafer/Lug/Gray Loc Hub não existem na norma
    #   - junção corpo/castelo (5.5.1 lista 4 métodos: bolting/welding/threaded com
    #     seal weld/threaded union nut): APARAFUSADO/SOLDADO/ROSCADO E SOLDADO sem
    #     limite extra de classe; ROSCADO (union nut) só até classe 800 ("threaded
    #     union nut, provided the valve is class <= 800")
    #   - vedação (junta) não pode ser Pressure Seal — não é um dos 4 métodos de 5.5.1
    #     e a palavra não aparece em nenhum lugar do texto da norma
    _CLASSES_API602 = {"150", "300", "600", "800", "1500"}
    if tipo_valvula in ("GAVETA", "GLOBO", "RETENCAO") and data.get("norma") == "API 602":
        _classe_602 = data.get("classe") or ""
        _ext_602 = data.get("tipo_extremidade") or ""
        _sw_rosca_602 = (_ext_602 == "SOCKET-WELDING" or _ext_602.startswith("ROSCA")
                         or _ext_602.startswith("NIPLE"))
        _flange_602 = _ext_602.startswith("FLANGE")
        _bw_602 = _ext_602.startswith("BUTT-WELDING")
        if _classe_602 and _classe_602 not in _CLASSES_API602:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma API 602, a classe deve ser 150, 300, 600, 800 ou 1500"}}, status=400)
        if _ext_602 and not (_sw_rosca_602 or _flange_602 or _bw_602):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma API 602, a extremidade deve ser Flange, Butt-Welding, Socket-Welding ou Rosca (a norma não cobre Wafer/Lug/Gray Loc Hub)"}}, status=400)
        if _classe_602 == "800" and _flange_602:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma API 602 com extremidade Flange, a classe não pode ser 800 (a norma exclui flange de classe 800 — 800 só existe em Butt-Welding, Socket-Welding ou Rosca)"}}, status=400)
        if _classe_602 in ("150", "300", "600") and _sw_rosca_602:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma API 602 com extremidade Socket-Welding/Rosca/Niple, a classe deve ser 800 ou 1500"}}, status=400)
        _diam_602 = data.get("diametro")
        _dn_602 = _parse_diametro(_diam_602) if _diam_602 else None
        if _dn_602 is not None:
            if (_flange_602 or _bw_602) and _dn_602 > 4:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma API 602 com extremidade Flange/Butt-Welding, o diâmetro máximo é 4"'}}, status=400)
            if _sw_rosca_602 and _dn_602 > 2.5:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma API 602 com extremidade Socket-Welding/Rosca/Niple, o diâmetro máximo é 2 1/2"'}}, status=400)
        if data.get("juncao_corpo_castelo") == "ROSCADO" and _classe_602 == "1500":
            return JsonResponse({"success": False, "errors": {"juncao_corpo_castelo": "Para norma API 602, junção Roscado (union nut) só é permitida até classe 800 (não cobre classe 1500)"}}, status=400)
        if _junta_categoria(data) == "PRESSURE SEAL":
            return JsonResponse({"success": False, "errors": {"materiais": "Para norma API 602, a categoria da junta não pode ser Pressure Seal (5.5.1 lista só 4 métodos de junção corpo/castelo — bolting, welding, threaded com seal weld, threaded union nut — pressure seal não é um deles)"}}, status=400)

    # ── BS 1868 (Cláusulas 1, 4, 6, 9.3) ─────────────────────────────────────
    # Steel check valves (flanged and butt-welding ends) — só Retenção oferece a norma
    # (NORMA_RETENCAO). Escopo da Cláusula 1 ("cast or forged steel check valves with
    # flanged or butt-welding ends"):
    #   - tipo_extremidade só Flange ou Butt-Welding — a norma não cobre Socket-Welding/
    #     Rosca/Niple/Wafer/Lug/Gray Loc Hub (mesmo escopo restrito da API 600/ISO 10434)
    #   - classe (Cláusula 4): "Classes 150, 300, 400, 600, 900, 1 500 and 2 500" — sem 800
    #     (designação de forjado — fora do escopo desta série cast/BW) nem PN. Classe 400
    #     adicionada ao CLASSES_RETENCAO_GLOBO do modelo especificamente pra essa norma —
    #     ASME B16.34 e API 6D (as outras normas de Retenção que usam essa mesma lista de
    #     classe) foram atualizadas pra bloquear 400 explicitamente, já que nenhuma das
    #     duas cobre essa designação (só a BS 1868 usa "Class 400")
    #   - diâmetro (Cláusula 6): DN 15 a DN 600 (NPS 1/2" a 24"), com buraco no NPS 22"
    #     (a lista da Cláusula 6 pula de 20" pra 24", igual ao buraco já visto na API 600/
    #     ISO 10434) — não cobre acima de 24" (o modelo tem RETENCAO até 42")
    #   - 9.3: "the body-to-cover connection shall be male-and-female, tongue-and-groove,
    #     or ring-joint type" — as três são junta parafusada (9.5/14 exigem stud bolts);
    #     não há bonnet soldado nem pressure-seal no escopo → vedação (junta) não pode ser
    #     Castelo Soldado nem Pressure Seal (mesma leitura já usada na ISO 10434)
    _CLASSES_BS1868 = {"150", "300", "400", "600", "900", "1500", "2500"}
    _DIAMETROS_BS1868 = {
        '1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"', '2"', '2 1/2"', '3"', '4"',
        '6"', '8"', '10"', '12"', '14"', '16"', '18"', '20"', '24"',
    }
    if tipo_valvula == "RETENCAO" and data.get("norma") == "BS 1868":
        _ext_1868 = data.get("tipo_extremidade") or ""
        if _ext_1868 and not (_ext_1868.startswith("FLANGE") or _ext_1868.startswith("BUTT-WELDING")):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma BS 1868, a extremidade deve ser Flange ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/Niple/Wafer/Lug/Gray Loc Hub)"}}, status=400)
        _classe_1868 = data.get("classe") or ""
        if _classe_1868 and _classe_1868 not in _CLASSES_BS1868:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma BS 1868, a classe deve ser 150, 300, 400, 600, 900, 1500 ou 2500 (a norma não cobre 800 nem PN)"}}, status=400)
        _diam_1868 = data.get("diametro") or ""
        if _diam_1868 and _diam_1868 not in _DIAMETROS_BS1868:
            return JsonResponse({"success": False, "errors": {"diametro": 'Para norma BS 1868, o diâmetro deve ser de 1/2" a 24" (a norma não cobre 22" nem acima de 24")'}}, status=400)
        if _junta_categoria(data) in ("CASTELO SOLDADO", "PRESSURE SEAL"):
            return JsonResponse({"success": False, "errors": {"materiais": "Para norma BS 1868, a categoria da junta deve ser Junta Espiralada ou RTJ (FJA) (9.3: body-to-cover connection é male-and-female/tongue-and-groove/ring-joint, sem opção soldada nem pressure-seal)"}}, status=400)

    # ── API 594 (Seção 1) ──────────────────────────────────────────────────
    # Check Valves: Flanged, Lug, Wafer, and Butt-welding — só Retenção oferece a norma
    # (NORMA_RETENCAO). O próprio título já é o escopo de extremidade: Flange, Lug, Wafer
    # ou Butt-Welding — sem Socket-Welding/Rosca/Niple/Gray Loc Hub. A norma descreve dois
    # tipos construtivos, agora um campo próprio (`categoria_594`, Tipo A/Tipo B):
    #   - Tipo A (curto): "wafer, lug, or double flanged; single plate or dual plate"
    #     (Seção 1/5.1.3) — sem Butt-Welding. Faixa (Seção 1): a) Classe 125/250 e
    #     b) 150/300: DN 50-1200/NPS 2-48; c) Classe 600: DN 50-1050/NPS 2-42;
    #     d) Classe 900/1500: DN 50-600/NPS 2-24; e) Classe 2500: DN 50-300/NPS 2-12
    #     (Classes 125/250 não existem no CLASSES_RETENCAO_GLOBO do modelo)
    #   - Tipo B (longo): "bolted cover swing check valves... flanged or butt-welding"
    #     (Seção 1/5.1.3) — sem Wafer/Lug. Faixa: a) Classe 150-1500: DN 50-600/NPS 2-24;
    #     b) Classe 2500: DN 50-300/NPS 2-12 (mesmo teto pra Flange ou Butt-Welding —
    #     a norma não diferencia dimensão por extremidade dentro do Tipo B)
    #   - Sem `categoria_594` preenchido (dado legado/campo ainda não selecionado):
    #     mantém o comportamento antigo — extremidade aceita as 4 opções, teto usa a
    #     coluna B (Butt-Welding) ou a mais permissiva (demais) por inferência
    #   - classe (Seção 1, comum aos dois tipos): 150/300/600/900/1500/2500 — sem 800
    #     (a norma nunca lista Class 800) nem PN. Classes 125/250 (ferro fundido/nodular)
    #     não existem no CLASSES_RETENCAO_GLOBO do modelo (mesmo buraco do Class 400 na
    #     ISO 14313/BS 1873)
    #   - diâmetro mínimo DN 50 (NPS 2) pra qualquer classe/tipo/extremidade — "Sizes:
    #     NPS 2, 2 1/2, 3, ..." não desce de NPS 2 (DN 90/NPS 3 1/2 e DN 125/NPS 5 são
    #     "non-preferred... usage is discouraged" — linguagem de recomendação, não trava;
    #     e nem existem no DIAMETROS_RETENCAO do modelo)
    #   - 5.1.14/6.3 (Tipo B apenas — "Type 'B' valves shall have a bolted cover design"):
    #     body-to-cover joint é flat face (só Classe 150)/raised face/tongue-and-groove/
    #     spigot-and-recess/ring-joint — nenhum soldado nem pressure-seal → categoria da
    #     junta não pode ser Castelo Soldado nem Pressure Seal (mesma leitura já usada na
    #     BS 1868/ISO 10434). Tipo A não tem bolted cover no escopo — regra não se aplica
    _CLASSES_594 = {"150", "300", "600", "900", "1500", "2500"}
    _TETO_594 = {
        "150": {"BW": 24, "OUTROS": 48}, "300": {"BW": 24, "OUTROS": 48},
        "600": {"BW": 24, "OUTROS": 42},
        "900": {"BW": 24, "OUTROS": 24}, "1500": {"BW": 24, "OUTROS": 24},
        "2500": {"BW": 12, "OUTROS": 12},
    }
    if tipo_valvula == "RETENCAO" and data.get("norma") == "API 594":
        _classe_594 = data.get("classe") or ""
        _cat_594 = data.get("categoria_594") or ""
        _ext_594 = (data.get("tipo_extremidade") or "").upper()
        _wl_594 = _ext_594 in ("WAFER", "LUG")
        _flange_594 = _ext_594.startswith("FLANGE")
        _bw_594 = _ext_594.startswith("BUTT-WELDING")
        if _classe_594 and _classe_594 not in _CLASSES_594:
            return JsonResponse({"success": False, "errors": {"classe": "Para norma API 594, a classe deve ser 150, 300, 600, 900, 1500 ou 2500"}}, status=400)
        if _ext_594 and not (_wl_594 or _flange_594 or _bw_594):
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma API 594, a extremidade deve ser Flange, Lug, Wafer ou Butt-Welding (a norma não cobre Socket-Welding/Rosca/Niple/Gray Loc Hub)"}}, status=400)
        if _cat_594 == "TIPO A" and _ext_594 and _bw_594:
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para API 594 Tipo A, a extremidade não pode ser Butt-Welding (5.1.3: Tipo A é wafer, lug ou duplo-flangeado — Butt-Welding só existe no Tipo B)"}}, status=400)
        if _cat_594 == "TIPO B" and _wl_594:
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para API 594 Tipo B, a extremidade deve ser Flange ou Butt-Welding (5.1.3: Tipo B é bolted cover — não cobre Wafer/Lug)"}}, status=400)
        _diam_594 = data.get("diametro")
        _dn_594 = _parse_diametro(_diam_594) if _diam_594 else None
        if _dn_594 is not None:
            if _dn_594 < 2:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma API 594, o diâmetro mínimo é 2"'}}, status=400)
            if _classe_594 in _TETO_594:
                if _cat_594 == "TIPO B":
                    _teto_594 = _TETO_594[_classe_594]["BW"]
                    _desc_594 = "Tipo B"
                elif _cat_594 == "TIPO A":
                    _teto_594 = _TETO_594[_classe_594]["OUTROS"]
                    _desc_594 = "Tipo A"
                else:
                    _teto_594 = _TETO_594[_classe_594]["BW" if _bw_594 else "OUTROS"]
                    _desc_594 = "Butt-Welding" if _bw_594 else "Flange/Lug/Wafer"
                if _dn_594 > _teto_594:
                    return JsonResponse({"success": False, "errors": {"diametro": f'Para norma API 594 com classe {_classe_594} e {_desc_594}, o diâmetro máximo é {_teto_594}"'}}, status=400)
        if _cat_594 == "TIPO B" and _junta_categoria(data) in ("CASTELO SOLDADO", "PRESSURE SEAL"):
            return JsonResponse({"success": False, "errors": {"materiais": "Para norma API 594 Tipo B, a categoria da junta deve ser Junta Espiralada ou RTJ (FJA) (5.1.14 restringe o body-to-cover joint a facings de flange — flat face/raised face/tongue-and-groove/spigot-and-recess/ring-joint — sem opção soldada nem pressure-seal)"}}, status=400)

    # ── ASME B16.34 (2.1.1) ────────────────────────────────────────────────
    # Ficam por último de propósito: valem para qualquer tipo de válvula, então as regras
    # específicas acima (que sabem o tipo/corpo/categoria) devem dar a mensagem primeiro.
    #
    # A norma designa rating por "Class" + número (150, 300, 600, 900, 1500, 2500, 4500).
    # Não cobre 400 (designação da BS 1868, que não é "Class" da B16.34), 800 (designação
    # de forjado, API 602/ISO 15761), PN (designação métrica), 125 (ferro fundido, ASME
    # B16.1) nem PMT (CWP, que é Categoria A da API 609).
    _classe_b1634 = data.get("classe") or ""
    _ext_b1634 = data.get("tipo_extremidade") or ""
    if data.get("norma") == "ASME B16.34":
        if _classe_b1634.startswith("PN") or _classe_b1634 in ("125", "400", "800", "PMT"):
            return JsonResponse({"success": False, "errors": {"classe": "Para norma ASME B16.34, a classe não pode ser 400, 800, 125, PMT nem PN. Use 150, 300, 600, 900, 1500, 2500 ou 4500"}}, status=400)

        # 2.1.1 (b): "Class 4500 applies only to welding-end valves".
        _welding_b1634 = _ext_b1634.startswith("BUTT-WELDING") or _ext_b1634 == "SOCKET-WELDING"
        if _classe_b1634 == "4500" and _ext_b1634 and not _welding_b1634:
            return JsonResponse({"success": False, "errors": {"tipo_extremidade": "Para norma ASME B16.34, a classe 4500 só existe em extremidade de solda (Butt-Welding ou Socket-Welding)"}}, status=400)

        # 2.1.1 (d): "Threaded and socket welding-end valves larger than NPS 2 1/2 are
        # beyond the scope of this Standard". Butt-welding não tem esse teto.
        if _ext_b1634.startswith("ROSCA") or _ext_b1634 == "SOCKET-WELDING":
            _diam_b1634 = data.get("diametro")
            _dn_b1634 = _parse_diametro(_diam_b1634) if _diam_b1634 else None
            if _dn_b1634 is not None and _dn_b1634 > 2.5:
                return JsonResponse({"success": False, "errors": {"diametro": 'Para norma ASME B16.34 com extremidade Rosca ou Socket-Welding, o diâmetro máximo é 2 1/2"'}}, status=400)

    return None


def _encontrar_duplicata(tipo_valvula, data, excluir_pk=None):
    """Procura válvula existente com a mesma especificação (campos + materiais +
    vedações + componentes). Retorna a válvula existente ou None.
    Duplicata é global: a mesma spec só pode existir 1x (projetos são associações à parte)."""
    campos_visiveis_dup = Valvula.CAMPOS_POR_TIPO.get(tipo_valvula, [])

    # anexo_nbr não vem do frontend (campo forçado no save); reproduzir aqui o
    # mesmo valor que o save vai gravar, senão válvula com NBR nunca bate no check
    _anexo_dup = (Valvula.ANEXO_NBR_POR_TIPO.get(tipo_valvula) or "") if data.get("nbr") else ""

    # Campos nullable no modelo (null=True) podem estar salvos como NULL ou "".
    # Usar Q para aceitar ambos quando o valor esperado é "".
    q_dup = Q(tipo_valvula=tipo_valvula)
    for campo in CAMPOS_TEXTO_DUP:
        if campo == "anexo_nbr":
            val = _anexo_dup
        else:
            val = (data.get(campo, "") or "") if campo in campos_visiveis_dup else ""
        if val == "":
            q_dup &= Q(**{campo: ""}) | Q(**{campo + "__isnull": True})
        else:
            q_dup &= Q(**{campo: val})
    for campo in CAMPOS_BOOL_VALVULA:
        val = bool(data.get(campo, False)) if campo in campos_visiveis_dup else False
        q_dup &= Q(**{campo: val})
    # QSL é campo de norma (API 6D), não de tipo — checar sempre com valor real
    _qsl_val = data.get("qsl", "") or ""
    if _qsl_val == "":
        q_dup &= Q(qsl="") | Q(qsl__isnull=True)
    else:
        q_dup &= Q(qsl=_qsl_val)

    # Função (Bloqueio/Controle) é sempre visível — compara com valor real (default Bloqueio)
    q_dup &= Q(funcao=data.get("funcao") or "BLOQUEIO")

    candidatos = Valvula.objects.filter(q_dup)
    if excluir_pk is not None:
        candidatos = candidatos.exclude(pk=excluir_pk)

    materiais_novos = sorted(
        [(m.get("tipo_material"), m.get("material")) for m in data.get("materiais", []) if m.get("tipo_material") and m.get("material")]
    )
    vedacoes_novas = sorted(
        (v.get("vedacao_junta") or v.get("vedacao_corpo_tampa"))
        for v in data.get("vedacoes", [])
        if (v.get("vedacao_junta") or v.get("vedacao_corpo_tampa"))
    )
    componentes_novos = sorted([c.get("inserto_rede") for c in data.get("componentes", []) if c.get("inserto_rede")])

    for valvula_existente in candidatos:
        # Comparar materiais
        materiais_existentes = sorted(
            ValvulaMaterial.objects.filter(valvula=valvula_existente)
            .values_list("tipo_material", "material__nome")
        )
        if materiais_existentes != materiais_novos:
            continue

        # Comparar vedação (campo único salvo em vedacao_junta; cai para
        # vedacao_corpo_tampa em dados antigos). Normaliza ambos os lados.
        vedacoes_existentes = sorted(
            (vj or vc) for vc, vj in Vedacao.objects.filter(valvula=valvula_existente)
            .values_list("vedacao_corpo_tampa", "vedacao_junta")
            if (vj or vc)
        )
        if vedacoes_existentes != vedacoes_novas:
            continue

        # Comparar componentes
        componentes_existentes = sorted(
            ComponentesInternos.objects.filter(valvula=valvula_existente)
            .values_list("inserto_rede", flat=True)
        )
        if componentes_existentes != componentes_novos:
            continue

        return valvula_existente
    return None


def _resposta_duplicata(valvula_existente):
    return JsonResponse({
        "success": False,
        "duplicata": True,
        "valvula_existente": {
            "id": valvula_existente.id_valvula,
            "codigo": valvula_existente.codigo,
            "tipo_valvula": valvula_existente.tipo_valvula,
            "tipo_label": TIPO_LABEL.get(valvula_existente.tipo_valvula, valvula_existente.tipo_valvula),
        },
        "errors": {
            "__all__": f"Já existe uma válvula idêntica cadastrada com o código {valvula_existente.codigo}."
        },
    }, status=409)


def _limpar_campos_por_tipo(valvula, tipo_valvula):
    """Zera campos não aplicáveis ao tipo selecionado e força campos derivados
    (classe_pmt, anexo_nbr)."""
    campos_visiveis = Valvula.CAMPOS_POR_TIPO.get(tipo_valvula, [])
    for field in Valvula._meta.get_fields():
        # Pular reverse relations (ForeignKey reverso), M2M e campos sem name
        if not hasattr(field, 'name') or getattr(field, 'many_to_many', False) or (hasattr(field, 'related_model') and not hasattr(field, 'column')):
            continue
        fname = field.name
        if fname in ["id_valvula", "codigo", "criado_em", "atualizado_em"]:
            continue
        if fname not in campos_visiveis and fname not in ["tipo_valvula", "funcao", "observacao"]:
            if fname in CAMPOS_BOOL_VALVULA:
                setattr(valvula, fname, False)
            elif isinstance(field, db_models.CharField):
                setattr(valvula, fname, "")
            else:
                setattr(valvula, fname, None)

    # PMT só faz sentido quando a classe é PMT
    if valvula.classe != "PMT":
        valvula.classe_pmt = ""

    # Anexo NBR só se aplica quando NBR está ativo; força valor correto por tipo
    _anexo_forca = Valvula.ANEXO_NBR_POR_TIPO.get(tipo_valvula)
    if valvula.nbr and _anexo_forca:
        valvula.anexo_nbr = _anexo_forca
    else:
        valvula.anexo_nbr = ""


def _salvar_relacionados(valvula, data, substituir=False):
    """Cria materiais, vedações e componentes internos a partir do payload.
    Com substituir=True remove os registros atuais antes (edição)."""
    if substituir:
        ValvulaMaterial.objects.filter(valvula=valvula).delete()
        Vedacao.objects.filter(valvula=valvula).delete()
        ComponentesInternos.objects.filter(valvula=valvula).delete()

    _tipos_vistos = set()
    for mat in data.get("materiais", []):
        tipo_mat = mat.get("tipo_material")
        material_nome = mat.get("material")
        if tipo_mat and material_nome and tipo_mat not in _tipos_vistos:
            _tipos_vistos.add(tipo_mat)
            try:
                material_obj = Material.objects.get(nome=material_nome)
            except Material.DoesNotExist:
                material_obj = Material.objects.create(nome=material_nome)
            ValvulaMaterial.objects.create(
                valvula=valvula,
                tipo_material=tipo_mat,
                material=material_obj,
            )

    for ved in data.get("vedacoes", []):
        ved_val = ved.get("vedacao_corpo_tampa")
        ved_junta = ved.get("vedacao_junta") or ""
        if ved_val or ved_junta:
            Vedacao.objects.create(valvula=valvula, vedacao_corpo_tampa=ved_val or "", vedacao_junta=ved_junta)

    for comp in data.get("componentes", []):
        inserto = comp.get("inserto_rede")
        if inserto:
            ComponentesInternos.objects.create(valvula=valvula, inserto_rede=inserto)


def _resposta_sucesso_valvula(valvula):
    return JsonResponse({
        "success": True,
        "valvula": {
            "id": valvula.id_valvula,
            "codigo": valvula.codigo,
            "tipo_valvula": valvula.tipo_valvula,
            "tipo_label": TIPO_LABEL.get(valvula.tipo_valvula, valvula.tipo_valvula),
            "projetos": [{"id": p.id_projeto, "nome": p.nome} for p in valvula.projetos.all()],
        },
    })


@require_POST
def valvula_criar(request):
    """Cria válvula via AJAX com geração automática de código. Disponível a qualquer usuário autenticado."""
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "errors": {"__all__": "Sem permissão"}}, status=403)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError as e:
        return JsonResponse({"success": False, "errors": {"__all__": f"Dados inválidos: {str(e)}"}}, status=400)

    tipo_valvula = data.get("tipo_valvula")
    if not tipo_valvula:
        return JsonResponse({"success": False, "errors": {"tipo_valvula": "Tipo de válvula é obrigatório"}}, status=400)

    _aplicar_regras_automaticas(tipo_valvula, data)
    erro = _validar_regras_valvula(tipo_valvula, data)
    if erro:
        return erro

    # Seção crítica: dup-check + geração de código + save serializados por tipo via
    # lock (advisory no Postgres, named lock no MySQL), dentro de uma transação. Sem
    # isso, 2+ requests concorrentes passam juntos no exists() (TOCTOU -> válvulas
    # duplicadas) e geram o mesmo código (colisão no unique -> IntegrityError/500).
    with transaction.atomic(), _lock_tipo_valvula(tipo_valvula):
        duplicata = _encontrar_duplicata(tipo_valvula, data)
        if duplicata:
            # Registra a tentativa para estatísticas
            TentativaDuplicata.objects.create(
                tipo_valvula=tipo_valvula,
                valvula_existente=duplicata,
                usuario=request.user if request.user.is_authenticated else None,
            )
            return _resposta_duplicata(duplicata)

        # Gerar código automático
        codigo = gerar_codigo(tipo_valvula)
        data["codigo"] = codigo

        form = ValvulaForm(data=data)
        if not form.is_valid():
            return JsonResponse({"success": False, "errors": form.errors}, status=400)

        valvula = form.save(commit=False)
        valvula.codigo = codigo
        _limpar_campos_por_tipo(valvula, tipo_valvula)

        # Válvula nasce sem projeto; a atribuição a projetos é feita depois (M2M).

        # Autor da criação (para estatísticas por usuário)
        valvula.criado_por = request.user if request.user.is_authenticated else None

        valvula.save()

        # Filhos dentro da transação: o dup-check compara materiais/vedações/
        # componentes, então precisam estar commitados junto com o pai
        _salvar_relacionados(valvula, data)

    return _resposta_sucesso_valvula(valvula)


@require_POST
@especial_required
def valvula_editar(request, pk):
    """Edita válvula via AJAX."""
    valvula = get_object_or_404(Valvula, pk=pk)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "errors": {"__all__": "Dados inválidos"}}, status=400)

    # Manter o código original (não muda na edição)
    data["codigo"] = valvula.codigo

    # Garantir valores booleanos quando não enviados
    for campo in CAMPOS_BOOL_VALVULA:
        if campo not in data:
            data[campo] = False

    tipo_valvula = data.get("tipo_valvula", valvula.tipo_valvula)
    bypass_validacao = bool(data.get("bypass_validacao", False))

    _aplicar_regras_automaticas(tipo_valvula, data)

    if not bypass_validacao:
        erro = _validar_regras_valvula(tipo_valvula, data)
        if erro:
            return erro

        duplicata = _encontrar_duplicata(tipo_valvula, data, excluir_pk=valvula.pk)
        if duplicata:
            return _resposta_duplicata(duplicata)

    form = ValvulaForm(data=data, instance=valvula)
    if not form.is_valid():
        return JsonResponse({"success": False, "errors": form.errors}, status=400)

    valvula = form.save(commit=False)
    _limpar_campos_por_tipo(valvula, tipo_valvula)

    # Projetos não são alterados aqui — a associação é feita pelos endpoints de atribuição (M2M).

    valvula.save()

    _salvar_relacionados(valvula, data, substituir=True)

    return _resposta_sucesso_valvula(valvula)


@require_POST
@especial_required
def valvula_excluir(request, pk):
    """Exclui válvula via AJAX."""
    valvula = get_object_or_404(Valvula, pk=pk)
    codigo = valvula.codigo
    valvula.delete()
    return JsonResponse({"success": True, "codigo": codigo})


@require_POST
@especial_required
def valvula_excluir_lote(request):
    """Exclui várias válvulas de uma vez."""
    data = json.loads(request.body)
    ids = data.get("ids", [])
    if not ids:
        return JsonResponse({"success": False, "error": "Nenhuma válvula selecionada"}, status=400)
    valvulas = Valvula.objects.filter(pk__in=ids)
    total = valvulas.count()
    codigos = list(valvulas.values_list("codigo", flat=True))
    valvulas.delete()
    return JsonResponse({"success": True, "total": total, "codigos": codigos})


# ── PDF ──────────────────────────────────────────────────────────────────────

def _calc_rate_api6d(valvula, materiais, componentes):
    """Retorna o Rate de vazamento ISO 5208 ('A', 'C', 'D', 'G') ou '' quando a
    norma não define rate (nem API 6D, ISO 14313 nem BS 1868).

    API 6D (Anexo I, tabela por tipo de válvula):
      Rate A: Esfera com inserto de sede não-metálico (PEEK/PTFE/DEVLON/…).
      Rate D: Esfera sem inserto macio (assento metálico).
      Rate C: demais válvulas (fora Esfera/Retenção) sem inserto macio.
      Rate G: Retenção sem inserto macio.

    ISO 14313 (11.4.3 + Anexo B.3.1, cita ISO 5208 direto): só distingue
    macio(A)/metálico(D) por tipo de sede — sem tabela por tipo de válvula
    como a API 6D (sem Rate C nem G no texto). Só calculável em Esfera/Retenção
    (únicos tipos com campo de inserto de sede no modelo); demais tipos ficam
    sem rate (sem como saber se a sede é macia ou metálica).

    BS 1868 (só Retenção oferece a norma): decisão de negócio (2026-07-21) —
    limitada a Rate A/C, sem G (a norma não usa a tabela por tipo da API 6D):
    com inserto de sede macio → Rate A; sem inserto → Rate C.

    Verifica duas fontes: ValvulaMaterial.tipo_material=='INSERTO_SEDE'
    e ComponentesInternos.inserto_rede — qualquer uma basta para Rate A.
    """
    if valvula.norma not in ("API 6D", "ISO 14313", "BS 1868"):
        return ""
    _inserto_mat = next(
        (m.material.nome for m in materiais if m.tipo_material == "INSERTO_SEDE"),
        "",
    )
    _inserto_comp = next((c.inserto_rede for c in componentes if c.inserto_rede), "")
    _tem_inserto = (
        (_inserto_mat and _inserto_mat != "N/A") or
        (_inserto_comp and _inserto_comp != "N/A")
    )
    if valvula.norma == "ISO 14313":
        if valvula.tipo_valvula in ("ESFERA", "RETENCAO"):
            return "A" if _tem_inserto else "D"
        return ""
    if valvula.norma == "BS 1868":
        return "A" if _tem_inserto else "C"
    # API 6D daqui pra baixo
    if valvula.tipo_valvula == "RETENCAO" and not _tem_inserto:
        return "G"
    if valvula.tipo_valvula == "ESFERA" and _tem_inserto:
        return "A"
    if valvula.tipo_valvula == "ESFERA" and not _tem_inserto:
        return "D"
    if valvula.tipo_valvula != "RETENCAO" and not _tem_inserto:
        return "C"
    return ""


@require_POST
def valvula_preview(request):
    """Renderiza o HTML da folha de dados (mesmo template do PDF) a partir dos
    dados do formulário, SEM salvar nada. Usado pelo botão de pré-visualização
    no cadastro. Retorna HTML pronto para exibir em iframe."""
    if not request.user.is_authenticated:
        return HttpResponse("Sem permissão", status=403)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponse("Dados inválidos", status=400)

    tipo_valvula = data.get("tipo_valvula")
    if not tipo_valvula:
        return HttpResponse("Selecione o tipo de válvula", status=400)

    # Retenção → junção corpo/castelo sempre Aparafusado (espelha valvula_criar)
    if tipo_valvula == "RETENCAO":
        data["juncao_corpo_castelo"] = "APARAFUSADO"

    # Acionamento manual → sem posição de falha (N/A) e sem dados elétricos (espelha valvula_criar)
    if data.get("tipo_acionamento") in ACIONAMENTOS_MANUAIS:
        data["posicao_falha"] = "N/A"
        data["marca_atuador"] = "PADRÃO FABRICANTE"
        data["flange_acoplamento"] = ""
        data["tensao"] = ""
        data["fase"] = ""
        data["frequencia"] = ""
        data["hot_disconnect"] = False

    # Acabamento da face do flange só existe em conexão Flange/Wafer/Lug (espelha valvula_criar)
    if "tipo_extremidade" in Valvula.CAMPOS_POR_TIPO.get(tipo_valvula, []):
        _ext = (data.get("tipo_extremidade") or "").upper()
        if _ext and not (_ext.startswith("FLANGE") or _ext in ("WAFER", "LUG")):
            data["tipo_ranhura"] = "N/A"

    # Regra: norma de conexão (campo "Flange") derivada do tipo de extremidade.
    # Butt-Welding → ASME B16.25; Socket-Welding → ASME B16.11; Rosca NPT →
    # ASME B16.20; Niple → ASME B36.10 (corpo carbono) ou B36.19 (corpo inox
    # austenítico) — mesmos conjuntos de corpo usados nas regras de parafuso/
    # porca por corpo NBR (CORPOS_NBR_B7 / CORPOS_NBR_INOX). Flange/Wafer/Lug/
    # Gray Loc Hub não têm norma derivada — campo continua livre.
    _FLANGE_CORPOS_CARBONO = ("ASTM A105", "ASTM A105N", "ASTM A181", "ASTM A216 WCB", "ASTM A216 GR WCB")
    _FLANGE_CORPOS_INOX = ("ASTM A182 F304", "ASTM A351 CF8", "ASTM A182 F316", "ASTM A351 CF8M",
                            "ASTM A182 F317", "ASTM A351 CG8M", "ASTM A182 F347", "ASTM A351 CF8C")
    _ext_flange = (data.get("tipo_extremidade") or "").upper()
    if _ext_flange.startswith("BUTT-WELDING"):
        data["flange"] = "ASME B16.25"
    elif _ext_flange == "SOCKET-WELDING":
        data["flange"] = "ASME B16.11"
    elif _ext_flange == "ROSCA NPT":
        data["flange"] = "ASME B16.20"
    elif _ext_flange.startswith("NIPLE"):
        _materiais_flange = data.get("materiais", [])
        _corpo_flange = next((m.get("material") for m in _materiais_flange if m.get("tipo_material") == "CORPO_TAMPA"), None)
        if _corpo_bate_base(_corpo_flange, _FLANGE_CORPOS_CARBONO):
            data["flange"] = "ASME B36.10"
        elif _corpo_bate_base(_corpo_flange, _FLANGE_CORPOS_INOX):
            data["flange"] = "ASME B36.19"

    # Esfera + diâmetro ≤2" → construção da sede trava em DBB (espelha valvula_criar)
    if tipo_valvula == "ESFERA":
        _diam_dib = data.get("diametro") or ""
        if _diam_dib and _parse_diametro(_diam_dib) <= 2:
            data["dib"] = "DBB"

    # Pintura (já é a norma) define cor e condição de pintura (espelha valvula_criar)
    _pintura_val = data.get("pintura") or ""
    data["norma_pintura"] = ""
    if _pintura_val == "PADRÃO FABRICANTE":
        data["cor"] = "PADRÃO DO FABRICANTE"
        data["condicao_pintura"] = "PADRÃO FABRICANTE"
    elif _pintura_val == "SEM PINTURA":
        data["cor"] = "N/A"
        data["condicao_pintura"] = "N/A"

    # Pintura do atuador (espelha valvula_criar)
    _pintura_atuador_val = data.get("pintura_atuador") or ""
    data["norma_pintura_atuador"] = ""
    if _pintura_atuador_val == "PADRÃO FABRICANTE":
        data["cor_atuador"] = "PADRÃO DO FABRICANTE"
        data["condicao_pintura_atuador"] = "PADRÃO FABRICANTE"
    elif _pintura_atuador_val == "SEM PINTURA":
        data["cor_atuador"] = "N/A"
        data["condicao_pintura_atuador"] = "N/A"

    # Função = Bloqueio → sem instrumentação + característica On - Off (espelha valvula_criar)
    if data.get("funcao") != "CONTROLE":
        for _campo in CAMPOS_INSTRUMENTACAO:
            data[_campo] = False if _campo == "indicador_posicao" else ""
        data["caracteristicas"] = "On - Off"
    else:
        for _campo in ("valvula_solenoide", "chave_fim_curso", "sensor_posicao"):
            if _campo in Valvula.CAMPOS_POR_TIPO.get(tipo_valvula, []):
                data[_campo] = "SIM"

    campos_texto = [
        "fabricante", "pintura", "cor", "norma_pintura", "condicao_pintura", "norma", "iogp", "qsl", "diametro", "classe", "classe_pmt", "tipo_extremidade", "tipo_ranhura",
        "tipo_passagem", "tipo_acionamento", "marca_atuador", "flange_acoplamento", "tipo_montagem", "construcao_corpo",
        "pintura_atuador", "cor_atuador", "norma_pintura_atuador", "condicao_pintura_atuador",
        "tipo_castelo", "juncao_corpo_castelo", "dib", "uso_geral",
        "tipo_retencao", "configuracao_corpo_retencao", "orientacao_instalacao", "categoria_594",
        "certificacao_sil", "nace", "revestimento", "categoria_borboleta", "face_a_face",
        "configuracao_disco", "posicionador", "ip", "ip_posicionador", "ip_solenoide", "ip_chave_fim_curso", "ip_sensor_posicao", "filtro", "tubing",
        "chave_fim_curso", "valvula_solenoide", "valvula_lock_up", "sensor_posicao", "valvula_escape_rapido",
        "caracteristicas", "dreno", "vent", "alivio_externo", "placa_identificacao", "flange", "anexo_nbr",
        "posicao_falha", "tensao", "fase", "frequencia",
    ] + CAMPOS_CE + CAMPOS_ELET
    campos_bool = ["nbr", "valvula_alivio", "dispositivo_antiestatico", "baixa_emissao_fugitiva", "indicador_posicao", "dreno", "vent", "alivio_externo", "hot_disconnect", "contra_peso"]

    # Válvula em memória (não salva) só para alimentar o template
    valvula = Valvula(tipo_valvula=tipo_valvula)
    valvula.funcao = data.get("funcao") or "BLOQUEIO"
    valvula.codigo = "PRÉVIA"
    campos_visiveis = Valvula.CAMPOS_POR_TIPO.get(tipo_valvula, [])
    for campo in campos_texto:
        setattr(valvula, campo, (data.get(campo, "") or "") if campo in campos_visiveis else "")
    for campo in campos_bool:
        setattr(valvula, campo, bool(data.get(campo, False)) if campo in campos_visiveis else False)
    if valvula.classe != "PMT":
        valvula.classe_pmt = ""

    # Anexo NBR só se aplica quando NBR está ativo; força valor correto por tipo
    # (espelha valvula_criar — não depende do front mandar o valor certo).
    _anexo_forca = Valvula.ANEXO_NBR_POR_TIPO.get(tipo_valvula)
    valvula.anexo_nbr = _anexo_forca if (valvula.nbr and _anexo_forca) else ""

    valvula.observacao = (data.get("observacao") or "")[:1000]

    # Materiais/vedações/componentes em memória
    materiais = [
        ValvulaMaterial(tipo_material=m.get("tipo_material"), material=Material(nome=m.get("material")))
        for m in data.get("materiais", []) if m.get("tipo_material") and m.get("material")
    ]
    vedacoes = [
        Vedacao(vedacao_corpo_tampa=v.get("vedacao_corpo_tampa") or "", vedacao_junta=v.get("vedacao_junta") or "")
        for v in data.get("vedacoes", []) if v.get("vedacao_corpo_tampa") or v.get("vedacao_junta")
    ]
    componentes = [
        ComponentesInternos(inserto_rede=c.get("inserto_rede"))
        for c in data.get("componentes", []) if c.get("inserto_rede")
    ]

    from django.template.loader import render_to_string
    LABELS = _folha_labels_bi()
    rate_api6d = _calc_rate_api6d(valvula, materiais, componentes)
    folha_grupos = _build_folha_grupos(valvula, materiais, vedacoes, componentes, rate_api6d, LABELS)
    folha_notas = _build_folha_notas(valvula, LABELS)
    folha_grupos, folha_notas = _numerar_folha(folha_grupos, folha_notas)
    # Notas vira mais um grupo com rótulo lateral vertical, igual Corpo/Atuador/Instrumentação.
    if folha_notas:
        folha_grupos = folha_grupos + [(LABELS["lbl_cat_notas"], folha_notas)]
    folha_grupos_ctx = _folha_grupos_ctx(folha_grupos, LABELS)
    context_base = {
        "valvula": valvula,
        "materiais": materiais,
        "vedacoes": vedacoes,
        "componentes": componentes,
        "rate_api6d": rate_api6d,
        "folha_grupos": folha_grupos_ctx,
        "folha_notas": folha_notas,
        "observacao": valvula.observacao or "",
        # No navegador o logo carrega via URL estática (não caminho de disco)
        "logo_path": settings.STATIC_URL + "assets/logo.png",
        **LABELS,
    }
    # Auto-fit medido: escolhe a maior fonte que cabe em 1 página (mesma lógica do PDF).
    escala, _ = _folha_autofit(context_base)
    html_string = render_to_string("core/valvula_pdf.html", {**context_base, **escala})
    return HttpResponse(html_string)


# Tradução de tipo de válvula PT→EN (usa a chave salva no banco).
_TIPO_VALVULA_EN = {
    "ESFERA": "Ball",
    "GAVETA": "Gate",
    "GLOBO": "Globe",
    "RETENCAO": "Check",
    "BORBOLETA": "Butterfly",
    "GLOBO_CONTROLE": "Control Globe",
}

# Tradução do tipo de material PT→EN (chave do ValvulaMaterial.TIPO_MATERIAL).
_TIPO_MATERIAL_EN = {
    "CORPO_TAMPA": "Body / Cover",
    "OBTURADOR": "Closure Member",
    "SEDE": "Seat / Seat Holder",
    "INSERTO_SEDE": "Seat Insert",
    "HASTE": "Stem",
    "MOLAS": "Springs",
    "JUNTA": "Joint Category",
    "MATERIAL_JUNTA": "Gasket Material",
    "GAXETA": "Packing",
    "PARAFUSOS": "Bolts",
    "PORCAS": "Nuts",
}

# Tradução de valores descritivos PT→EN para a folha de dados em inglês.
# Chave = valor armazenado no banco (uppercase). Valores técnicos (ASTM, AISI,
# ISO, normas, classes, diâmetros) não traduzem — ausência no dict = mantém original.
_VALOR_PT_EN = {
    # Pintura
    "PADRÃO FABRICANTE": "Manufacturer Standard",
    "ESPECIAL": "Special",
    "SEM PINTURA": "Unpainted",
    "Zona Submersa": "Submerged Zone",
    "Zona de Transição": "Splash Zone",
    "Zona Atmosférica": "Atmospheric Zone",
    # Passagem
    "PLENA": "Full Bore",
    "REDUZIDA": "Reduced Bore",
    # Montagem
    "TRUNNION": "Trunnion",
    "FLUTUANTE": "Floating",
    # Construção do corpo
    "BI-PARTIDO": "Two-piece",
    "TRI-PARTIDO": "Three-piece",
    # Castelo
    "NORMAL": "Standard",
    "EXTENDIDO": "Extended",
    # Junção corpo/castelo
    "APARAFUSADO": "Bolted",
    "SOLDADO": "Welded",
    "ROSCADO": "Threaded",
    "ROSCADO E SOLDADO": "Threaded and Welded",
    # Categoria borboleta
    "CATEGORIA A": "Category A",
    "CATEGORIA B": "Category B",
    # Categoria API 594 (Retenção) — Tipo A/Tipo B
    "TIPO A": "Type A",
    "TIPO B": "Type B",
    # Retenção: tipo construtivo / configuração do corpo / orientação de instalação
    # (mesmas chaves ESFERA/DISCO de outros dicionários, mas aqui é o tipo construtivo
    # do obturador da Retenção, não o tipo de válvula)
    "PISTAO": "Piston",
    "ESFERA": "Ball",
    "DISCO": "Disc",
    "ANGULAR": "Angle",
    "RETO": "Straight",
    "HORIZONTAL": "Horizontal",
    "VERTICAL": "Vertical",
    # Face a face
    "LUG": "Lug",
    "WAFER": "Wafer",
    "FLANGEADA PADRÃO CURTO": "Flanged Short Pattern",
    "FLANGEADA PADRÃO LONGO": "Flanged Long Pattern",
    # Configuração do disco
    "CONCÊNTRICA": "Concentric",
    "BI-EXCÊNTRICA": "Double Offset",
    "TRI-EXCÊNTRICA": "Triple Offset",
    # Extremidade
    "FLANGE FACE PLANA": "Flat Face Flange",
    "FLANGE RF": "RF Flange",
    "FLANGE RTJ (FJA)": "FJA Flange (RTJ)",
    "BUTT-WELDING": "Butt-Welding",
    "SOCKET-WELDING": "Socket-Welding",
    "ROSCA NPT": "NPT Threaded",
    "ROSCA BSP": "BSP Threaded",
    'NIPLE 4" COMP. SCH 40': 'Nipple 4" Length SCH 40',
    'NIPLE 4" COMP. SCH 80': 'Nipple 4" Length SCH 80',
    'NIPLE 4" COMP. SCH 160': 'Nipple 4" Length SCH 160',
    'NIPLE 4" COMP. SCH XS': 'Nipple 4" Length SCH XS',
    'NIPLE 4" COMP. SCH XXS': 'Nipple 4" Length SCH XXS',
    # Ranhura
    "125-250 μin ESPIRAL": "125-250 μin Spiral",
    "125-250 μin CONCÊNTRICA": "125-250 μin Concentric",
    "LISO (125 μin)": "Smooth (125 μin)",
    # Acionamento
    "ALAVANCA": "Lever",
    "VOLANTE": "Handwheel",
    "VOLANTE COM ENGRENAGEM DE REDUÇÃO": "Handwheel with Reduction Gear",
    "ATUADOR ELÉTRICO": "Electric Actuator",
    "ATUADOR ELÉTRICO COM VOLANTE": "Electric Actuator with Handwheel",
    "ATUADOR PNEUMÁTICO RETORNO POR MOLA TIPO PISTÃO": "Spring-Return Pneumatic Actuator, Piston Type",
    "ATUADOR PNEUMÁTICO RETORNO POR MOLA TIPO DIAFRAGMA": "Spring-Return Pneumatic Actuator, Diaphragm Type",
    "ATUADOR PNEUMÁTICO RETORNO POR MOLA COM VOLANTE TIPO PISTÃO": "Spring-Return Pneumatic Actuator with Handwheel, Piston Type",
    "ATUADOR PNEUMÁTICO RETORNO POR MOLA COM VOLANTE TIPO DIAFRAGMA": "Spring-Return Pneumatic Actuator with Handwheel, Diaphragm Type",
    "ATUADOR PNEUMÁTICO DUPLA AÇÃO TIPO PISTÃO": "Double-Acting Pneumatic Actuator, Piston Type",
    "ATUADOR PNEUMÁTICO DUPLA AÇÃO TIPO DIAFRAGMA": "Double-Acting Pneumatic Actuator, Diaphragm Type",
    "ATUADOR ELETROHIDRÁULICO RETORNO POR MOLA": "Spring-Return Electro-Hydraulic Actuator",
    "ATUADOR ELETROHIDRÁULICO RETORNO POR MOLA COM VOLANTE": "Spring-Return Electro-Hydraulic Actuator with Handwheel",
    # Posição em caso de falha (display PT → EN)
    "Aberto": "Open",
    "Fechada": "Closed",
    "Última Posição": "Last Position",
    # Uso geral
    "USO GERAL": "General Use",
    # Revestimento
    "DACROMETIZAÇÃO": "Dacromet Coating",
    "ZINCO NÍQUEL": "Zinc-Nickel",
    # Posicionador
    "ELETROPNEUMÁTICO 4–20 mA": "Electro-Pneumatic 4–20 mA",
    "PROTOCOLO HART": "HART Protocol",
    "FIELDBUS": "Fieldbus",
    "PROFIBUS": "Profibus",
    # Instrumentação Sim/Não (Solenoide, Chave Fim de Curso, Sensor Posição, Lock-Up, Escape Rápido)
    "SIM": "Yes",
    "NÃO": "No",
    # Válvula Lock-Up / Escape Rápido
    "2 VIAS": "2-Way",
    "3 VIAS": "3-Way",
    # Filtro
    "SIM, COM MANÔMETROS": "Yes, with Gauges",
    "SIM, SEM MANÔMETROS": "Yes, without Gauges",
    "SIM, COM MAN. ELEMENTO FILTRANTE 5 μ": "Yes, with Gauges, 5 μ Filter Element",
    "SIM, COM MAN. ELEMENTO FILTRANTE 5 μ CORPO ALUMÍNIO": "Yes, with Gauges, 5 μ Filter Element, Aluminum Body",
    "SIM, COM MAN. ELEMENTO FILTRANTE 5 μ CORPO AÇO INOX": "Yes, with Gauges, 5 μ Filter Element, Stainless Steel Body",
    # Vedação corpo/tampa
    "JUNTA ESPIRALADA": "Spiral Wound Gasket",
    "O'RING VITON": "Viton O-Ring",
    "PRESSURE SEAL": "Pressure Seal",
    "CASTELO SOLDADO": "Welded Bonnet",
    # Tubing
    "INOX 316": "SS 316",
    "INOX 304": "SS 304",
    "COBRE": "Copper",
    "COBRE REVESTIDO": "Coated Copper",
    # Gaxeta / junta descritivas
    "GRAFITE": "Graphite",
    "GRAFITE FLEXÍVEL + FIO DE INCONEL": "Flexible Graphite + Inconel Wire",
    "GRAFITE FLEXÍVEL + FIO DE INCONEL C/ INIBIDOR DE CORROSÃO": "Flexible Graphite + Inconel Wire w/ Corrosion Inhibitor",
    "GRAFITE FLEXÍVEL + FIO DE INCONEL C/ INIBIDOR (molibdato de bário e/ou fios de zinco)": "Flexible Graphite + Inconel Wire w/ Inhibitor (barium molybdate and/or zinc wire)",
    "AISI 304 + GRAFITE FLEXÍVEL": "AISI 304 + Flexible Graphite",
    "AISI 316 + GRAFITE FLEXÍVEL": "AISI 316 + Flexible Graphite",
    # Junta (novos itens) — "PADRÃO FABRICANTE" já traduzido acima (Pintura)
    "S32750 + GRAFITE": "S32750 + Graphite",
    "GRAFITE + LIP SEAL": "Graphite + Lip Seal",
    "HNBR + GRAFITE": "HNBR + Graphite",
    "SELADO À PRESSÃO": "Pressure Sealed",
    "NÃO APLICÁVEL": "Not Applicable",
    "AISI 347 + GRAFITE": "AISI 347 + Graphite",
    "AISI 317 + GRAFITE": "AISI 317 + Graphite",
    "AISI 321 + GRAFITE": "AISI 321 + Graphite",
    "AISI 321H + GRAFITE": "AISI 321H + Graphite",
    "S31803 + GRAFITE": "S31803 + Graphite",
    "PAPELÃO HIDRÁULICO C/ BORRACHA SBR": "Hydraulic Cardboard w/ SBR Rubber",
    "AISI 304 + GRAFITE": "AISI 304 + Graphite",
    "AISI 316 + GRAFITE": "AISI 316 + Graphite",
    "AISI 304L + GRAFITE": "AISI 304L + Graphite",
    "AISI 316L + GRAFITE": "AISI 316L + Graphite",
    "VITON + GRAFITE": "Viton + Graphite",
    # Materiais renomeados de "SS<n>" para "AISI <n>" (mesmo aço, grafia antiga
    # mantida entre parênteses no inglês em vez do fallback genérico "SS <n>"
    # de _t_valor_bi, que inseriria espaço e não bateria com a grafia original).
    "AISI 304 + ENP": "SS304 + ENP",
    "AISI 304 + Stellite": "SS304 + Stellite",
    "AISI 304 + carbeto de tungstênio": "SS304 + carbeto de tungstênio",
    "AISI 304 + cromo duro": "SS304 + cromo duro",
    "AISI 304H": "SS304H",
    "AISI 316 + ENP": "SS316 + ENP",
    "AISI 316 + Ni60": "SS316 + Ni60",
    "AISI 316 + PTFE": "SS316 + PTFE",
    "AISI 316 + Stellite": "SS316 + Stellite",
    "AISI 316 + carbeto de tungstênio": "SS316 + carbeto de tungstênio",
    "AISI 316 + cromo duro": "SS316 + cromo duro",
    "AISI 316 SF": "SS316 SF",
    "AISI 316L + PTFE": "SS316L + PTFE",
    "AISI 317": "SS317",
    "AISI 317 + Stellite": "SS317 + Stellite",
    "AISI 321H": "SS321H",
    "AISI 321H + Stellite": "SS321H + Stellite",
    "AISI 410 + CP": "SS410 + CP",
    "AISI 410 + Carbeto de Cromo": "SS410 + Carbeto de Cromo",
    "AISI 410 + ENP": "SS410 + ENP",
    "AISI 410 + INCONEL (UNS N06625)": "SS410 + INCONEL (UNS N06625)",
    "AISI 410 + NYLON": "SS410 + NYLON",
    "AISI 410 + PEEK": "SS410 + PEEK",
    "AISI 410 + RPTFE (25% C)": "SS410 + RPTFE (25% C)",
    "AISI 410 + SF": "SS410 + SF",
    "AISI 410 + Stellite": "SS410 + Stellite",
    "AISI 410 + carbeto de tungstênio": "SS410 + carbeto de tungstênio",
    "AISI 410 + cromo duro": "SS410 + cromo duro",
    "AISI 416": "SS416",
    "AISI 420": "SS420",
    "AISI 430": "SS430",
    "AISI 431": "SS431",
}


def _t_valor(valor, is_en):
    """Traduz um valor descritivo PT→EN. Mantém o original se não houver tradução
    (valores técnicos como ASTM/AISI/ISO) ou se não for inglês."""
    if not is_en or not valor:
        return valor
    return _VALOR_PT_EN.get(valor, valor)


def _bi(pt, en):
    """Combina PT + EN em 'PT (EN)'. Se EN vazio ou igual ao PT, retorna só PT."""
    if en and en != pt:
        return f"{pt} ({en})"
    return pt


def _folha_labels_bi():
    """Rótulos bilíngues 'PT (EN)' para o documento único. Fonte única PDF/Excel/preview."""
    pt = _folha_labels(False)
    en = _folha_labels(True)
    labels = {k: _bi(pt[k], en[k]) for k in pt}
    labels["lbl_item"] = "ITEM (ITEM)"
    return labels


# Renomeações só de exibição nos documentos (PDF/Excel/preview) — o valor no
# banco/formulário não muda. Chave = valor armazenado, valor = como imprimir.
_VALOR_DISPLAY_DOC = {
    "RTJ (FJA)": "FJA (RTJ)",
    "FLANGE RTJ (FJA)": "FLANGE FJA (RTJ)",
}


# Cores: prefixo do nome PT → EN para a folha de dados bilíngue.
# O número e o código (RAL/Munsell/SINAL) não mudam entre idiomas.
_COR_PREFIX_EN = {
    "PADRÃO DO FABRICANTE": "MANUFACTURER STANDARD",
    "AMARELO SEGURANÇA": "SAFETY YELLOW",
    "AMARELO PETROBRAS": "PETROBRAS YELLOW",
    "AMARELO CATERPILLAR": "CATERPILLAR YELLOW",
    "AMARELO": "YELLOW",
    "AREIA": "SAND",
    "AZUL PASTEL": "PASTEL BLUE",
    "AZUL-PASTEL": "PASTEL BLUE",
    "AZUL SEGURANÇA": "SAFETY BLUE",
    "AZUL-SEGURANÇA": "SAFETY BLUE",
    "AZUL PETROBRAS": "PETROBRAS BLUE",
    "AZUL FRANÇA": "FRENCH BLUE",
    "AZUL DEL REY": "ROYAL BLUE",
    "AZUL": "BLUE",
    "BRANCO-GELO": "ICE WHITE",
    "BRANCO": "WHITE",
    "CARAMELO": "CARAMEL",
    "CASTANHO": "CHESTNUT",
    "COR-DE-ALUMÍNIO": "ALUMINUM COLOR",
    "ALUMÍNIO": "ALUMINUM",
    "VINHO": "WINE",
    "ÓXIDO DE FERRO": "IRON OXIDE",
    "VIOLETA": "VIOLET",
    "AZUL-MARINHO": "NAVY BLUE",
    "BORDÔ": "BURGUNDY",
    "ROSA-SECO": "DUSTY ROSE",
    "TURQUESA": "TURQUOISE",
    "LILÁS": "LILAC",
    "CINZA GELO": "ICE GRAY",
    "CINZA-GELO": "ICE GRAY",
    "CINZA CLARO": "LIGHT GRAY",
    "CINZA-CLARO": "LIGHT GRAY",
    "CINZA MÉDIO": "MEDIUM GRAY",
    "CINZA-MÉDIO": "MEDIUM GRAY",
    "CINZA ESCURO": "DARK GRAY",
    "CINZA-ESCURO": "DARK GRAY",
    "CINZA COSTADO": "HULL GRAY",
    "CINZA": "GRAY",
    "CREME CLARO": "LIGHT CREAM",
    "CREME-CLARO": "LIGHT CREAM",
    "CREME CANALIZAÇÕES": "PIPING CREAM",
    "CREME": "CREAM",
    "LARANJA SEGURANÇA": "SAFETY ORANGE",
    "LARANJA-SEGURANÇA": "SAFETY ORANGE",
    "LARANJA": "ORANGE",
    "MARFIM": "IVORY",
    "MARROM CANALIZAÇÕES": "PIPING BROWN",
    "MARROM": "BROWN",
    "PÉROLA": "PEARL",
    "PRETO": "BLACK",
    "PÚRPURA SEGURANÇA": "SAFETY PURPLE",
    "PÚRPURA": "PURPLE",
    "SÂNDALO": "SANDALWOOD",
    "VERDE PASTEL": "PASTEL GREEN",
    "VERDE-PASTEL": "PASTEL GREEN",
    "VERDE SEGURANÇA": "SAFETY GREEN",
    "VERDE-SEGURANÇA": "SAFETY GREEN",
    "VERDE PETROBRAS": "PETROBRAS GREEN",
    "VERDE-PETROBRAS": "PETROBRAS GREEN",
    "VERDE MÁQUINA": "MACHINE GREEN",
    "VERDE EMBLEMA": "EMBLEM GREEN",
    "VERDE-EMBLEMA": "EMBLEM GREEN",
    "VERDE": "GREEN",
    "VERMELHO SEGURANÇA": "SAFETY RED",
    "VERMELHO-SEGURANÇA": "SAFETY RED",
    "VERMELHO BOMBEIRO": "FIRE RED",
    "VERMELHO ÓXIDO": "OXIDE RED",
    "VERMELHO": "RED",
}


def _t_cor_bi(cor):
    """Cor bilíngue 'PT (EN)': traduz o prefixo do nome e mantém o restante
    (número e código RAL/Munsell). 'VERDE OTAN / NATO GREEN' já é bilíngue."""
    if not cor or cor == "VERDE OTAN / NATO GREEN":
        return cor
    for prefixo in sorted(_COR_PREFIX_EN, key=len, reverse=True):
        if cor.startswith(prefixo):
            return _bi(cor, _COR_PREFIX_EN[prefixo] + cor[len(prefixo):])
    return cor


def _t_valor_bi(valor):
    """Valor descritivo bilíngue 'PT (EN)'. Sem tradução no dicionário: técnicos
    (ASTM/ISO) ficam iguais; 'AISI <n>' vira 'SS <n>' em inglês."""
    if not valor:
        return valor
    en = _VALOR_PT_EN.get(valor)
    if not en:
        import re
        alt = re.sub(r'\bAISI\b', 'SS', valor, flags=re.IGNORECASE)
        if alt != valor:
            en = alt
    valor = _VALOR_DISPLAY_DOC.get(valor, valor)
    return _bi(valor, en) if en else valor


def _folha_labels(is_en):
    """Rótulos da folha de dados PT/EN. Fonte única usada pelo PDF, Excel e preview."""
    return {
        "title_valve":            "Valve" if is_en else "Valvula",
        "title_sheet":            "Valve Data Sheet" if is_en else "Ficha de Cadastro de Valvula",
        "lbl_main_data":          "Main Data" if is_en else "Dados Principais",
        "lbl_valve_type":         "Valve Type" if is_en else "Tipo de Valvula",
        "lbl_funcao":             "Function" if is_en else "Função",
        "lbl_standard":           "Construction Standard" if is_en else "Norma de Construção",
        "lbl_iogp":               "IOGP" if is_en else "IOGP",
        "lbl_diameter":           "Diameter" if is_en else "Diametro",
        "lbl_class":              "Pressure Class" if is_en else "Classe de Pressão",
        "lbl_end_type":           "Connection" if is_en else "Conexão",
        "lbl_groove":             "Flange Face Finish" if is_en else "Acabamento da Face do Flange",
        "lbl_bore":               "Bore" if is_en else "Passagem",
        "lbl_actuation":          "Valve Actuation" if is_en else "Acionamento da Válvula",
        "lbl_marca_atuador":      "Actuator Brand" if is_en else "Marca do Atuador",
        "lbl_mounting":           "Mounting" if is_en else "Montagem",
        "lbl_body_construction":  "Body Construction" if is_en else "Construcao Corpo",
        "lbl_bonnet":             "Bonnet" if is_en else "Castelo",
        "lbl_body_bonnet_joint":  "Body/Bonnet Joint" if is_en else "Juncao Corpo/Castelo",
        "lbl_category":           "Category" if is_en else "Categoria",
        "lbl_face_to_face":       "Face to Face" if is_en else "Tipo de Conexão",
        "lbl_disc_config":        "Disc Config." if is_en else "Config. Disco",
        "lbl_retencao_tipo":      "Check Type" if is_en else "Tipo",
        "lbl_retencao_config":    "Body Configuration" if is_en else "Configuração do Corpo",
        "lbl_retencao_orientacao": "Installation Orientation" if is_en else "Orientação de Instalação",
        "lbl_positioner":         "Power Supply Signal" if is_en else "Sinal de Alimentação",
        "lbl_ip":                 "Protection Rating" if is_en else "Grau de Proteção",
        "lbl_filter":             "Filter" if is_en else "Filtro",
        "lbl_tubing":             "Tubing" if is_en else "Tubing",
        "lbl_chave_fim_curso":    "Limit Switch" if is_en else "Chave Fim de Curso",
        "lbl_valvula_solenoide":  "Solenoid Valve" if is_en else "Válvula Solenoide",
        "lbl_valvula_lock_up":    "Lock-Up Valve" if is_en else "Válvula Lock-Up",
        "lbl_sensor_posicao":     "Position Sensor" if is_en else "Sensor de Posição",
        "lbl_valvula_escape_rapido": "Quick Exhaust Valve" if is_en else "Válvula de Escape Rápido",
        "lbl_features":           "Features" if is_en else "Caracteristicas",
        "lbl_relief_valve":       "Relief Valve" if is_en else "Valvula de Alivio",
        "lbl_antistatic":         "Antistatic Device" if is_en else "Dispositivo Antiestático",
        "lbl_low_emission":       "Low Fugitive Emission" if is_en else "Baixa Emissao Fugitiva",
        "lbl_position_indicator": "Position Indicator" if is_en else "Indicador de Posicao",
        "lbl_hot_disconnect":     "Hot Disconnect" if is_en else "Hot Disconnect",
        "lbl_general_use":        "Fire-Tested" if is_en else "Testada a Fogo",
        "lbl_sil_cert":           "SIL Certification" if is_en else "Certificacao SIL",
        "lbl_coating":            "Bolt and Nut Coating" if is_en else "Revestimento Parafuso e Porca",
        "lbl_materials":          "Materials" if is_en else "Materiais",
        "lbl_material_type":      "Material Type" if is_en else "Tipo de Material",
        "lbl_material":           "Material" if is_en else "Material",
        "lbl_sealing":            "Sealing" if is_en else "Vedacao",
        "lbl_body_cover_seal":    "Seat/Cover Seal" if is_en else "Vedação Sede/Tampa",
        "lbl_junta_seal":         "Gasket Seal" if is_en else "Vedação da Junta",
        "lbl_junta_material_categoria": "Joint Material/Category" if is_en else "Material/Categoria da Junta",
        "lbl_internal_components":"Internal Components" if is_en else "Componentes Internos",
        "lbl_seat_insert":        "Seat Insert" if is_en else "Inserto da Sede",
        "lbl_yes":                "Yes" if is_en else "Sim",
        "lbl_no":                 "No" if is_en else "Não",
        "lbl_generated_at":       "Document generated on" if is_en else "Documento gerado em",
        "lbl_system_name":        "Valve Management System" if is_en else "Sistema de Gestao de Valvulas",
        # Cabeçalhos e rótulos adicionais
        "lbl_data_sheet":         "DATA SHEET" if is_en else "FOLHA DE DADOS",
        "lbl_item":               "ITEM" if is_en else "ITEM",
        "lbl_description":        "DESCRIPTION" if is_en else "DESCRIÇÃO",
        "lbl_acceptance_criteria":"Acceptance Criteria" if is_en else "Critério de Aceitação",
        "lbl_qsl":                "QSL" if is_en else "QSL",
        "lbl_id_plate":           "Identification Plate" if is_en else "Placa de Identificação",
        "lbl_rate":               "Rate" if is_en else "Rate",
        "lbl_nbr":                "ABNT NBR 15827" if is_en else "ABNT NBR 15827",
        "lbl_pmt":                "PMT" if is_en else "PMT",
        "lbl_dib":                "Seat Construction" if is_en else "Construção da Sede",
        "lbl_nace":               "NACE" if is_en else "NACE",
        # Seção de notas
        "lbl_notes":              "NOTES" if is_en else "NOTAS",
        "lbl_observation":        "Remarks" if is_en else "Observação",
        "lbl_corpo_tampa":        "Body and Cover" if is_en else "Corpo e Tampa",
        "lbl_fabricante":         "Manufacturer" if is_en else "Fabricante",
        "lbl_pintura":            "Painting" if is_en else "Pintura",
        "lbl_norma_pintura":      "Painting Standard" if is_en else "Norma de Pintura",
        "lbl_condicao_pintura":   "Painting Condition" if is_en else "Condição de Pintura",
        "lbl_color":              "Color" if is_en else "Cor",
        "lbl_pintura_atuador":    "Actuator Painting" if is_en else "Pintura do Atuador",
        "lbl_norma_pintura_atuador": "Actuator Painting Standard" if is_en else "Norma de Pintura do Atuador",
        "lbl_condicao_pintura_atuador": "Actuator Painting Condition" if is_en else "Condição de Pintura do Atuador",
        "lbl_color_atuador":      "Actuator Color" if is_en else "Cor do Atuador",
        "lbl_caracteristicas":    "Characteristics" if is_en else "Características",
        "lbl_dreno":              "Drain" if is_en else "Dreno",
        "lbl_vent":               "Vent" if is_en else "Vent",
        "lbl_alivio_externo":     "External Relief" if is_en else "Alívio Externo",
        "lbl_contra_peso":        "Counterweight" if is_en else "Contrapeso",
        "lbl_placa_identificacao":"Identification Plate" if is_en else "Placa de Identificação",
        "lbl_flange":             "Flange" if is_en else "Flange",
        "lbl_flange_acoplamento": "ISO 5211 Mounting Flange" if is_en else "Flange de Acoplamento ISO 5211",
        "lbl_posicao_falha":      "Fail Position" if is_en else "Posição em Caso de Falha",
        "lbl_tensao":             "Voltage" if is_en else "Tensão",
        "lbl_fase":               "Phase" if is_en else "Fase",
        "lbl_frequencia":         "Frequency" if is_en else "Frequência",
        # Categorias (coluna vertical de agrupamento) — PT na 1ª linha, EN embaixo
        "lbl_cat_corpo":          "Corpo e Internos\n(Body and Internals)",
        "lbl_cat_atuador":        "Atuador\n(Actuator)",
        "lbl_cat_instrumentacao": "Instrum.\n(Instrum.)",
        "lbl_cat_posicionador":   "Posic\n(Posit)",
        "lbl_cat_solenoide":      "Sol.\n(Sol.)",
        "lbl_cat_chave_fim_curso":"Ch. FC\n(Lim. Sw)",
        "lbl_cat_sensor_posicao": "Sens. Pos\n(Pos. Sens)",
        "lbl_ce_generic":         "Electrical Char." if is_en else "Caract. Elétricas",
        "lbl_el_generic":         "Power Supply" if is_en else "Alimentação Elétrica",
        "lbl_ip_posicionador":    "IP Positioner" if is_en else "IP Posicionador",
        "lbl_ip_solenoide":       "IP Solenoid" if is_en else "IP Solenoid",
        "lbl_ip_chave_fim_curso": "IP Limit Switch" if is_en else "IP Chave Fim de Curso",
        "lbl_ip_sensor_posicao":  "IP Position Sensor" if is_en else "IP Sensor de Posição",
        "lbl_ce_posicionador":    "Electrical Char. Positioner" if is_en else "Caract. Elétricas Posicionador",
        "lbl_ce_solenoide":       "Electrical Char. Solenoid" if is_en else "Caract. Elétricas Solenoid",
        "lbl_ce_chave_fim_curso": "Electrical Char. Limit Switch" if is_en else "Caract. Elétricas Chave Fim de Curso",
        "lbl_ce_sensor_posicao":  "Electrical Char. Position Sensor" if is_en else "Caract. Elétricas Sensor de Posição",
        "lbl_el_posicionador":    "Power Supply Positioner" if is_en else "Elétrica Posicionador",
        "lbl_el_solenoide":       "Power Supply Solenoid" if is_en else "Elétrica Solenoid",
        "lbl_el_chave_fim_curso": "Power Supply Limit Switch" if is_en else "Elétrica Chave Fim de Curso",
        "lbl_el_sensor_posicao":  "Power Supply Position Sensor" if is_en else "Elétrica Sensor de Posição",
        "lbl_cat_notas":          "Notas\n(Notes)",
    }


def _vertical_label_image(text, target_width_pt=20.0):
    """<img> com o texto virado 90° p/ coluna estreita do PDF. xhtml2pdf não
    suporta writing-mode/transform em CSS — só sabe desenhar imagem, então o
    rótulo é desenhado com PIL e rotacionado antes de virar PNG embutido.
    Aceita múltiplas linhas separadas por '\\n' (ex.: PT + tradução EN entre
    parênteses), todas no mesmo tamanho/negrito, empilhadas antes da rotação."""
    from PIL import Image, ImageDraw, ImageFont
    from django.utils.safestring import mark_safe
    import base64
    import reportlab

    font_path = os.path.join(os.path.dirname(reportlab.__file__), "fonts", "VeraBd.ttf")
    linhas = text.split("\n")
    font = ImageFont.truetype(font_path, 40)

    probe = ImageDraw.Draw(Image.new("RGBA", (1, 1)))
    # Altura fixa da fonte (ascent+descent), não o bbox do texto — textos sem
    # descendente (ex: "Atuador") têm bbox mais baixo que "Corpo e Internos"
    # (tem "p"), o que fazia a escala final variar e o texto parecer maior/menor.
    ascent, descent = font.getmetrics()
    lh = ascent + descent
    metricas = []
    for linha in linhas:
        l, t, r, b = probe.textbbox((0, 0), linha, font=font)
        metricas.append((linha, r - l, l))

    gap = 4
    pad = 6
    w = max(m[1] for m in metricas)
    h = lh * len(metricas) + gap * (len(metricas) - 1)

    img = Image.new("RGBA", (w + pad * 2, h + pad * 2), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)
    y = pad
    for linha, lw, lx in metricas:
        draw.text((pad - lx + (w - lw) // 2, y), linha, font=font, fill=(0, 0, 0, 255))
        y += lh + gap

    img = img.rotate(90, expand=True)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    uri = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")
    height_pt = target_width_pt * (img.height / img.width)
    return mark_safe(
        f'<img src="{uri}" style="width:{target_width_pt:.1f}pt;height:{height_pt:.1f}pt">'
    )


def _folha_grupos_ctx(folha_grupos, L):
    """Prepara os grupos para o template: (categoria, imagem_vertical, itens, is_header).
    Todos os grupos usam rótulo lateral vertical (mesmo layout de 'Corpo e Internos'
    e 'Atuador'), incluindo as subcategorias de instrumentação (Posicionador,
    Solenoide, Chave, Sensor) — rótulos curtos agora cabem na coluna vertical."""
    ctx = []
    for cat, itens in folha_grupos:
        ctx.append((cat, _vertical_label_image(cat), itens, False))
    return ctx


_DN_NBR14788 = {
    '1/2"': "15", '3/4"': "20", '1"': "25", '1 1/4"': "32", '1 1/2"': "40",
    '2"': "50", '2 1/2"': "65", '3"': "80", '4"': "100", '6"': "150",
    '8"': "200", '10"': "250", '12"': "300", '14"': "350", '16"': "400",
    '18"': "450", '20"': "500",
}


def _formatar_diametro(valvula):
    """NBR 14788 designa o diâmetro por DN, não NPS — exibe 'DN (NPS)' na folha."""
    if valvula.norma == "NBR 14788":
        dn = _DN_NBR14788.get(valvula.diametro)
        if dn:
            return f"{dn} ({valvula.diametro})"
    return valvula.diametro


def _build_folha_grupos(valvula, materiais, vedacoes, componentes, rate_api6d, L):
    """Monta os grupos da folha: 'Corpo e Internos' e 'Atuador'. Retorna lista de
    (categoria, [(item, valor), ...]) na ordem definida. SIL e demais campos de
    nota vão para NOTAS. Documento único bilíngue. Fonte única PDF/Excel."""
    campos_visiveis = Valvula.CAMPOS_POR_TIPO.get(valvula.tipo_valvula, [])
    sim, nao = L["lbl_yes"], L["lbl_no"]
    mat_by_tipo = {m.tipo_material: (m.material.nome if m.material else "") for m in materiais}
    tipo_mat_display = dict(ValvulaMaterial.TIPO_MATERIAL)

    def mat_label(tipo):
        if tipo == "CORPO_TAMPA":
            return L["lbl_corpo_tampa"]
        if tipo == "INSERTO_SEDE":
            return L["lbl_seat_insert"]
        return _bi(tipo_mat_display.get(tipo, tipo), _TIPO_MATERIAL_EN.get(tipo))

    corpo = []
    # Ordem definida: Corpo e Internos
    corpo.append((L["lbl_valve_type"], _bi(valvula.get_tipo_valvula_display(), _TIPO_VALVULA_EN.get(valvula.tipo_valvula))))
    corpo.append((L["lbl_funcao"], _bi(valvula.get_funcao_display(), "Control" if valvula.funcao == "CONTROLE" else "On / Off")))
    if valvula.fabricante:
        corpo.append((L["lbl_fabricante"], valvula.fabricante))
    if valvula.tipo_montagem:
        # Construção da Sede (DIB) entra junto da Montagem: "Trunnion - DIB-1".
        _montagem_val = _t_valor_bi(valvula.tipo_montagem)
        if valvula.dib and valvula.dib != "N/A":
            _montagem_val = f"{_montagem_val} - {valvula.dib}"
        corpo.append((L["lbl_mounting"], _montagem_val))
    if "caracteristicas" in campos_visiveis:
        corpo.append((L["lbl_caracteristicas"], valvula.caracteristicas or "N/A"))
    if valvula.diametro:
        corpo.append((L["lbl_diameter"], _formatar_diametro(valvula)))
    if valvula.classe:
        corpo.append((L["lbl_class"], valvula.classe))
    if "flange" in campos_visiveis:
        corpo.append((L["lbl_flange"], valvula.flange or "N/A"))
    if valvula.tipo_extremidade:
        corpo.append((L["lbl_end_type"], _t_valor_bi(valvula.tipo_extremidade)))
    if valvula.face_a_face:
        corpo.append((L["lbl_face_to_face"], _t_valor_bi(valvula.face_a_face)))
    if valvula.tipo_ranhura:
        corpo.append((L["lbl_groove"], _t_valor_bi(valvula.tipo_ranhura)))
    # Norma de Construção + IOGP + QSL na mesma célula, separados por " - "
    # (ex.: "API 6D - IOGP AS-562 - QSL3/3G").
    if valvula.norma:
        _norma_partes = [valvula.norma]
        if "iogp" in campos_visiveis and valvula.iogp and valvula.iogp != "N/A":
            _norma_partes.append(valvula.iogp)
        if valvula.qsl and valvula.qsl != "N/A":
            _norma_partes.append(valvula.qsl)
        corpo.append((L["lbl_standard"], " - ".join(_norma_partes)))
    if "flange_acoplamento" in campos_visiveis and valvula.flange_acoplamento:
        corpo.append((L["lbl_flange_acoplamento"], valvula.flange_acoplamento))
    if valvula.tipo_passagem:
        corpo.append((L["lbl_bore"], _t_valor_bi(valvula.tipo_passagem)))
    if valvula.construcao_corpo:
        corpo.append((L["lbl_body_construction"], _t_valor_bi(valvula.construcao_corpo)))
    if valvula.norma in ("API 6D", "ISO 14313", "BS 1868"):  # critério de aceitação
        criterio = f"ISO 5208 — Rate {rate_api6d}" if rate_api6d else "ISO 5208"
        corpo.append((L["lbl_acceptance_criteria"], criterio))
    # Vedação sede/tampa — entra logo abaixo do "Inserto da Sede". Só Esfera.
    _ved_val = ""
    if valvula.tipo_valvula == "ESFERA":
        for v in vedacoes:
            _ved_val = getattr(v, "vedacao_junta", "") or v.vedacao_corpo_tampa
            if _ved_val:
                break
    # Materiais na ordem pedida
    for tipo in ["CORPO_TAMPA", "OBTURADOR", "HASTE", "SEDE", "INSERTO_SEDE",
                 "MOLAS", "GAXETA", "PARAFUSOS", "PORCAS"]:
        nome = mat_by_tipo.get(tipo)
        if nome:
            lbl = mat_label(tipo)
            item = lbl if tipo == "INSERTO_SEDE" else f"{L['lbl_material']} {lbl}"
            corpo.append((item, _t_valor_bi(nome)))
        # Vedação logo após o Inserto da Sede (independe de o inserto ter material)
        if tipo == "INSERTO_SEDE" and _ved_val:
            corpo.append((L["lbl_body_cover_seal"], _t_valor_bi(_ved_val)))
        # Categoria da Junta (JUNTA) + Material da Junta (MATERIAL_JUNTA) na mesma
        # linha, logo após Molas — mesma posição que a JUNTA ocupava antes da
        # separação em 2 tipo_material (entre Molas e Gaxeta).
        if tipo == "MOLAS":
            _junta_categoria = mat_by_tipo.get("JUNTA")
            _junta_material = mat_by_tipo.get("MATERIAL_JUNTA")
            if _junta_categoria or _junta_material:
                _junta_valor = " / ".join(_t_valor_bi(v) for v in (_junta_material, _junta_categoria) if v)
                corpo.append((L["lbl_junta_material_categoria"], _junta_valor))
    # Revestimento (parafuso e porca) logo abaixo do material Porca.
    if valvula.revestimento:
        corpo.append((L["lbl_coating"], _t_valor_bi(valvula.revestimento)))
    if "dreno" in campos_visiveis:
        corpo.append((L["lbl_dreno"], sim if valvula.dreno else nao))
    if "vent" in campos_visiveis:
        corpo.append((L["lbl_vent"], sim if valvula.vent else nao))
    if "alivio_externo" in campos_visiveis:
        corpo.append((L["lbl_alivio_externo"], sim if valvula.alivio_externo else nao))
    if "placa_identificacao" in campos_visiveis:
        corpo.append((L["lbl_placa_identificacao"], valvula.placa_identificacao or "N/A"))
    if valvula.pintura:
        corpo.append((L["lbl_pintura"], _t_valor_bi(valvula.pintura)))
    if valvula.condicao_pintura:
        corpo.append((L["lbl_condicao_pintura"], _t_valor_bi(valvula.condicao_pintura)))
    if valvula.cor:
        corpo.append((L["lbl_color"], _t_cor_bi(valvula.cor)))
    # Não-listados → também no Corpo e Internos
    if valvula.classe_pmt:
        corpo.append((L["lbl_pmt"], valvula.classe_pmt))
    if valvula.tipo_castelo:
        corpo.append((L["lbl_bonnet"], _t_valor_bi(valvula.tipo_castelo)))
    if valvula.juncao_corpo_castelo:
        corpo.append((L["lbl_body_bonnet_joint"], _t_valor_bi(valvula.juncao_corpo_castelo)))
    if valvula.tipo_retencao:
        corpo.append((L["lbl_retencao_tipo"], _t_valor_bi(valvula.tipo_retencao)))
    if valvula.configuracao_corpo_retencao:
        corpo.append((L["lbl_retencao_config"], _t_valor_bi(valvula.configuracao_corpo_retencao)))
    if valvula.orientacao_instalacao:
        corpo.append((L["lbl_retencao_orientacao"], _t_valor_bi(valvula.orientacao_instalacao)))
    if valvula.categoria_594:
        corpo.append((L["lbl_category"], _t_valor_bi(valvula.categoria_594)))
    if valvula.categoria_borboleta:
        corpo.append((L["lbl_category"], _t_valor_bi(valvula.categoria_borboleta)))
    if valvula.configuracao_disco:
        corpo.append((L["lbl_disc_config"], _t_valor_bi(valvula.configuracao_disco)))

    # Instrumentação — 4 subcategorias (item + seu IP + Caract. Elétricas) + gerais.
    # Só preenchida em Função = Controle (Bloqueio zera esses campos).
    def _ce_valor(suf):
        partes = [getattr(valvula, f"{p}_{suf}") or "" for p in ("ex", "protecao", "grupo", "temp", "epl")]
        return " ".join(x for x in partes if x)

    def _ce_tem_dado_real(suf):
        # "Ex" sozinho não conta: o campo vem sempre travado em "Ex" (única opção
        # do select) e não indica por si só que o dispositivo existe/foi usado.
        return any(getattr(valvula, f"{p}_{suf}") for p in ("protecao", "grupo", "temp", "epl"))

    def _elet_valor(suf):
        partes = [getattr(valvula, f"{p}_{suf}") or "" for p in ("tensao", "corrente", "potencia")]
        return " / ".join(x for x in partes if x)

    # Cada subcategoria vira uma categoria própria na folha (item principal, IP,
    # Caract. Elétricas e Alimentação Elétrica). (campo_principal, suf, rótulo_item, categoria)
    _subcats = [
        (valvula.posicionador, "posicionador", L["lbl_positioner"], L["lbl_cat_posicionador"]),
        (valvula.valvula_solenoide, "solenoide", L["lbl_valvula_solenoide"], L["lbl_cat_solenoide"]),
        (valvula.chave_fim_curso, "chave_fim_curso", L["lbl_chave_fim_curso"], L["lbl_cat_chave_fim_curso"]),
        (valvula.sensor_posicao, "sensor_posicao", L["lbl_sensor_posicao"], L["lbl_cat_sensor_posicao"]),
    ]
    # Posicionador só existe em Função = Controle. Solenoide/Chave Fim de Curso/
    # Sensor de Posição são opcionais em qualquer Função (Bloqueio inclusive).
    _subcats_ativos = _subcats if valvula.funcao == "CONTROLE" else [s for s in _subcats if s[1] != "posicionador"]
    grupos_subcat = []
    for _principal, _suf, _lbl_item, _lbl_cat in _subcats_ativos:
        _linhas = []
        if valvula.funcao == "CONTROLE":
            # Item sempre aparece (valor "N/A" se dispositivo não selecionado), igual aos
            # demais campos da folha (Flange, Placa de Identificação etc).
            _valor_item = _t_valor_bi(_principal) if _principal and _principal != "N/A" else "N/A"
            _linhas.append((_lbl_item, _valor_item))
        elif _principal and _principal != "N/A":
            _linhas.append((_lbl_item, _t_valor_bi(_principal)))
        _ip = getattr(valvula, f"ip_{_suf}")
        if _ip:
            # Prefixa "IP - " ao número (ex.: "24" → "IP - 24"); mantém "N/A" como está.
            _ip_val = _ip if _ip == "N/A" else f"IP - {_ip}"
            _linhas.append((L["lbl_ip"], _ip_val))
        if _ce_tem_dado_real(_suf):
            _linhas.append((L["lbl_ce_generic"], _ce_valor(_suf)))
        if _elet_valor(_suf):
            _linhas.append((L["lbl_el_generic"], _elet_valor(_suf)))
        if _linhas:
            grupos_subcat.append((_lbl_cat, _linhas))

    # Instrumentação geral (fora das subcategorias)
    instrumentacao = []
    if valvula.filtro:
        instrumentacao.append((L["lbl_filter"], _t_valor_bi(valvula.filtro)))
    if valvula.tubing:
        instrumentacao.append((L["lbl_tubing"], _t_valor_bi(valvula.tubing)))
    if valvula.valvula_lock_up and valvula.valvula_lock_up != "N/A":
        instrumentacao.append((L["lbl_valvula_lock_up"], _t_valor_bi(valvula.valvula_lock_up)))
    if valvula.valvula_escape_rapido and valvula.valvula_escape_rapido != "N/A":
        instrumentacao.append((L["lbl_valvula_escape_rapido"], _t_valor_bi(valvula.valvula_escape_rapido)))

    # Atuador
    atuador = []
    if valvula.tipo_acionamento:
        atuador.append((L["lbl_actuation"], _t_valor_bi(valvula.tipo_acionamento)))
    if "marca_atuador" in campos_visiveis and valvula.marca_atuador:
        atuador.append((L["lbl_marca_atuador"], _t_valor_bi(valvula.marca_atuador)))
    if valvula.pintura_atuador:
        atuador.append((L["lbl_pintura_atuador"], _t_valor_bi(valvula.pintura_atuador)))
    if valvula.condicao_pintura_atuador:
        atuador.append((L["lbl_condicao_pintura_atuador"], _t_valor_bi(valvula.condicao_pintura_atuador)))
    if valvula.cor_atuador:
        atuador.append((L["lbl_color_atuador"], _t_cor_bi(valvula.cor_atuador)))
    if "tensao" in campos_visiveis and valvula.tensao:
        atuador.append((L["lbl_tensao"], f"{valvula.tensao} V"))
    if "fase" in campos_visiveis and valvula.fase:
        atuador.append((L["lbl_fase"], valvula.fase))
    if "frequencia" in campos_visiveis and valvula.frequencia:
        atuador.append((L["lbl_frequencia"], f"{valvula.frequencia} Hz"))
    if "posicao_falha" in campos_visiveis and valvula.posicao_falha:
        atuador.append((L["lbl_posicao_falha"], _t_valor_bi(valvula.get_posicao_falha_display())))

    # Ordem: Corpo e Internos, Atuador, Posicionador, Solenoid, Chave Fim de Curso,
    # Sensor de Posição e, por fim, Instrumentação geral (filtro/tubing/lock-up/escape).
    grupos = [(L["lbl_cat_corpo"], corpo)]
    if atuador:
        grupos.append((L["lbl_cat_atuador"], atuador))
    grupos.extend(grupos_subcat)  # Posicionador, Solenoid, Chave Fim de Curso, Sensor
    if instrumentacao:
        grupos.append((L["lbl_cat_instrumentacao"], instrumentacao))
    return grupos


def _build_folha_notas(valvula, L):
    """Monta as linhas (rótulo, valor) da seção de NOTAS: QSL, construção da sede
    (DIB), dispositivo antiestático, válvula de alívio, baixa emissão, indicador
    de posição, contrapeso e uso geral. Documento único bilíngue. Bool sempre
    presente (Sim/Não); texto só se preenchido."""
    sim = L["lbl_yes"]
    nao = L["lbl_no"]
    campos_visiveis = Valvula.CAMPOS_POR_TIPO.get(valvula.tipo_valvula, [])
    notas = []

    if valvula.certificacao_sil:  # SIL desceu para as Notas
        notas.append((L["lbl_sil_cert"], valvula.certificacao_sil))
    if valvula.nace:  # NACE desceu para as Notas
        notas.append((L["lbl_nace"], valvula.nace))
    # Construção da Sede (DIB) agora aparece junto da Montagem em Corpo e Internos.
    # Só cai aqui como fallback se não houver Montagem para acoplá-la.
    if valvula.dib and not valvula.tipo_montagem:
        notas.append((L["lbl_dib"], valvula.dib))
    if "dispositivo_antiestatico" in campos_visiveis:
        notas.append((L["lbl_antistatic"], sim if valvula.dispositivo_antiestatico else nao))
    if "valvula_alivio" in campos_visiveis:
        notas.append((L["lbl_relief_valve"], sim if valvula.valvula_alivio else nao))
    if "baixa_emissao_fugitiva" in campos_visiveis:
        notas.append((L["lbl_low_emission"], f"{_bi('Sim', 'Yes')} - ISO 15848-1" if valvula.baixa_emissao_fugitiva else nao))
    if "indicador_posicao" in campos_visiveis:
        notas.append((L["lbl_position_indicator"], sim if valvula.indicador_posicao else nao))
    if "hot_disconnect" in campos_visiveis:
        notas.append((L["lbl_hot_disconnect"], sim if valvula.hot_disconnect else nao))
    if "contra_peso" in campos_visiveis:
        notas.append((L["lbl_contra_peso"], sim if valvula.contra_peso else nao))
    # Testada a Fogo: uma norma real (≠ "USO GERAL"/"N/A") → "Sim, <norma>"; senão "Não".
    if "uso_geral" in campos_visiveis:
        if valvula.uso_geral and valvula.uso_geral not in ("USO GERAL", "N/A"):
            valor_fogo = _bi(f"Sim, {valvula.uso_geral}", f"Yes, {valvula.uso_geral}")
        else:
            valor_fogo = _bi("Não", "No")
        notas.append((L["lbl_general_use"], valor_fogo))
    # NBR + Anexo numa linha só (ex.: "Sim, Anexo C (Yes, Annex C)").
    if valvula.nbr and valvula.anexo_nbr:
        valor_nbr = _bi(f"Sim, {valvula.anexo_nbr}", f"Yes, {valvula.anexo_nbr.replace('Anexo', 'Annex')}")
    else:
        valor_nbr = _bi("Sim", "Yes") if valvula.nbr else _bi("Não", "No")
    notas.append((L["lbl_nbr"], valor_nbr))

    return notas


def _numerar_folha(grupos, notas):
    """Aplica Nº sequencial contínuo entre os grupos e as notas. Garante que PDF
    e Excel usem a mesma numeração. Retorna (grupos_num, notas_num) onde cada item
    vira (num, label, valor)."""
    n = 0
    grupos_num = []
    for cat, itens in grupos:
        rows = []
        for label, valor in itens:
            n += 1
            rows.append((n, label, valor))
        grupos_num.append((cat, rows))
    notas_num = []
    for label, valor in notas:
        n += 1
        notas_num.append((n, label, valor))
    return grupos_num, notas_num


# Escalas da folha, da MAIOR fonte p/ a MENOR. Auto-fit escolhe a maior que
# renderiza em 1 página. Cada item: (fonte, padding, line-height, título, cabeçalho).
_FOLHA_ESCALAS = [
    {"folha_fs": "14.5pt", "folha_pad": "5px 6px",   "folha_lh": "1.25", "folha_titulo_fs": "16.5pt", "folha_cab_fs": "14.5pt"},
    {"folha_fs": "13.5pt", "folha_pad": "5px 6px",   "folha_lh": "1.2",  "folha_titulo_fs": "15.5pt", "folha_cab_fs": "13.5pt"},
    {"folha_fs": "12.5pt", "folha_pad": "4px 6px",   "folha_lh": "1.2",  "folha_titulo_fs": "14.5pt", "folha_cab_fs": "13pt"},
    {"folha_fs": "11.5pt", "folha_pad": "4px 5px",   "folha_lh": "1.18", "folha_titulo_fs": "13.5pt", "folha_cab_fs": "12pt"},
    {"folha_fs": "10.5pt", "folha_pad": "3px 5px",   "folha_lh": "1.15", "folha_titulo_fs": "12.5pt", "folha_cab_fs": "11pt"},
    {"folha_fs": "9.5pt",  "folha_pad": "3px 4px",   "folha_lh": "1.1",  "folha_titulo_fs": "11.5pt", "folha_cab_fs": "10pt"},
    {"folha_fs": "8.5pt",  "folha_pad": "2px 4px",   "folha_lh": "1.05", "folha_titulo_fs": "10pt",  "folha_cab_fs": "9pt"},
    {"folha_fs": "7.5pt",  "folha_pad": "1.5px 4px", "folha_lh": "1.0",  "folha_titulo_fs": "9pt",   "folha_cab_fs": "8pt"},
    {"folha_fs": "7pt",    "folha_pad": "1px 3px",   "folha_lh": "1.0",  "folha_titulo_fs": "8.5pt", "folha_cab_fs": "7.5pt"},
    {"folha_fs": "6.5pt",  "folha_pad": "1px 3px",   "folha_lh": "0.98", "folha_titulo_fs": "8pt",   "folha_cab_fs": "7pt"},
    {"folha_fs": "6pt",    "folha_pad": "0px 3px",   "folha_lh": "0.95", "folha_titulo_fs": "7.5pt", "folha_cab_fs": "6.5pt"},
    {"folha_fs": "5.5pt",  "folha_pad": "0px 2px",   "folha_lh": "0.95", "folha_titulo_fs": "7pt",   "folha_cab_fs": "6pt"},
    {"folha_fs": "5pt",    "folha_pad": "0px 2px",   "folha_lh": "0.92", "folha_titulo_fs": "6.5pt", "folha_cab_fs": "5.5pt"},
    {"folha_fs": "4.5pt",  "folha_pad": "0px 2px",   "folha_lh": "0.9",  "folha_titulo_fs": "6pt",   "folha_cab_fs": "5pt"},
]


def _folha_num_paginas(pdf_bytes):
    """Conta páginas de um PDF em memória. 999 se ilegível (força usar menor fonte)."""
    import io
    try:
        from pypdf import PdfReader
        return len(PdfReader(io.BytesIO(pdf_bytes)).pages)
    except Exception:
        return 999


def _folha_render_pdf(context):
    """Renderiza o template da folha em PDF (bytes) via xhtml2pdf. (bytes, ok).

    Único ponto por onde passa todo render, então é aqui que o código universal é
    quebrado em linhas: a quebra depende de folha_cab_fs, que só existe depois de
    a escala ser escolhida (ver _folha_autofit)."""
    import io
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    context = {
        **context,
        "codigo_universal_linhas": _cu_linhas_pdf(
            context.get("codigo_universal", ""), context.get("folha_cab_fs", "8.5pt")
        ),
    }
    buf = io.BytesIO()
    try:
        status = pisa.CreatePDF(render_to_string("core/valvula_pdf.html", context), dest=buf)
        return buf.getvalue(), (not status.err)
    except Exception:
        return b"", False


def _folha_autofit(context_base):
    """Escolhe, por MEDIÇÃO, a maior escala de fonte que cabe em 1 página A4.
    Busca binária sobre _FOLHA_ESCALAS (maior→menor). Independe do tamanho da
    folha: cresce em folha curta, encolhe em folha cheia. Garante 1 página.
    Retorna (escala_dict, pdf_bytes_dessa_escala). pdf_bytes pode ser reaproveitado."""
    lo, hi = 0, len(_FOLHA_ESCALAS) - 1
    melhor_idx, melhor_bytes = None, None
    while lo <= hi:
        mid = (lo + hi) // 2
        pdf_bytes, ok = _folha_render_pdf({**context_base, **_FOLHA_ESCALAS[mid]})
        if ok and _folha_num_paginas(pdf_bytes) <= 1:
            melhor_idx, melhor_bytes = mid, pdf_bytes  # coube: tenta fonte maior (idx menor)
            hi = mid - 1
        else:
            lo = mid + 1  # não coube (ou erro): fonte menor (idx maior)
    if melhor_bytes is None:
        # Nem a menor fonte coube (folha gigante improvável) — usa a menor mesmo assim.
        esc = _FOLHA_ESCALAS[-1]
        melhor_bytes, _ = _folha_render_pdf({**context_base, **esc})
        return esc, melhor_bytes
    return _FOLHA_ESCALAS[melhor_idx], melhor_bytes


def _valvula_fire_safe(valvula):
    """Fire-safe / testada a fogo: uso_geral preenchido com uma norma de fogo
    (qualquer valor ≠ 'USO GERAL'/'N/A'/vazio)."""
    ug = (valvula.uso_geral or "").strip().upper()
    return ug not in ("", "USO GERAL", "N/A")


def _codigo_universal_aaa(valvula):
    """Parte AAA (3 letras) do código universal
    AAAB.CCCD.EEE.FFF.GGHHII.JKLLMM(NNN-NNN) — derivada do tipo/subtipo.
    (Demais partes serão adicionadas depois.)"""
    tipo = valvula.tipo_valvula
    if tipo == "GAVETA":
        return "GVS"
    if tipo == "GLOBO":
        return "GLB"
    if tipo == "GLOBO_CONTROLE":
        return "ESC"  # esfera segmentada p/ controle (tipo V)
    if tipo == "RETENCAO":
        return "REI"  # todos os subtipos usam REI por enquanto
    if tipo == "BORBOLETA":
        cfg = valvula.configuracao_disco
        if cfg == "CONCÊNTRICA":
            return "BOC"
        if cfg == "TRI-EXCÊNTRICA":
            return "BOT"
        if cfg == "BI-EXCÊNTRICA":
            return "BOF" if _valvula_fire_safe(valvula) else "BOB"
        return "BOB"  # fallback borboleta
    if tipo == "ESFERA":
        tri = valvula.construcao_corpo == "TRI-PARTIDO"  # corpo
        if valvula.tipo_montagem == "FLUTUANTE":
            return "FL3" if tri else "FL2"
        if valvula.tipo_montagem == "TRUNNION":
            dib = valvula.dib  # construção da sede
            if dib == "DIB-1":
                return "TR2" if tri else "T2B"
            if dib == "DIB-2":
                return "TR3" if tri else "TR4"
            return "TR1"  # DBB (bi; tri não existe na prática → TR1)
        return "FL2"  # fallback esfera
    return "XXX"


# Fundido (cast) x Forjado (forged) pela spec ASTM/ASME do material do corpo.
_CORPO_FORJADO_SPECS = {"A105", "A182", "A350", "B564", "B865"}
_CORPO_FUNDIDO_SPECS = {"A216", "A217", "A351", "A352", "A395", "A487", "A536", "A995"}


def _corpo_material_nome(materiais):
    """Nome do material CORPO_TAMPA (string original) ou '' se não houver."""
    for vm in materiais:
        if vm.tipo_material == "CORPO_TAMPA":
            return vm.material.nome or ""
    return ""


def _corpo_fundido_ou_forjado(materiais):
    """'fundido' / 'forjado' / None — derivado da spec ASTM (A216, A105...) do
    material CORPO_TAMPA. None se não houver material ou spec desconhecida."""
    import re
    nome = _corpo_material_nome(materiais).upper()
    m = re.search(r"\b([AB]\d{3})[A-Z]?\b", nome)  # sufixo opcional (ex.: A105N)
    spec = m.group(1) if m else None
    if spec in _CORPO_FORJADO_SPECS:
        return "forjado"
    if spec in _CORPO_FUNDIDO_SPECS:
        return "fundido"
    return None


# (corpo, juncao_corpo_castelo) → letra B. Combos não listados caem em 'X'.
_CODIGO_B_MAPA = {
    ("fundido", "APARAFUSADO"): "C",
    ("forjado", "APARAFUSADO"): "F",
    ("fundido", "SOLDADO"): "W",
    ("forjado", "SOLDADO"): "S",
    ("forjado", "ROSCADO"): "R",
    ("fundido", "ROSCADO"): "T",
    ("fundido", "ROSCADO E SOLDADO"): "U",
    ("forjado", "ROSCADO E SOLDADO"): "V",
}


def _codigo_universal_b(valvula, materiais):
    """Parte B (1 letra): corpo (fundido/forjado) × junção corpo/castelo.
    'X' (não aplicável) se sem junção, material indeterminado ou combo não previsto
    (ex.: Esfera/Borboleta não têm junção; 'Roscado e Soldado')."""
    corpo = _corpo_fundido_ou_forjado(materiais)
    juncao = valvula.juncao_corpo_castelo
    if not corpo or not juncao:
        return "X"
    return _CODIGO_B_MAPA.get((corpo, juncao), "X")


# Diâmetro (polegadas) → parte C (CCC). Até 38" = mm nominal 3 dígitos; ≥40" = "NNP".
# Chaves na forma normalizada (ponto→espaço). Diâmetros fora da tabela não entram no código.
_CODIGO_C_DIAMETRO = {
    '1/4"': "010", '1/2"': "015", '3/4"': "020", '1"': "025",
    '1 1/4"': "030", '1 1/2"': "040", '2"': "050", '2 1/2"': "065",
    '3"': "080", '4"': "100", '5"': "125", '6"': "150", '7 1/16"': "179",
    '8"': "200", '10"': "250", '12"': "300", '14"': "350", '16"': "400",
    '18"': "450", '20"': "500", '22"': "550", '24"': "600", '26"': "650",
    '28"': "700", '30"': "750", '32"': "800", '34"': "850", '36"': "900",
    '38"': "950", '40"': "40P", '42"': "42P", '44"': "44P", '46"': "46P",
    '48"': "48P", '50"': "50P", '52"': "52P", '54"': "54P", '56"': "56P",
    '58"': "58P", '60"': "60P", '62"': "62P",
}


def _norm_diametro(d):
    """Normaliza o diâmetro p/ lookup: ponto→espaço, colapsa espaços (ex.:
    '1.1/4\"' e '1 1/4\"' → '1 1/4\"')."""
    import re
    return re.sub(r"\s+", " ", (d or "").replace(".", " ")).strip()


def _codigo_universal_c(valvula):
    """Parte C (CCC): código do diâmetro. '' se o diâmetro não estiver na tabela
    (esses não entram no código)."""
    return _CODIGO_C_DIAMETRO.get(_norm_diametro(valvula.diametro), "")


# Classe de pressão → parte D (1 char). Chaves normalizadas (sem '#'/espaços,
# maiúsculas). Classes sem código (ex.: 4500, PMT) ou vazias → 'X'.
_CODIGO_D_CLASSE = {
    "125": "0", "150": "1", "200": "2", "300": "3", "250": "5",
    "600": "6", "800": "8", "900": "9", "1500": "A", "2500": "B", "4500": "4",
    "PN10": "C", "PN16": "D", "PN20": "E", "PN25": "F", "PN30": "G",
    "PN40": "H", "PN50": "I", "PN64": "J", "PN68": "K", "PN80": "L",
    "PN100": "M", "PN120": "N", "PN150": "O", "PN200": "P", "PN250": "Q",
    "PN420": "R",
}


def _codigo_universal_d(valvula):
    """Parte D (1 char): classe de pressão. 'X' se a classe não tiver código
    (ex.: 4500, PMT) ou estiver vazia."""
    cl = (valvula.classe or "").upper().replace("#", "").replace(" ", "").strip()
    return _CODIGO_D_CLASSE.get(cl, "X")


# tipo_extremidade (valores fixos, sem depender de ranhura/schedule) → parte EEE.
_CODIGO_E_DIRETO = {
    "SOCKET-WELDING": "SOW",
    "ROSCA NPT": "NPT",
    "ROSCA BSP": "BSP",
    "GRAY LOC HUB": "GRL",
    "LUG": "LUG",
    "Wafer": "WAF",
    "FLANGE RTJ (FJA)": "RTJ",
}

# Butt-Welding: sufixo do schedule (ex.: "BUTT-WELDING 40" → "40") → parte EEE.
# Schedules sem código próprio na tabela (50/70/110/130/150/170) e "BUTT-WELDING"
# sem sufixo (Globo Controle) caem em "BWS" (schedule a confirmar).
_CODIGO_E_BW_SCHEDULE = {
    "10": "BW1", "20": "BW2", "30": "BW3", "40": "BW4", "60": "BW6", "80": "BW8",
    "100": "B10", "120": "B12", "140": "B14", "160": "B16", "180": "B18",
}

# Flange RF/Face Plana: tipo_ranhura (acabamento da face) → parte EEE.
_CODIGO_E_FLANGE_RF = {
    "125-250 μin ESPIRAL": "FRE",
    "125-250 μin CONCÊNTRICA": "FRC",
    "LISO (125 μin)": "FRL",
    "N/A": "FRL",
}
_CODIGO_E_FLANGE_FACE_PLANA = {
    "125-250 μin ESPIRAL": "FSE",
    "125-250 μin CONCÊNTRICA": "FSC",
    "LISO (125 μin)": "FSL",
    "N/A": "FSL",
}


def _codigo_universal_e(valvula):
    """Parte E (3 chars): tipo de extremidade/conexão.
    Flangeadas (RF/Face Plana) usam tipo_ranhura p/ achar ranhura concêntrica/
    espiral/lisa. Butt-Welding usa o schedule (sufixo numérico de tipo_extremidade).
    Niple sempre 'NIP' (extensão 100mm, sem distinção por schedule na tabela).
    '' se não mapeado (não entra no código)."""
    ext = (valvula.tipo_extremidade or "").strip()
    ranhura = (valvula.tipo_ranhura or "").strip()
    if ext == "FLANGE RF":
        return _CODIGO_E_FLANGE_RF.get(ranhura, "")
    if ext == "FLANGE FACE PLANA":
        return _CODIGO_E_FLANGE_FACE_PLANA.get(ranhura, "")
    if ext.startswith("BUTT-WELDING"):
        sch = ext.replace("BUTT-WELDING", "").strip()
        return _CODIGO_E_BW_SCHEDULE.get(sch, "BWS")
    if ext.startswith("NIPLE"):
        return "NIP"
    return _CODIGO_E_DIRETO.get(ext, "")


# Material do corpo (CORPO_TAMPA) → parte FFF (3 chars). Chaves = nome do material
# como cadastrado; casadas por forma normalizada (maiúsculas, sem acento, só
# alfanumérico). Material fora desta tabela não entra no código.
_CODIGO_F_MATERIAL_RAW = {
    # Aliases exatos dos choices de MATERIAIS_CORPO_TAMPA → mesmo código do material equivalente.
    "ASTM A182 F11 CL1": "111",
    "ASTM A182 F11 CL2": "112",
    "ASTM A105N revestimento interno de INCONEL": "15I",
    "ASTM A105N revestimento em SS316": "15S",
    "ASTM A216 WCB revestimento em EPOXY": "CB2",
    "ASTM A216 WCB revestimento INCONEL": "WCI",
    "ASTM A350 LF2 Cl1 revestimento interno de INCONEL": "L2I",
    # Specs sem código prévio.
    "ASTM A182 F6a": "F6A",
    "ASTM A182 F54": "F54",
    "ASTM A217 WC5": "WC5",
    "ASTM A352 LCA": "LCA",
    "ASTM A352 CA6NM": "CA7",
    "ASTM A995 3A": "03A",
    "ASTM B564 N04400": "56N",
    "ASTM B865 N05500": "865",
    "ASTM A995 1B": "01B",
    "ASTM A995 4A": "04A",
    "ASTM A995 5A": "05A",
    "ASTM A995 6A": "06A",
    "ASTM A217 C5": "0C5",
    "ASTM A182 F5": "0F5",
    "ASTM A182 F9": "0F9",
    "ASTM A182 F11 class 1": "111",
    "ASTM A182 F11 class 2": "112",
    "ASTM A182 F11 class 3": "113",
    "ASTM A217 C12A": "12A",
    "ASTM A105N (ENP - Níquel plating)": "15E",
    "ASTM A105N (revestimento interno de INCONEL - área de vedação)": "15I",
    "ASTM A105N": "15N",
    "ASTM A105N (revestimento interno orgânico)": "15O",
    "ASTM A105N (revestimento interno de PTFE)": "15P",
    "ASTM A105": "15N",  # A105 puro = mesmo código do A105N
    "ASTM A182 F22 class 1": "221",
    "ASTM A182 F22 class 3": "223",
    "ASTM A182 F304": "304",
    "ASTM A182 F316": "316",
    "ASTM A182 F321": "321",
    "ASTM A182 F347": "347",
    "ASTM A182 F304L": "34L",
    "ASTM A182 F316L": "36L",
    "ASTM B564": "564",
    "ASTM B148 C95500": "955",
    "ASTM B148 C95800": "958",
    "ASTM A217 CA15": "A15",
    "ASTM B62": "B62",
    "ASTM A351 CF10": "C10",
    "ASTM A217 C12": "C12",
    "ASTM A351 CF10M": "C1M",
    "AISI 4130": "413",
    "ASTM A351 CF3M": "C3M",
    "ASTM A351 CF8C": "C8C",
    "ASTM A351 CF8M": "C8M",
    "ASTM A487 CA6NM": "CA6",
    "ASTM A216 WCB (revestimento interno em PTFE)": "CB1",
    "ASTM A216 WCB (revestimento interno em EPOXY)": "CB2",
    "ASTM A216 WCB (revestimento em BUNA-N)": "CB3",
    "ASTM A216 WCB (revestimento em VITON)": "CB4",
    "ASTM A216 WCB (revestimento em NEOPRENE)": "CB5",
    "ASTM A216 WCB (revestimento em EPDM)": "CB6",
    "ASTM A216 WCB ( C95800 internal overlay)": "CB7",
    "ASTM A351 CF3": "CF3",
    "ASTM A351 CF8": "CF8",
    "ASTM A182 F51": "F51",
    "ASTM A182 F53": "F53",
    "ASTM A182 F55": "F55",
    "ASTM A182 F5a": "F5A",
    "ASTM A182 F60": "F60",
    "ASTM A182 F61": "F61",
    "ASTM A182 F71": "F71",
    "ASTM A182 F91": "F91",
    "ASTM A352 LC1": "LC1",
    "ASTM A352 LC2": "LC2",
    "ASTM A352 LC3": "LC3",
    "ASTM A352 LCB": "LCB",
    "ASTM A352 LCC": "LCC",
    "ASTM A350 LF1": "LF1",
    "ASTM A350 LF2": "LF2",
    "ASTM A350 LF3": "LF3",
    "ASTM A217 WC6": "WC6",
    "ASTM A217 WC9": "WC9",
    "ASTM A216 WCB": "WCB",
    "ASTM A216 WCC": "WCC",
    "ASTM A522 type 1": "522",
    "ASTM A536 GR 65-45-12": "536",
    "ASTM A216 WCB (revestimento interno de INCONEL - área de vedação)": "WCI",
    "ASTM A351 CG8M": "G8M",
    "AASTM A351 CG8M": "G8M",  # tolera typo de cadastro
    "ASTM A182 F304H": "34H",
    "ASTM A126 cl B": "126",
    "ASTM A182 F317": "317",
    "ASTM A182 F321H": "21H",
    "ASTM A494-M35-1(Monel 400)": "494",
    "ASTM A182 gr F1": "0F1",
    "ASTM A312 Gr tp 316L": "312",
    "ASTM A358 Gr tp 316L Cl 1 or 3": "358",
    "ASTM B444 Gr1 (UNS N06625)": "444",
    "API 5L Gr B PSL 1": "A5L",
    "ASTM A333 Gr 3": "333",
    "ASTM A403 Gr WP 316L": "403",
    "ASTM B564 (UNS N06625)": "56I",
    "ASTM A234 Gr WPB": "234",
    "ASTM A420 Gr WPL3 (UNS N06625)": "42I",
    "ASTM A420 Gr WPL3": "420",
    "Aço carbono (revestido Zn-Ni)": "CSZ",
    "ASTM A105N (Galvanizado)": "15G",
    "ASTM A350 LF2 Cl1 (revestimento interno de INCONEL - área de vedação)": "L2I",
    "ASTM A351 gr CN7M": "C7M",
    "ASTM A182 F20": "F20",
    "ASTM A105N (revestimento nas áreas de vedação do corpo em SS316)": "15S",
    "ASTM A395 60-40-18": "395",
    "ASTM A182 F317L": "37L",
    "ASTM A182 310H": "31H",
    "ASTM A694 F65": "F65",
    "ASTM A694 F60": "A60",
    "ASTM A216 WCB (revestimento interno orgânico)": "CBO",
    "API 6A tipo 60K": "60K",
    "UNS N08811": "811",
    "ASTM A216 WCB (revestimento orgânico)": "WCO",
    "4140/ENP": "41E",
    "ASTM A352 LC3 + INCONEL 625": "52L",
    "ASTM A890 5A (UNS J93404)": "89A",
    "ASTM A352 LCC + INCONEL 625": "LC6",
    "ASTM A860 WPHY65": "860",
    "ASTM A403 WP347": "4C7",
    "Não aplicável (tubos e conexões)": "ZZZ",
    "ASTM A106 B": "16N",
    "API 5L X65M PSL2": "X65",
    "A350 LF2+ INCONEL 625 CLADDING(ALL WETTED PARTS)": "LFF",
    "ASTM A352 gr LCC weld overlay N00625": "LCF",
    "INCONEL 625": "INC",
    "ASTM A216 WCB+PFA": "WFA",
    "ASTM A351 CF8M+PFA": "CFA",
    "ASTM A105+13CR": "A13",
    "ASTM A350 LCB (revestimento interno de INCONEL - área de vedação)": "LCD",
    "AISI 316L": "A3L",
    "Não aplicável": "HHH",
}


def _norm_material(s):
    """Normaliza nome de material p/ lookup robusto: sem acentos, maiúsculas,
    remove token 'GR'/'GRADE' (filler, ex.: 'A216 GR WCB' = 'A216 WCB'),
    só letras/dígitos (ignora espaços/pontuação)."""
    import unicodedata
    import re
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\b(?:GRADE|GR)\b", " ", s.upper())
    return re.sub(r"[^A-Z0-9]", "", s)


_CODIGO_F_MATERIAL = {_norm_material(k): v for k, v in _CODIGO_F_MATERIAL_RAW.items()}


def _codigo_universal_f(valvula, materiais):
    """Parte FFF (3 chars): código do material do corpo (CORPO_TAMPA). '' se o
    material não estiver na tabela (esses não entram no código)."""
    return _CODIGO_F_MATERIAL.get(_norm_material(_corpo_material_nome(materiais)), "")


def _material_por_tipo(materiais, tipo):
    """Nome do material de um componente (tipo_material) ou '' se não houver."""
    for vm in materiais:
        if vm.tipo_material == tipo:
            return vm.material.nome or ""
    return ""


# Material da HASTE → parte GG (2 chars). Chaves = nome do material; casadas por
# forma normalizada com AISI≡SS. Material fora da tabela → GG vazio.
_CODIGO_G_MATERIAL_RAW = {
    "STELLITE 6": "ST", "STELLITE 21": "S2",
    "ASTM A182 F6NM": "6N",
    "SS304": "A0", "SS304L": "A1", "SS316": "A2", "SS316L": "A3",
    "SS317": "A4", "SS321": "A5", "SS347": "A6",
    "ASTM A105N": "C1", "ASTM A105N (ENP)": "C2", "ASTM A216 WCB": "C3",
    "ASTM A216 WCB (revestido em PTFE)": "C4",
    "ASTM A182 F51": "D0", "ASTM A182 F53": "D1", "ASTM A182 F55": "D2",
    "ASTM A182 F60": "D3", "ASTM A182 F61": "D4", "ASTM A182 F71": "D5",
    "INCONEL N06625": "I6", "INCONEL 718": "I8", "INCONEL N07750": "IN",
    "ASTM A182 F5": "L6", "ASTM A182 F5A": "L7", "ASTM A182 F9": "L8", "ASTM A182 F91": "L9",
    "SS410": "M1", "SS420": "M2", "SS430": "M3", "XM-19": "M4", "17-4PH": "M5",
    "MONEL": "MO", "B62": "B1", "ASTM B148 C95500": "B2", "ASTM B148 C95800": "B3",
    "ASTM B150 C63200": "B4", "ASTM A536 GR 65-45-12": "36", "ASTM A217 CA15": "15",
    "SS410 + ENP": "NI", "UNS 04400": "N1", "ASTM B148 C95200": "B5", "SS304H": "4H",
    "ASTM A182 F51 + ENP": "1E", "ASTM A182 F53 + ENP": "3E", "ASTM A182 F55 + ENP": "5E",
    "ASTM B564 (UNS N06625)": "56", "ASTM B564 type 630": "B0", "Ni UNS N10276": "N2",
    "UNS N08811": "81", "ASTM A350 LF2 (ENP)": "L2", "Padrão fabricante": "XX",
    "Não aplicável": "ZZ", "SS316 + Stellite": "6S", "4140/ENP": "4E",
    "ASTM A182 F316 + Carbeto de Tungstênio": "FC", "ASTM A182 F6A": "F6",
    "ASTM A182 F316": "F3", "SS316L + INCONEL N06625": "AI", "ASTM B574 (UNS N10276)": "57",
    "MONEL K500": "MK", "ASTM B381 F5": "B5",
    "ASTM A522 type 1 + Carbeto de Tungstênio": "5T", "A351 CF8": "B8",
    "HASTELAY C (ASTM A494 GR CW. 2MW)": "N2", "PTFE + 50% AISI 316": "R6", "CF8M": "8M",
    "A216 WCB+B62": "WB", "A216 WCB+13CR": "WC", "A105+B62": "1B", "F316L + STL": "AS",
    "CF8M + STL": "8S", "ASTM A182 F11": "F1", "F11 + Stellite": "1T", "F91 + Stellite": "9S",
    "B150 C63000": "BB", "SS431": "M6", "ASTM A182 F6": "F7",
    "ASTM A522 tipo 1 + Nitretação": "6T", "ASTM A105 + ENP + DEVLON": "5D",
    "ASTM A105 + ENP + PEEK": "5P", "ASTM A105 + ENP + RPTFE": "5R",
    "SS321H + Stellite": "1H", "SS321H": "3H", "410 SS c/ revestimento cromado": "MC",
    "AISI 4130": "I3", "ASTM A276 (UNS S32750)": "I4", "F6A+HFC": "FH",
    "INCONEL 718-API 6A": "I9", "A694 F60+TCC": "IT", "A182 F6A+PEEK+PTFE": "PP",
    "A182 F6A+RPTFE": "FT", "A276 316": "FW", "Monel 400": "MW",
    "ASTM A182 F51 + Boronizado": "D6", "A565 616HT": "H6", "17-4PH + Nitretação": "M7",
    "A182 F6A + PEEK": "FP", "A217 CA15 + Stellite": "CS", "CF8+PFA": "PF", "CF8M+PFA": "CA",
    "ASTM A182 F55 + PEEK": "5K", "ASTM A182 F53 + NYLON": "FN", "ASTM A182 F304": "04",
    "17-4PH + INCONEL 625 (todas as áreas de vedação)": "M8",
    "SS410 + INCONEL 625 (todas as áreas de vedação)": "M9",
    "SS410 + Carbeto de Tungstênio": "MD", "ASTM A276 T316": "FX",
}


def _norm_material_ss(s):
    """Como _norm_material, mas tratando AISI ≡ SS (AISI 304 → SS304)."""
    return _norm_material(s).replace("AISI", "SS")


_CODIGO_G_MATERIAL = {_norm_material_ss(k): v for k, v in _CODIGO_G_MATERIAL_RAW.items()}


def _codigo_universal_g(valvula, materiais):
    """Parte GG (2 chars): código do material da HASTE. '' se o material não
    estiver na tabela (ex.: STELLITE 6/21 → omitido)."""
    return _CODIGO_G_MATERIAL.get(_norm_material_ss(_material_por_tipo(materiais, "HASTE")), "")


# Material do OBTURADOR → parte HH (2 chars). Mesma normalização do GG (AISI≡SS;
# 'Stellite .6' casa 'STELLITE 6'). Material fora da tabela → HH vazio.
_CODIGO_H_MATERIAL_RAW = {
    "MONEL K500": "MK", "ASTM A106": "4M",
    "ASTM A217 CA15": "15", "ASTM A182 F51 + Grafite": "1G",
    "ASTM A182 F316 + Grafite": "2G", "ASTM A536 GR 65-45-12": "36",
    "ASTM A182 F53 + Grafite": "3G", "Inconel 625 + Grafite": "4G",
    "ASTM A182 F55 + Grafite": "5G", "ASTM A182 F6NM": "6N", "CF3M": "3M",
    "SS304": "A0", "SS304L": "A1", "SS316": "A2", "SS316L": "A3", "SS317": "A4",
    "SS321": "A5", "SS347": "A6", "B62": "B1", "ASTM B148 C95500": "B2",
    "ASTM B148 C95800": "B3", "BUNA N": "BU", "ASTM A105N": "C1",
    "ASTM A105N + ENP": "C2", "ASTM A216 WCB": "C3",
    "ASTM A216 WCB (revestido em PTFE)": "C4", "Cromo duro": "CR",
    "Carbeto de Tungstênio": "CT", "ASTM A182 F51": "D0", "ASTM A182 F53": "D1",
    "ASTM A182 F55": "D2", "ASTM A182 F60": "D3", "ASTM A182 F61": "D4",
    "ASTM A182 F71": "D5", "SS316 + Stellite": "6S", "DEVLON": "DV", "EPDM": "EP",
    "INCONEL 625 (UNS N06625)": "I6", "INCONEL X-750 (UNS N07750)": "IN",
    "ASTM A182 F5": "L6", "ASTM A182 F5A": "L7", "ASTM A182 F9": "L8",
    "ASTM A182 F91": "L9", "UNS 04400": "N1", "SS410": "M1", "SS420": "M2",
    "SS430": "M3", "XM-19": "M4", "17-4PH": "M5", "MONEL": "MO", "SS304 + ENP": "N4",
    "SS410 + ENP": "NI", "Neoprene": "NO", "Nylon 12": "NY", "PCTFE": "PC",
    "PEEK": "PE", "PTFE": "PT", "RPTFE (25% C)": "RC", "RPTFE (25% FV)": "RP",
    "Stellite .12": "S1", "Stellite .21": "S2", "Stellite .6": "ST", "VITON": "VI",
    "UNS N06001": "N2", "SS304H": "4H", "INCONEL 718": "I8", "ASTM A890/A995 4A": "4A",
    "ASTM A890/A995 5A": "5A", "ASTM A890/A995 6A": "6A", "SS316 + ENP": "E6",
    "ASTM A182 F51 + ENP": "1E", "ASTM A182 F53 + ENP": "3E", "ASTM A182 F55 + ENP": "5E",
    "ASTM B564 (UNS N06625)": "56", "ASTM B564 type 630": "B0", "UNS N08811": "81",
    "SS304 + Stellite": "4S", "SS410 + Stellite": "0S", "CF8M": "8M",
    "Não aplicável": "ZZ", "SS304 + carbeto de tungstênio": "4C",
    "SS316 + carbeto de tungstênio": "6C", "SS410 + carbeto de tungstênio": "0C",
    "SS410 + cromo duro": "1C", "SS304 + cromo duro": "3C", "SS316 + cromo duro": "2C",
    "ASTM A350 LF2 (ENP)": "L2", "ASTM A536 GR 65-45-12 + NYLON 11": "3N",
    "SS316 + Ni60": "6I", "Padrão fabricante": "XX", "ASTM A216 WCC": "C5",
    "ASTM A182 F316 + Carbeto de Tungstênio": "FC", "ASTM A182 F6A": "F6",
    "ASTM A182 F316": "F3", "ASTM A747 C": "A7",
    "ASTM A352 LCC + INCONEL (UNS N06625)": "A8", "A217 CA15 + Stellite": "CS",
    "A182 F6A + Stellite": "FS", "A217 WC6 + Stellite": "WS", "A217 WC9 + Stellite": "S9",
    "SS410 + INCONEL (UNS N06625)": "MI", "INCONEL 625 + Carbeto de Tungstênio": "I7",
    "ASTM A182 F55 + Carbeto de Tungstênio": "D6", "ASTM A182 F51 + TCC": "D7",
    "A995 6A + STL": "7A", "ASTM A216 WCB + STL": "A9",
    "ASTM A182 F6A + Carbeto de Tungstênio": "F9",
    "ASTM A522 type 1 + Carbeto de Tungstênio": "5T", "A105 + 13CR": "C6",
    "ASTM A479 410": "41", "ASTM A182 F304": "04", "ASTM A216 WCB + B62": "WB",
    "ASTM A216 WCB + 13CR": "WC", "ASTM A105 + B62": "1B", "ASTM A182 F316L + STL": "AS",
    "ASTM A351 CF8M + STL": "8S", "ASTM A182 F11": "F1", "ASTM A182 F11 + Stellite": "1T",
    "ASTM A182 F91 + Stellite": "9S", "M5 (17-4PH) + ST (Stellite .6)": "MT",
    "B150 C63000": "BB", "A182 F55 + STL": "FT", "ASTM A522 Type I": "FW", "SS431": "M6",
    "ASTM A105 + ENP + DEVLON": "5D", "ASTM A105 + ENP + PEEK": "5P",
    "ASTM A105 + ENP + RPTFE": "5R", "ASTM A351 CF8": "9M", "SS321H + Stellite": "1H",
    "SS321H": "3H", "INCONEL 718 + Carbeto de Tungstênio": "I9",
    "ASTM A350 Gr LF3 + revestimento de solda INCONEL 625 + Carbeto de Tungstênio": "II",
    "INCONEL 625 + RPTFE (25% C)": "IJ", "ASTM A182 F6A Cl 2 + ENP": "FZ",
    "ASTM A350 LF2 + INCONEL 625": "LF", "ASTM A352 LCB": "LL",
    "UNS 32750 + Hard Chrome": "LC", "AISI 4130": "I3", "MONEL 400 (UNS N04400)": "M7",
    "INTEGRAL+MONEL 400 (UNS N04400)": "IM", "F316+STL.6": "FG", "C5+SS304": "CB",
    "F5+STL.6": "FL", "CI+B62": "BL", "ASTM A276 (UNS S32750)": "I4", "F6A+HFC": "FH",
    "INCONEL 718-API 6A": "I9", "A694 F60+TCC": "IT", "A182 F6A+PEEK+PTFE": "PP",
    "A182 F6A+RPTFE": "FT", "ASTM A182 F316 + ENP": "6G", "ASTM A217 CA15 + SF": "16",
    "SS410 + SF": "M8", "SS316 SF": "2A", "ASTM A995 4A + SF": "8A", "A276 316": "FW",
    "Monel 400": "MW", "ASTM A995 4A + RAM 21 (TCC)": "9A", "A565 616HT": "H6",
    "AISI 410 Hardened": "H7", "ASTM A351 CF8M + FSF": "8T", "A182 F6A + PEEK": "FP",
    "ASTM A522 type 1 + carbeto de Tungstênio": "S3", "CF8+PFA": "PF", "CF8M+PFA": "CA",
    "17-4PH+Cr": "MR", "F51+Cr": "MZ", "ASTM A182 F55 + PEEK": "5K",
    "ASTM A182 F53 + NYLON": "FN", "13 CR": "PR", "SS410 + Carbeto de Cromo": "7C",
    "ASTM A890/A995 6A + Carbeto de Tungstênio": "6T", "SS316L + PTFE": "K0",
    "AISI 410 Nitretado": "MN",
}

# setdefault: primeira ocorrência vence em colisão de chave normalizada
# (ex.: 'Carbeto'/'carbeto' de Tungstênio → mantém 5T, ignora S3 redundante).
_CODIGO_H_MATERIAL = {}
for _k, _v in _CODIGO_H_MATERIAL_RAW.items():
    _CODIGO_H_MATERIAL.setdefault(_norm_material_ss(_k), _v)


def _codigo_universal_h(valvula, materiais):
    """Parte HH (2 chars): código do material do OBTURADOR. '' se o material
    não estiver na tabela."""
    return _CODIGO_H_MATERIAL.get(_norm_material_ss(_material_por_tipo(materiais, "OBTURADOR")), "")


# Material da SEDE → parte II (2 chars). Lista de (material, código) porque a
# tabela tem materiais repetidos com códigos diferentes; setdefault → 1ª vence.
_CODIGO_I_MATERIAL_RAW = [
    ("MONEL K500", "MK"),
    ("ASTM A217 CA15", "15"), ("ASTM A182 F51 + Grafite", "1G"),
    ("ASTM A182 F316 + Grafite", "2G"), ("ASTM A536 GR 65-45-12", "36"),
    ("ASTM A182 F53 + Grafite", "3G"), ("INCONEL 625 + Grafite", "4G"),
    ("ASTM A182 F55 + Grafite", "5G"), ("ASTM A182 F6NM", "6N"),
    ("SS304 + Stellite", "4S"), ("SS304", "A0"), ("SS304L", "A1"), ("SS316", "A2"),
    ("SS316L", "A3"), ("SS317", "A4"), ("SS321", "A5"), ("SS347", "A6"),
    ("B62", "B1"), ("ASTM B148 C95500", "B2"), ("ASTM B148 C95800", "B3"),
    ("BUNA N", "BU"), ("ASTM A105N", "C1"), ("ASTM A105N + ENP", "C2"),
    ("ASTM A216 WCB", "C3"), ("ASTM A216 WCB (revestido em PTFE)", "C4"),
    ("Cromo duro", "CR"), ("Carbeto de Tungstênio", "CT"), ("ASTM A182 F51", "D0"),
    ("ASTM A182 F53", "D1"), ("ASTM A182 F55", "D2"), ("ASTM A182 F60", "D3"),
    ("ASTM A182 F61", "D4"), ("ASTM A182 F71", "D5"),
    ("ASTM A182 F51 + Stellite .21", "1S"), ("DEVLON", "DV"), ("EPDM", "EP"),
    ("INCONEL (UNS N06625)", "I6"), ("INCONEL X-750 (UNS N07750)", "IN"),
    ("ASTM A182 F5", "L6"), ("ASTM A182 F5A", "L7"), ("ASTM A182 F9", "L8"),
    ("ASTM A182 F91", "L9"), ("UNS 04400", "N1"), ("SS410", "M1"), ("SS420", "M2"),
    ("SS430", "M3"), ("XM-19", "M4"), ("17-4PH", "M5"), ("MONEL", "MO"),
    ("SS304 + ENP", "N4"), ("SS410 + ENP", "NI"), ("Neoprene", "NO"),
    ("Nylon 12", "NY"), ("PCTFE", "PC"), ("PEEK", "PE"), ("PTFE", "PT"),
    ("RPTFE (25% C)", "RC"), ("RPTFE (25% FV)", "RP"), ("Stellite .12", "S1"),
    ("Stellite .21", "S2"), ("Stellite .6", "ST"), ("VITON", "VI"),
    ("UNS N06001", "N2"), ("SS304H", "4H"), ("INCONEL 718", "I8"),
    ("ASTM A890/A995 4A", "4A"), ("ASTM A890/A995 5A", "5A"),
    ("ASTM A890/A995 6A", "6A"), ("ASTM A182 F51 + ENP", "1E"),
    ("ASTM A182 F53 + ENP", "3E"), ("ASTM A182 F55 + ENP", "5E"),
    ("ASTM B564 (UNS N06625)", "56"), ("ASTM B564 type 630", "B0"),
    ("Ni UNS N10276+Grafite", "NG"), ("UNS N08811", "81"), ("SS316 + Stellite", "6S"),
    ("Sede resiliente", "SS"), ("Sede metal/metal", "MM"), ("Não aplicável", "ZZ"),
    ("SS410 + Stellite", "0S"), ("CF8M", "8M"), ("SS304 + carbeto de tungstênio", "4C"),
    ("SS316 + carbeto de tungstênio", "6C"), ("SS410 + carbeto de tungstênio", "0C"),
    ("NBR", "NB"), ("Padrão fabricante", "XX"),
    ("ASTM A182 F316 + carbeto de Tungstênio", "FC"), ("ASTM A182 F6A", "F6"),
    ("ASTM A217 CA15 + Stellite", "CS"), ("ASTM A182 F6A + Stellite", "FS"),
    ("ASTM A494 UNS N26625", "AU"),
    ("ASTM B148 C95800 + Revestimento Metálico", "BM"),
    ("ASTM B148 C95800/TC4 + Grafite", "BG"), ("SS317 + Stellite", "A7"),
    ("INTEGRAL + STL", "A8"), ("ASTM A105 + STL", "AA"),
    ("ASTM A522 type 1 + Carbeto de Tungstênio", "5T"),
    ("ASTM A351 CF8M + Carbeto de Tungstênio", "9M"),
    ("ASTM A182 F51 + Carbeto de Tungstênio", "D7"), ("ASTM A216 WCB + B62", "WB"),
    ("ASTM A216 WCB + 13CR", "WC"), ("ASTM A105 + B62", "1B"),
    ("ASTM A182 F316L + STL", "AS"), ("ASTM A351 CF8M + STL", "8S"),
    ("ASTM A182 F11", "F1"), ("ASTM A182 F11 + Stellite", "1T"),
    ("ASTM A182 F91 + Stellite", "9S"), ("INCONEL 625 + RPTFE (25% C)", "I7"),
    ("ASTM A182 F6A + DEVLON V", "I9"), ("ASTM A182 F51 + DEVLON", "I0"),
    ("ASTM A182 F316 + RPTFE", "II"), ("ASTM A182 F51 + PEEK", "P1"),
    ("B150 C63000", "BB"), ("INTEGRAL+STL. UNS 32760 + GRAPHITE", "2B"),
    ("ASTM A182 F51 + RPTFE (25% C)", "D9"),
    ("ASTM F53 + Carbeto de Tungstênio", "DO"), ("SS431", "M6"),
    ("ASTM A182 F55 + Carbeto de Tungstênio", "D6"), ("ASTM A105 + ENP + DEVLON", "5D"),
    ("ASTM A105 + ENP + PEEK", "5P"), ("ASTM A105 + ENP + RPTFE", "5R"),
    ("INTEGRAL", "IT"), ("SS321H + Stellite", "1H"), ("SS321H", "3H"),
    ("INCONEL 625 + Carbeto de Tungstênio", "I5"),
    ("INCONEL 718 + Carbeto de Tungstênio", "I4"), ("ASTM A522 Type I", "FW"),
    ("PEEK + METAL", "FM"), ("INTEGRAL + SS316", "TT"),
    ("PEEK XP108 com anéis de inserto em Inconel X750", "PI"), ("AISI 4130", "I3"),
    ("MONEL 400 (UNS N04400)", "M7"), ("INTEGRAL+MONEL 400 (UNS N04400)", "IM"),
    ("F316+STL.6", "FG"), ("C5+SS304", "CB"), ("F5+STL.6", "FL"), ("INT+B62", "BT"),
    ("ASTM A182 F6A + HFC", "FH"), ("INCONEL 718-API 6A", "I9"),
    ("A694 F60+TCC", "IT"), ("ASTM A182 F6A + PEEK + PTFE", "PP"),
    ("ASTM A182 F6A+RPTFE", "FT"),
    ("UNS S32760 (SUPER DUPLEX) + CoCr Alloy (Stellite)", "CJ"),
    ("ASTM A182 F316 + PEEK", "IJ"), ("ASTM A182 F53 + RPTFE (25% C)", "7G"),
    ("SS416", "M7"), ("ASTM A182 F53 + Stellite", "6G"), ("ASTM A217 CA15 + CP", "17"),
    ("SS410 + SF", "M8"), ("ASTM A890/A995 6A + PTFE", "7A"), ("SS410 + CP", "M9"),
    ("SS316 SF", "2A"), ("ASTM A995 4A + SF", "8A"), ("STL/UNS S32760 + Grafite", "3B"),
    ("STL/UNS S31803 + Grafite", "4B"), ("A276 316", "FW"), ("Monel 400", "MW"),
    ("SS410 + RPTFE (25% C)", "4M"), ("SS410 + Carbeto de Cromo", "2M"),
    ("ASTM A995 4A + RAM 21 (TCC)", "9A"), ("ASTM A182 F55 + PEEK", "D8"),
    ("UNS S32750 + RPTFE (25% C)", "D9"), ("A565 616HT", "H6"),
    ("ASTM A351 CF8M + CP", "8T"), ("B148 C95800 + STL.6", "9T"),
    ("ASTM A182 F6A Cl 2 + PEEK", "F2"), ("A182 F6A + PEEK", "FP"),
    ("ASTM A522 type 1 + carbeto de Tungstênio", "S3"), ("CF8+PFA", "PF"),
    ("CF8M+PFA", "CA"), ("ASTM A182 GR F316", "Q7"), ("17-4PH+PEEK", "MP"),
    ("ASTM A182 F55 + PEEK", "5K"), ("ASTM A182 F53 + NYLON", "FN"),
    ("WCB + STL", "A9"), ("ASTM B150 C63200", "B4"), ("ASTM A182 F304", "04"),
    ("ASTM A182 F55 + STL", "C7"), ("SS410 + PEEK", "3M"),
    ("ASTM A182 F55 + RPTFE (25% C)", "6K"), ("SS410 + PEEK", "4N"),
    ("SS410 + NYLON", "5N"), ("ASTM A182 F316 + NYLON", "IK"), ("SS316 + PTFE", "6D"),
    ("MONEL + PTFE", "MQ"), ("ASTM A274 T316", "16"), ("TEFLON", "TF"),
    # Aliases p/ nomes de sede que diferem da tabela base:
    ("N/A", "ZZ"),                              # valor gravado "N/A" = Não aplicável
    ("AISI 410 (ASTM A182 F6A)", "M1"),         # sede esfera → 410 inox
    ("STELLITE 6 (ASTM A564 T630)", "ST"),      # sede esfera → Stellite .6
]

_CODIGO_I_MATERIAL = {}
for _mat, _cod in _CODIGO_I_MATERIAL_RAW:
    _CODIGO_I_MATERIAL.setdefault(_norm_material_ss(_mat), _cod)


def _codigo_universal_i(valvula, materiais):
    """Parte II (2 chars): código do material da SEDE. '' se o material não
    estiver na tabela."""
    return _CODIGO_I_MATERIAL.get(_norm_material_ss(_material_por_tipo(materiais, "SEDE")), "")


# Material da GAXETA → parte J (1 char).
_CODIGO_J_GAXETA = {
    _norm_material("GRAFITE"): "0",
    _norm_material("GRAFITE FLEXÍVEL + FIO DE INCONEL"): "1",
    _norm_material("PTFE"): "2",
    _norm_material("GRAFITE FLEXÍVEL + FIO DE INCONEL C/ INIBIDOR DE CORROSÃO"): "U",
    # Código NOVO (criado aqui, não veio da tabela): "Z" segue a convenção de N/A
    # das partes GG/LL/MM ("ZZ"), reduzida ao 1 char do J.
    _norm_material("N/A"): "Z",
    # Mesmo material do "C/ INIBIDOR DE CORROSÃO" acima, com o inibidor nomeado:
    # é o rótulo usado em MATERIAIS_GAXETA_* (e forçado pela regra Gaveta + NBR).
    _norm_material("GRAFITE FLEXÍVEL + FIO DE INCONEL C/ INIBIDOR (molibdato de bário e/ou fios de zinco)"): "U",
}


def _codigo_universal_j(valvula, materiais):
    """Parte J (1 char): código do material da GAXETA. '' se não mapeado
    (molibdato, N/A → omitidos)."""
    return _CODIGO_J_GAXETA.get(_norm_material(_material_por_tipo(materiais, "GAXETA")), "")


# tipo_acionamento → parte K (1 char). "0" = N/A (eixo livre, sem atuador/volante/
# alavanca). "Z" = tipo de válvula sem esse campo (ex.: Retenção) ou vazio.
_CODIGO_K_ACIONAMENTO_RAW = {
    "N/A": "0",
    "ALAVANCA": "L",
    "VOLANTE": "V",
    "VOLANTE COM ENGRENAGEM DE REDUÇÃO": "G",
    "ATUADOR ELÉTRICO": "E",
    "ATUADOR ELÉTRICO COM VOLANTE": "E",
    "ATUADOR PNEUMÁTICO RETORNO POR MOLA TIPO PISTÃO": "P",
    "ATUADOR PNEUMÁTICO RETORNO POR MOLA COM VOLANTE TIPO PISTÃO": "P",
    "ATUADOR PNEUMÁTICO DUPLA AÇÃO TIPO PISTÃO": "P",
    "ATUADOR PNEUMÁTICO RETORNO POR MOLA TIPO DIAFRAGMA": "D",
    "ATUADOR PNEUMÁTICO RETORNO POR MOLA COM VOLANTE TIPO DIAFRAGMA": "D",
    "ATUADOR PNEUMÁTICO DUPLA AÇÃO TIPO DIAFRAGMA": "D",
    "ATUADOR ELETROHIDRÁULICO RETORNO POR MOLA": "F",
    "ATUADOR ELETROHIDRÁULICO RETORNO POR MOLA COM VOLANTE": "F",
    "ATUADOR ELETROHIDRÁULICO DUPLA AÇÃO": "F",
    "ATUADOR ELETROHIDRAULICO RETORNO POR MOLA": "F",  # variante sem acento cadastrada
    "ATUADOR HIDRÁULICO": "H",
    "ATUADOR ELETROPNEUMÁTICO": "I",
}
_CODIGO_K_ACIONAMENTO = {_norm_material(k): v for k, v in _CODIGO_K_ACIONAMENTO_RAW.items()}


def _codigo_universal_k(valvula):
    """Parte K (1 char): tipo de acionamento/atuador. 'Z' se o tipo de válvula
    não tem esse campo (ex.: Retenção) ou o campo estiver vazio."""
    ac = (valvula.tipo_acionamento or "").strip()
    if not ac:
        return "Z"
    return _CODIGO_K_ACIONAMENTO.get(_norm_material(ac), "")


# Material/vedação da JUNTA → parte LL (2 chars). Até a divisão de tipo_material
# em JUNTA (categoria: RTJ/Junta Espiralada/Pressure Seal/Castelo Soldado) e
# MATERIAL_JUNTA (composição), o select único misturava os dois — por isso
# "RTJ (FJA)" está aqui junto dos materiais de composição. N/A e "Não Aplicável"
# → mesmo código (ZZ).
_CODIGO_L_JUNTA_RAW = {
    "S32750 + GRAFITE": "53",
    "PADRÃO FABRICANTE": "00",
    "AISI 304L + GRAFITE": "1G",
    "AISI 304L + PTFE": "1P",
    "AISI 316L + GRAFITE": "3G",
    "AISI 316L + PTFE": "3P",
    "AISI 304 + GRAFITE": "4G",
    "AISI 304 + PTFE": "4P",
    "AISI 316 + GRAFITE": "6G",
    "AISI 316 + PTFE": "6P",
    "EPDM": "EP",
    "GRAFITE": "GG",
    "GRAFITE + LIP SEAL": "GL",
    "HNBR + GRAFITE": "HG",
    "HNBR": "HN",
    "SELADO À PRESSÃO": "PS",
    "PTFE": "PT",
    "RTJ (FJA)": "RT",
    "VITON + GRAFITE": "VG",
    "VITON": "VI",
    "NÃO APLICÁVEL": "ZZ",
    "N/A": "ZZ",
    "AISI 347 + GRAFITE": "7G",
    "AISI 317 + GRAFITE": "F7",
    "AISI 321 + GRAFITE": "2G",
    "AISI 321H + GRAFITE": "2H",
    "S31803 + GRAFITE": "54",
    "AISI + PTFE": "SP",
    "S32760 + PTFE": "55",
    "S32550 + PTFE": "56",
    "PAPELÃO HIDRÁULICO C/ BORRACHA SBR": "40",
    # Aliases: rótulos usados nos choices (MATERIAIS_JUNTA / VEDACAO_*) que nomeiam
    # o mesmo material das entradas acima → mesmo código.
    "AISI 304 + GRAFITE FLEXÍVEL": "4G",
    "AISI 316 + GRAFITE FLEXÍVEL": "6G",
    "PRESSURE SEAL": "PS",
    "O'RING VITON": "VI",
    # Códigos NOVOS (criados aqui, não vieram da tabela de referência): estas duas
    # vedações são opção de VEDACAO_CORPO_TAMPA_* mas não tinham linha na tabela.
    # Reconciliar com o padrão da empresa antes de tratar como oficial.
    "JUNTA ESPIRALADA": "ES",
    "CASTELO SOLDADO": "CS",
}
_CODIGO_L_JUNTA = {_norm_material(k): v for k, v in _CODIGO_L_JUNTA_RAW.items()}


def _codigo_universal_l(valvula, materiais):
    """Parte LL (2 chars): material/vedação da JUNTA. '' se não mapeado.

    Prioriza MATERIAL_JUNTA (composição, ex.: "AISI 316 + Grafite") e cai para
    JUNTA (categoria, ex.: "Castelo Soldado", sem material de composição próprio)
    quando MATERIAL_JUNTA não foi informado."""
    _mat = _material_por_tipo(materiais, "MATERIAL_JUNTA") or _material_por_tipo(materiais, "JUNTA")
    return _CODIGO_L_JUNTA.get(_norm_material(_mat), "")


# (PARAFUSOS, PORCAS) → parte MM (2 chars). Par fixo por norma (o parafuso sozinho
# não define o código — ex.: mesmo parafuso B16 casa com porca 7 (código "16") ou
# porca 2H (código "6H") dependendo da porca escolhida).
_CODIGO_M_PARAFUSO_PORCA_RAW = {
    ("ASTM A193 Gr B16", "ASTM A194 Gr 7"): "16",
    ("ASTM A193 B7", "ASTM A194 2H"): "B7",
    ("ASTM A193 B8", "ASTM A194 Gr 8"): "B8",
    ("PADRÃO FABRICANTE", "PADRÃO FABRICANTE"): "00",
    ("ASTM A193 Grade B8MA, Class 1A", "ASTM A194 Grade 8MA"): "MA",
    ("ASTM A193 B8A", "ASTM A194 8A"): "8A",
    ("ASTM A193 Gr B8M", "ASTM A194 Gr 8M"): "8M",
    ("ASTM A320 Gr L7", "ASTM A194 Gr 7"): "L7",
    ("ASTM A320 Grade L7M", "ASTM A194 Grade 7M"): "LM",
    ("ASTM A320 Grade L7M", "ASTM A194 Gr 7L"): "LL",
    ("ASTM A193 Grade B7M", "ASTM A194 Grade 2HM"): "7M",
    ("ZERON 100 FG", "ZERON 100 FG"): "FG",
    ("ASTM A193 Gr B16", "ASTM A194 2H"): "6H",
    ("ASTM A193 Gr B8M CL2", "ASTM A194 Gr 8M"): "CM",
    ("ASTM A193 B8T", "ASTM A194 8T"): "8T",
    ("UNS S32760", "UNS S32760"): "54",
    ("UNS S32550", "UNS S32550"): "55",
    ("ASTM A320 Gr L7", "ASTM A194 Gr 7L"): "7L",
    ("ASTM A193 B7", "ASTM A194 2H + HDG"): "B9",
    ("N/A", "N/A"): "ZZ",
    # Códigos NOVOS (criados aqui, não vieram da tabela de referência) — pares que
    # as regras NBR liga / NACE (ver _validar_regras_valvula) aceitam e que não
    # tinham linha. Só entram pares da mesma família metalúrgica do parafuso: as
    # regras validam parafuso e porca em sets independentes, então o cruzamento
    # também admite combos inexistentes (ex.: parafuso inox + porca carbono-molib),
    # que ficam sem código de propósito e saem "**".
    ("ASTM A320 Gr L7", "ASTM A194 Gr 4L"): "4L",
    ("ASTM A193 Gr B8M", "ASTM A194 Grade 8MA"): "MM",
    ("ASTM A193 Gr B8M CL2", "ASTM A194 Grade 8MA"): "CA",
    ("ASTM A193 Grade B7M", "ASTM A194 Grade 7M"): "M7",
    ("ASTM A193 Grade B8MA, Class 1A", "ASTM A194 Gr 8M"): "AM",
    ("ASTM A320 Grade L7M", "ASTM A194 Grade 2HM"): "LH",
    # Idem, para pares livres (fora de NBR/NACE, o formulário aceita qualquer
    # parafuso com qualquer porca) que aparecem em válvulas já cadastradas.
    ("ASTM A193 Gr B16", "ASTM A194 Gr 4"): "64",
    ("ASTM A193 Gr B16", "ASTM A194 Gr 4L"): "6L",
    ("ASTM A193 B8", "ASTM A194 Gr 8M"): "BM",
}
_CODIGO_M_PARAFUSO_PORCA = {
    (_norm_material(p), _norm_material(n)): v for (p, n), v in _CODIGO_M_PARAFUSO_PORCA_RAW.items()
}


def _codigo_universal_m(valvula, materiais):
    """Parte MM (2 chars): par (PARAFUSOS, PORCAS). '' se o par não estiver
    mapeado (peças com material só de um lado, ou combinação não catalogada)."""
    parafuso = _material_por_tipo(materiais, "PARAFUSOS")
    porca = _material_por_tipo(materiais, "PORCAS")
    if not parafuso or not porca:
        return ""
    return _CODIGO_M_PARAFUSO_PORCA.get((_norm_material(parafuso), _norm_material(porca)), "")


# Extremidade niple → código NNN. SCH 5 é opção do campo mas não tem linha na
# tabela de códigos; niple sem schedule definido (NAC) não tem campo p/ avaliar.
_CODIGO_N_NIPLE = {
    'NIPLE 4" COMP. SCH 10': "N10",
    'NIPLE 4" COMP. SCH 30': "N30",
    'NIPLE 4" COMP. SCH 40': "N40",
    'NIPLE 4" COMP. SCH 80': "N80",
    'NIPLE 4" COMP. SCH 160': "N16",
    'NIPLE 4" COMP. SCH XXS': "NXS",
}

# uso_geral só vale FSD nas normas que são de teste de fogo. As demais opções do
# campo (API 6D, ISO 14313, ISO 17292, ASME B16.34) são normas de projeto.
_USO_GERAL_FIRE_TEST = {"API 6FA", "API 607", "ISO 10497"}


def _codigo_universal_n(valvula):
    """Partes (NNN-NNN-...): notas codificadas. Lista vazia se Função ≠ Controle
    (o segmento não sai na folha). Sem limite de quantidade.

    A ordem segue a das seções da folha (ver _build_folha_grupos/_build_folha_notas):
    Corpo e Internos → Atuador → Posicionador → Solenoide → Chave Fim de Curso →
    Sensor de Posição → Instrumentação → Notas. Cada código sai na seção do campo
    que o dispara.

    Só entram os códigos da tabela que têm campo no modelo p/ avaliar; os demais
    (dreno posicionado, by-pass, dureza, rugosidade, pares de niple etc.) não têm
    origem de dado e ficam de fora."""
    if valvula.funcao != "CONTROLE":
        return []
    campos_visiveis = Valvula.CAMPOS_POR_TIPO.get(valvula.tipo_valvula, [])
    n = []

    # ── Corpo e Internos ──
    _niple = _CODIGO_N_NIPLE.get((valvula.tipo_extremidade or "").strip())
    if _niple:
        n.append(_niple)
    if "iogp" in campos_visiveis and valvula.iogp == "IOGP AS-562":
        n.append("IGP")
    if valvula.tipo_passagem == "REDUZIDA":
        n.append("REB")
    if valvula.revestimento == "ZINCO NÍQUEL":
        n.append("ZNI")
    if valvula.tipo_castelo == "EXTENDIDO":
        n.append("CAE")
    if valvula.dib == "DBB":
        n.append("DBB")

    # ── Solenoide / Chave Fim de Curso ──
    if valvula.valvula_solenoide == "SIM":
        n.append("SOL")
    if valvula.chave_fim_curso == "SIM":
        n.append("LMS")

    # ── Instrumentação ──
    if valvula.filtro == "SIM, COM MANÔMETROS":
        n.append("FIL")

    # ── Notas ──
    if valvula.certificacao_sil == "SIL 2":
        n.append("SL2")
    elif valvula.certificacao_sil == "SIL 3":
        n.append("SL3")
    # H2S e NMI compartilham o mesmo gatilho (NACE MR0175); NMI é o certificado,
    # que não tem campo próprio — só H2S sai.
    if valvula.nace == "MR0175 ISO 15156":
        n.append("H2S")
    elif valvula.nace == "MR0103 ISO 17495":
        n.append("SSC")
    if "dispositivo_antiestatico" in campos_visiveis and valvula.dispositivo_antiestatico:
        n.append("DAE")
    # FSD (design testado à fogo); FSC é o certificado do teste, sem campo próprio.
    if "uso_geral" in campos_visiveis and valvula.uso_geral in _USO_GERAL_FIRE_TEST:
        n.append("FSD")
    if valvula.nbr:
        n.append("NBR")
    return n


# Largura útil da célula do código universal, em pontos.
# PDF: col4 = 44.5% da largura útil A4 retrato (595.28 - 2×28.35 de margem), menos padding.
# Excel: coluna E, width 53 ≈ 53×7+5 px → ×72/96 pt, menos folga da borda.
_CU_CELULA_PT_PDF = (595.28 - 2 * 28.35) * 0.445 - 8
_CU_CELULA_PT_XLSX = (53 * 7 + 5) * 72 / 96 - 8
# Fonte do Excel na célula (title_font: Arial 19 bold). Arial ≈ Helvetica nas métricas.
_CU_FONT_SIZE_XLSX = 19


def _quebrar_codigo_universal(codigo, max_width_pt, font_size):
    """Quebra o código universal nas linhas que cabem em max_width_pt, medindo a
    largura real em Helvetica-Bold font_size (≈ Arial, a fonte da folha).

    A quebra tem de vir pronta do Python: o segmento (NNN-...) não tem limite de
    tamanho e o xhtml2pdf 0.2.17 ignora word-wrap/word-break — com
    text-align:center, uma string mais larga que a célula é desenhada a partir de
    um x negativo e vaza para fora da página, não só da célula.

    Prefere os separadores naturais ('.' entre segmentos, '-' entre notas) para
    não partir um código de 3 letras no meio; só corta à força um token que
    sozinho já não caiba na linha."""
    import re
    from reportlab.pdfbase.pdfmetrics import stringWidth
    if not codigo:
        return []

    def cabe(s):
        return stringWidth(s, "Helvetica-Bold", font_size) <= max_width_pt

    linhas, atual = [], ""
    for tok in (t for t in re.split(r"(?<=[.\-])", codigo) if t):
        if atual and not cabe(atual + tok):
            linhas.append(atual)
            atual = ""
        while not cabe(tok):  # token maior que a linha inteira → corta à força
            corte = len(tok) - 1
            while corte > 1 and not cabe(tok[:corte]):
                corte -= 1
            linhas.append(tok[:corte])
            tok = tok[corte:]
        atual += tok
    if atual:
        linhas.append(atual)
    return linhas


def _cu_linhas_pdf(codigo, cab_fs):
    """Linhas do código universal para a escala de fonte escolhida pelo auto-fit.
    cab_fs vem de _FOLHA_ESCALAS (varia de 14.5pt a 5pt), então o número de
    linhas muda com a escala — por isso é recalculado a cada render."""
    try:
        fs = float(str(cab_fs).replace("pt", "").strip())
    except (TypeError, ValueError):
        fs = 8.5
    return _quebrar_codigo_universal(codigo, _CU_CELULA_PT_PDF, fs)


def _gerar_pdf_excel(valvula, materiais, vedacoes, componentes):
    """Gera (pdf_data, excel_data) da folha bilíngue (PT + EN) a partir de
    objetos já carregados (válvula + relacionados). Não faz queries próprias —
    materiais/vedacoes/componentes devem vir prefetched. Retorna (None, None)
    se a geração do PDF falhar."""
    LABELS = _folha_labels_bi()
    # Código universal AAAB.CCCD.EEE.FFF... — montado em partes.
    # FFF omitido se não houver código.
    _seg1 = _codigo_universal_aaa(valvula) + _codigo_universal_b(valvula, materiais)
    _seg2 = _codigo_universal_c(valvula) + _codigo_universal_d(valvula)  # CCC + D (D sempre presente)
    _eee = _codigo_universal_e(valvula) or "***"
    _partes_cu = [_seg1, _seg2, _eee]
    _fff = _codigo_universal_f(valvula, materiais)
    if _fff:
        _partes_cu.append(_fff)
    # Segmento 5 = GGHHII: GG = haste, HH = obturador, II = sede.
    _gg = _codigo_universal_g(valvula, materiais) or "**"
    _hh = _codigo_universal_h(valvula, materiais) or "**"
    _ii = _codigo_universal_i(valvula, materiais) or "**"
    _partes_cu.append(_gg + _hh + _ii)
    # Segmento 6 = JKLLMM(NNN-NNN-...): J = gaxeta; K = acionamento; LL = junta;
    # MM = parafuso+porca; (NNN-...) = notas, só em Função = Controle.
    _j = _codigo_universal_j(valvula, materiais) or "*"
    _k = _codigo_universal_k(valvula) or "*"
    _ll = _codigo_universal_l(valvula, materiais) or "**"
    _mm = _codigo_universal_m(valvula, materiais) or "**"
    _nnn = _codigo_universal_n(valvula)
    _partes_cu.append(_j + _k + _ll + _mm + (f"({'-'.join(_nnn)})" if _nnn else ""))
    codigo_universal = ".".join(_partes_cu)

    logo_path = os.path.join(settings.BASE_DIR, "static", "assets", "logo.png")

    rate_api6d = _calc_rate_api6d(valvula, materiais, componentes)
    folha_grupos = _build_folha_grupos(valvula, materiais, vedacoes, componentes, rate_api6d, LABELS)
    folha_notas = _build_folha_notas(valvula, LABELS)
    folha_grupos, folha_notas = _numerar_folha(folha_grupos, folha_notas)
    # Notas vira mais um grupo com rótulo lateral vertical, igual Corpo/Atuador/Instrumentação.
    if folha_notas:
        folha_grupos = folha_grupos + [(LABELS["lbl_cat_notas"], folha_notas)]
    observacao = valvula.observacao or ""
    folha_grupos_ctx = _folha_grupos_ctx(folha_grupos, LABELS)

    context_base = {
        "valvula": valvula,
        "materiais": materiais,
        "vedacoes": vedacoes,
        "componentes": componentes,
        "rate_api6d": rate_api6d,
        "folha_grupos": folha_grupos_ctx,
        "folha_notas": folha_notas,
        "observacao": observacao,
        "logo_path": logo_path,
        "codigo_universal": codigo_universal,
        **LABELS,
    }

    # Auto-fit medido: renderiza em várias escalas e escolhe a maior fonte que
    # cabe em 1 página A4. Garante 1 página independente do tamanho da folha.
    _escala, pdf_data = _folha_autofit(context_base)
    if not pdf_data:
        return None, None

    # Gerar Excel idêntico ao PDF se openpyxl estiver disponível
    excel_data = None
    if OPENPYXL_AVAILABLE:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage

        wb = Workbook()
        ws = wb.active
        ws.title = "Folha de Dados"

        # Definir estilos
        thin_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # Cores do PDF
        header_bg = "D9D9D9"  # cinza (igual ao PDF)
        
        title_font = Font(size=19, bold=True, name='Arial')
        header_font = Font(size=16, bold=True, name='Arial')
        item_font = Font(size=12, name='Arial')
        desc_font = Font(size=12, name='Arial')
        footer_font = Font(size=7, color="6C757D", name='Arial')

        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Margem: coluna A (~18px) + linha 1 (~18px). Conteúdo desloca +1 col / +1 linha.
        from openpyxl.styles import Protection
        ROW_OFF = 1  # linha 1 = margem superior
        COL_OFF = 1  # coluna A = margem esquerda
        unlocked = Protection(locked=False)

        def xcell(r, c):
            """Célula lógica (1-based) com deslocamento de margem; conteúdo desbloqueado."""
            cell = ws.cell(row=r + ROW_OFF, column=c + COL_OFF)
            cell.protection = unlocked
            return cell

        def xrowdim(r):
            return ws.row_dimensions[r + ROW_OFF]

        def xmerge(r, c1, c2):
            ws.merge_cells(start_row=r + ROW_OFF, start_column=c1 + COL_OFF,
                           end_row=r + ROW_OFF, end_column=c2 + COL_OFF)

        # Larguras: A=margem; B=Categoria(vertical), C=Nº, D=ITEM, E=DESCRIÇÃO
        ws.column_dimensions[get_column_letter(1)].width = 1.86          # A margem ~18px
        ws.column_dimensions[get_column_letter(1 + COL_OFF)].width = 8   # B = Categoria (7% no PDF)
        ws.column_dimensions[get_column_letter(2 + COL_OFF)].width = 5   # C = Nº (4% no PDF)
        ws.column_dimensions[get_column_letter(3 + COL_OFF)].width = 53  # D = ITEM (44.5%)
        ws.column_dimensions[get_column_letter(4 + COL_OFF)].width = 53  # E = DESCRIÇÃO (44.5%)
        ws.row_dimensions[1].height = 13.5  # linha 1 = margem ~18px

        current_row = 1

        # Linha 1: Logo (cols 1-3, branco) + Título (col4). Linha 2: Código + valor.
        xmerge(1, 1, 3)
        logo_cell = xcell(1, 1)
        logo_cell.value = ""
        logo_cell.border = thin_border
        logo_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        xcell(1, 2).border = thin_border
        xcell(1, 3).border = thin_border
        xrowdim(1).height = 90
        xrowdim(2).height = 30

        # Rótulo "Código (Code)" (cinza, cols 1-3)
        xmerge(2, 1, 3)
        codigo_label_cell = xcell(2, 1)
        codigo_label_cell.value = "Código (Code)"
        codigo_label_cell.font = header_font
        codigo_label_cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        codigo_label_cell.alignment = center_alignment
        codigo_label_cell.border = thin_border
        xcell(2, 2).border = thin_border
        xcell(2, 3).border = thin_border

        # Tentar adicionar logo se existe (centralizado na célula do logo)
        if os.path.exists(logo_path):
            try:
                from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
                from openpyxl.utils.units import pixels_to_EMU

                img = XLImage(logo_path)
                # Manter proporção 596x283 (2.1:1)
                img_w, img_h = 210, 100  # 210 / (596/283)
                img.width = img_w
                img.height = img_h

                def _chars_to_px(chars, mdw=7):
                    """Conversão oficial Excel largura(chars) -> pixels p/ MDW=7 (Calibri 11)."""
                    return int(((256 * chars + int(128 / mdw)) / 256) * mdw)

                # Cols B, C, D (larguras 8+5+53 chars) mescladas p/ logo
                cols_px = [(COL_OFF, _chars_to_px(8)), (COL_OFF + 1, _chars_to_px(5)), (COL_OFF + 2, _chars_to_px(53))]
                row_px = 120  # 90pt * 96/72dpi
                off_x = max(0, (sum(w for _, w in cols_px) - img_w) // 2)
                off_y = max(0, (row_px - img_h) // 2)

                def _locate(px):
                    """Converte offset absoluto (px, a partir da col B) p/ (col 0-based, offset na col)."""
                    cum = 0
                    for idx, w in cols_px:
                        if px < cum + w or idx == cols_px[-1][0]:
                            return idx, max(0, px - cum)
                        cum += w
                    return cols_px[-1][0], 0

                start_col, start_off = _locate(off_x)
                end_col, end_off = _locate(off_x + img_w)

                marker_from = AnchorMarker(col=start_col, colOff=pixels_to_EMU(start_off),
                                           row=ROW_OFF, rowOff=pixels_to_EMU(off_y))
                marker_to = AnchorMarker(col=end_col, colOff=pixels_to_EMU(end_off),
                                         row=ROW_OFF, rowOff=pixels_to_EMU(off_y + img_h))
                # TwoCellAnchor com pontos explícitos em cada coluna real evita bugs de
                # colOff "estourando" a largura de uma única coluna em alguns visualizadores.
                img.anchor = TwoCellAnchor(editAs="oneCell", _from=marker_from, to=marker_to)
                ws.add_image(img)
            except Exception:
                pass  # Se falhar, continua sem logo

        # Título (DATA SHEET empilhado sob FOLHA DE DADOS) — col4
        titulo_cell = xcell(1, 4)
        titulo_cell.value = "FOLHA DE DADOS\n(DATA SHEET)"
        titulo_cell.font = title_font
        titulo_cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        titulo_cell.border = thin_border

        # Código (valor) (cinza) — col4
        codigo_cell = xcell(2, 4)
        codigo_cell.value = valvula.codigo
        codigo_cell.font = title_font
        codigo_cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        codigo_cell.alignment = center_alignment
        codigo_cell.border = thin_border

        # Linha 3: Código Universal (parcial AAA) — rótulo cols 1-3, valor col4.
        xmerge(3, 1, 3)
        cu_label_cell = xcell(3, 1)
        cu_label_cell.value = "Código Universal (Universal Code)"
        cu_label_cell.font = header_font
        cu_label_cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        cu_label_cell.alignment = center_alignment
        cu_label_cell.border = thin_border
        xcell(3, 2).border = thin_border
        xcell(3, 3).border = thin_border
        cu_val_cell = xcell(3, 4)
        # Quebra vinda do Python (não do wrap do Excel): o Excel só quebra em
        # espaço, e o código universal não tem nenhum — sem isto a string invade
        # as células vizinhas. Altura da linha acompanha o nº de linhas.
        _cu_linhas_xlsx = _quebrar_codigo_universal(
            codigo_universal, _CU_CELULA_PT_XLSX, _CU_FONT_SIZE_XLSX
        ) or [codigo_universal]
        cu_val_cell.value = "\n".join(_cu_linhas_xlsx)
        cu_val_cell.font = title_font
        cu_val_cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        cu_val_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cu_val_cell.border = thin_border
        xrowdim(3).height = max(30, 26 * len(_cu_linhas_xlsx))
        xrowdim(3).hidden = True  # oculto, nao removido (mesmo dado, so escondido)

        current_row = 4

        # Linha de cabeçalho: (Categoria vazia) | Nº | ITEM | DESCRIÇÃO
        header_cat = xcell(current_row, 1)
        header_cat.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        header_cat.border = thin_border
        header_num = xcell(current_row, 2)
        header_num.value = "Nº"
        header_num.font = header_font
        header_num.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        header_num.alignment = center_alignment
        header_num.border = thin_border

        header_item = xcell(current_row, 3)
        header_item.value = LABELS["lbl_item"]
        header_item.font = header_font
        header_item.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        header_item.alignment = center_alignment
        header_item.border = thin_border

        header_desc = xcell(current_row, 4)
        header_desc.value = LABELS["lbl_description"]
        header_desc.font = header_font
        header_desc.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
        header_desc.alignment = center_alignment
        header_desc.border = thin_border

        xrowdim(current_row).height = 25
        current_row += 1

        vertical_align = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)

        # Linha de dados (Nº | ITEM | DESCRIÇÃO) — categoria é escrita à parte.
        # O Nº já vem calculado (contínuo entre grupos e notas).
        def add_excel_row(num, item_label, item_value):
            nonlocal current_row

            num_cell = xcell(current_row, 2)
            num_cell.value = num
            num_cell.font = item_font
            num_cell.alignment = center_alignment
            num_cell.border = thin_border

            item_cell = xcell(current_row, 3)
            item_cell.value = item_label
            item_cell.font = item_font
            item_cell.alignment = left_alignment
            item_cell.border = thin_border

            desc_cell = xcell(current_row, 4)
            desc_cell.value = item_value
            desc_cell.font = desc_font
            desc_cell.alignment = left_alignment
            desc_cell.border = thin_border

            # Ajustar altura baseada no número de linhas
            desc_text = str(item_value) if item_value else ""
            lines = desc_text.count('\n') + 1 if desc_text else 1
            xrowdim(current_row).height = max(28, lines * 15)

            current_row += 1

        # Grupos (Corpo e Internos / Atuador): coluna 1 = categoria (texto vertical),
        # mesclada por todas as linhas do grupo.
        for categoria, itens in folha_grupos:
            if not itens:
                continue
            grupo_ini = current_row
            for num, item_label, item_value in itens:
                add_excel_row(num, item_label, item_value)
            grupo_fim = current_row - 1
            if grupo_fim > grupo_ini:
                ws.merge_cells(start_row=grupo_ini + ROW_OFF, start_column=1 + COL_OFF,
                               end_row=grupo_fim + ROW_OFF, end_column=1 + COL_OFF)
            cat_cell = xcell(grupo_ini, 1)
            cat_cell.value = categoria
            cat_cell.font = header_font
            cat_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cat_cell.alignment = vertical_align
            cat_cell.border = thin_border

        # Observação (Notas em si já entrou no loop de grupos acima, com rótulo
        # lateral vertical igual Corpo/Atuador/Instrumentação).
        if observacao:
            xmerge(current_row, 1, 4)
            obs_cell = xcell(current_row, 1)
            obs_cell.value = f"{LABELS['lbl_observation']}: {observacao}"
            obs_cell.font = desc_font
            obs_cell.alignment = left_alignment
            obs_cell.border = thin_border
            for cc in (2, 3, 4):
                xcell(current_row, cc).border = thin_border
            obs_lines = observacao.count('\n') + 1 + (len(observacao) // 70)
            xrowdim(current_row).height = max(40, obs_lines * 15)
            current_row += 1

        last_content_row = current_row - 1  # última linha com dado (sem rodapé vazio)

        # Garantir grade completa + fundo branco no conteúdo. Cabeçalhos/rótulos
        # cinza já foram pintados antes; só pinto branco onde ainda não há fill.
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for r in range(1, last_content_row + 1):
            for c in (1, 2, 3, 4):
                cell = xcell(r, c)
                cell.border = thin_border
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = white_fill

        # Moldura VISUAL (não protege; células editáveis). Gridlines escondidas.
        # Margem imediata (~18px) BRANCA; área não utilizada ao redor CINZA.
        ws.sheet_view.showGridLines = False
        left_col = 1                    # col A (margem esquerda)
        right_col = 4 + COL_OFF + 1     # col após o conteúdo (margem direita)
        top_row = 1                     # linha 1 (margem topo)
        bottom_row = last_content_row + ROW_OFF + 1  # linha após o conteúdo
        ws.column_dimensions[get_column_letter(right_col)].width = 1.86  # ~18px
        ws.row_dimensions[bottom_row].height = 13.5  # ~18px

        # Margem imediata = branca (limpa)
        for rr in range(top_row, bottom_row + 1):
            ws.cell(row=rr, column=left_col).fill = white_fill
            ws.cell(row=rr, column=right_col).fill = white_fill
        for cc in range(left_col, right_col + 1):
            ws.cell(row=top_row, column=cc).fill = white_fill
            ws.cell(row=bottom_row, column=cc).fill = white_fill

        # Área não utilizada ao redor (direita/abaixo) = cinza
        gray_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        gray_bottom = bottom_row + 40
        gray_right = right_col + 20
        for rr in range(top_row, gray_bottom + 1):
            for cc in range(left_col, gray_right + 1):
                if left_col <= cc <= right_col and top_row <= rr <= bottom_row:
                    continue  # dentro da tabela + margem branca
                ws.cell(row=rr, column=cc).fill = gray_fill

        # Impressão/PDF: caber toda a largura em 1 página (evita colunas em páginas
        # separadas). Retrato, A4, escala automática pela largura.
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_area = f"A1:{get_column_letter(right_col)}{bottom_row}"

        # Salvar Excel em buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()

    return pdf_data, excel_data


def _baixar_anexos_valvula(valvula):
    """Baixa os anexos da válvula do storage. Retorna lista de (nome, conteudo).
    Falha em um anexo (ou tabela ausente) não interrompe os demais."""
    try:
        anexos_qs = list(valvula.anexos.all())
    except Exception as exc:
        logger.warning("Falha ao consultar anexos da válvula %s: %s", valvula.pk, exc)
        return []
    resultado = []
    for anexo in anexos_qs:
        try:
            conteudo = storage.download(anexo.storage_key)
        except storage.StorageError as exc:
            logger.warning("Anexo %s indisponível no export: %s", anexo.storage_key, exc)
            continue
        resultado.append((anexo.nome_original, conteudo))
    return resultado


def valvula_pdf(request, pk):
    """Gera PDF bilíngue (PT + EN) da valvula via xhtml2pdf (template
    valvula_pdf.html) + Excel idêntico. Documento único."""
    valvula = get_object_or_404(Valvula.objects.prefetch_related("projetos"), pk=pk)
    materiais = ValvulaMaterial.objects.filter(valvula=valvula).select_related("material")
    vedacoes = Vedacao.objects.filter(valvula=valvula)
    componentes = ComponentesInternos.objects.filter(valvula=valvula)

    pdf_data, excel_data = _gerar_pdf_excel(valvula, materiais, vedacoes, componentes)
    if not pdf_data:
        return HttpResponse("Erro ao gerar PDF (Error generating PDF)", status=500)

    # Criar arquivo ZIP contendo PDF e Excel
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr(f"valvula_{valvula.codigo}.pdf", pdf_data)
        if excel_data:
            zip_file.writestr(f"valvula_{valvula.codigo}.xlsx", excel_data)
        for nome, conteudo in _baixar_anexos_valvula(valvula):
            zip_file.writestr(f"anexos/{nome}", conteudo)

    if request.GET.get("format") == "pdf":
        response = HttpResponse(pdf_data, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="valvula_{valvula.codigo}.pdf"'
        return response

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="valvula_{valvula.codigo}.zip"'
    return response


def _valvulas_para_export(ids):
    """Carrega válvulas + todos os relacionados usados na geração de folha numa
    única query com prefetch — evita N+1 (4 queries por válvula) no loop de export."""
    return (
        Valvula.objects.filter(pk__in=ids)
        .prefetch_related("projetos", "materiais__material", "vedacoes", "componentes", "anexos")
    )


# Workers do pool de geração paralela. O gargalo do export em lote é CPU-bound
# (xhtml2pdf faz ~4 renders por válvula no autofit) + I/O de rede (download de
# anexos). Rodar as válvulas em paralelo sobrepõe render e download.
_EXPORT_MAX_WORKERS = 4


def _gerar_export_valvula(valv):
    """Gera o pacote de export de UMA válvula já prefetched (relacionados vêm do
    cache — nenhuma query no worker). Isolado para rodar em ThreadPoolExecutor.
    Retorna (codigo, pdf_data, excel_data, anexos) ou None se o PDF falhar."""
    try:
        pdf_data, excel_data = _gerar_pdf_excel(
            valv, valv.materiais.all(), valv.vedacoes.all(), valv.componentes.all()
        )
        if not pdf_data:
            return None
        return (valv.codigo, pdf_data, excel_data, _baixar_anexos_valvula(valv))
    except Exception as exc:
        logger.warning("Falha ao gerar export da válvula %s: %s", valv.pk, exc)
        return None


def _gerar_export_lote(valvulas):
    """Gera os pacotes de export de várias válvulas em paralelo, preservando a
    ordem de entrada. Retorna lista de (codigo, pdf, excel, anexos) (sem falhas)."""
    from concurrent.futures import ThreadPoolExecutor
    workers = min(_EXPORT_MAX_WORKERS, len(valvulas)) or 1
    with ThreadPoolExecutor(max_workers=workers) as pool:
        return [r for r in pool.map(_gerar_export_valvula, valvulas) if r]


@require_POST
def valvula_export_lote(request):
    """Gera um único ZIP com PDF+Excel de várias válvulas."""
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "Sem permissão"}, status=403)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Dados inválidos"}, status=400)

    ids = body.get("ids") or []
    valvulas = list(_valvulas_para_export(ids))
    if not valvulas:
        return JsonResponse({"success": False, "error": "Nenhuma válvula selecionada"}, status=400)

    # Geração em paralelo (CPU render + I/O anexos); escrita do ZIP fica serial.
    resultados = _gerar_export_lote(valvulas)
    combined = io.BytesIO()
    with zipfile.ZipFile(combined, "w", zipfile.ZIP_DEFLATED) as zf:
        for codigo, pdf_data, excel_data, anexos in resultados:
            zf.writestr(f"{codigo}/valvula_{codigo}.pdf", pdf_data)
            if excel_data:
                zf.writestr(f"{codigo}/valvula_{codigo}.xlsx", excel_data)
            for nome, conteudo in anexos:
                zf.writestr(f"{codigo}/anexos/{nome}", conteudo)

    combined.seek(0)
    response = HttpResponse(combined.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="valvulas_lote.zip"'
    return response


@require_POST
def valvula_email(request):
    """Envia por email o(s) PDF(s) das válvulas selecionadas para o(s) destinatário(s)."""
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "Sem permissão"}, status=403)
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Dados inválidos"}, status=400)

    import re
    from django.core.validators import validate_email
    from django.core.exceptions import ValidationError as DjValidationError

    ids = body.get("ids") or []
    destinos = [e.strip() for e in re.split(r"[;,]", body.get("email", "")) if e.strip()]
    if not destinos:
        return JsonResponse({"success": False, "error": "Informe ao menos um email"}, status=400)
    for e in destinos:
        try:
            validate_email(e)
        except DjValidationError:
            return JsonResponse({"success": False, "error": f"Email inválido: {e}"}, status=400)

    valvulas = list(_valvulas_para_export(ids))
    if not valvulas:
        return JsonResponse({"success": False, "error": "Nenhuma válvula selecionada"}, status=400)

    email_msg = EmailMessage(
        subject="Folha(s) de Dados de Válvula — Imex Solutions",
        body="Segue(m) em anexo a(s) folha(s) de dados solicitada(s).",
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=destinos,
    )
    import mimetypes
    anexos = 0
    # Geração em paralelo; anexar ao email fica serial (EmailMessage não é thread-safe).
    for codigo, pdf_data, _excel_data, anexos_valv in _gerar_export_lote(valvulas):
        email_msg.attach(f"valvula_{codigo}.pdf", pdf_data, "application/pdf")
        anexos += 1
        for nome, conteudo in anexos_valv:
            mime = mimetypes.guess_type(nome)[0] or "application/octet-stream"
            email_msg.attach(f"{codigo}_{nome}", conteudo, mime)
            anexos += 1

    if anexos == 0:
        return JsonResponse({"success": False, "error": "Falha ao gerar PDFs"}, status=500)

    try:
        email_msg.send(fail_silently=False)
    except Exception as exc:
        return JsonResponse({"success": False, "error": f"Falha ao enviar email: {exc}"}, status=500)

    return JsonResponse({"success": True, "enviados": len(destinos), "anexos": anexos})



def material_criar(request):
    """Cria material via AJAX."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "errors": {"__all__": "Dados inválidos"}}, status=400)

    form = MaterialForm(data=data)
    if not form.is_valid():
        return JsonResponse({"success": False, "errors": form.errors}, status=400)

    material = form.save()
    return JsonResponse({"success": True, "material": {"id": material.id_material, "nome": material.nome}})


@require_GET
def material_lista_api(request):
    """Retorna lista de materiais em JSON."""
    materiais = Material.objects.all().order_by("nome")
    data = [{"id": m.id_material, "nome": m.nome} for m in materiais]
    return JsonResponse({"materiais": data})


def esqueci_senha_page(request):
    """Renderiza a tela de esqueci minha senha."""
    return render(request, "core/esqueci_senha.html")


@csrf_exempt
@require_POST
def esqueci_senha_api(request):
    """Solicita redefinição de senha enviando token por email."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "errors": {"__all__": "Dados inválidos"}}, status=400)

    email = data.get("email", "").strip()

    if not email:
        return JsonResponse({"success": False, "errors": {"__all__": "Email é obrigatório"}}, status=400)

    try:
        usuario = Tb_Usuario.objects.get(email=email)
    except Tb_Usuario.DoesNotExist:
        # Não revelar se o email existe ou não por segurança
        return JsonResponse({
            "success": True,
            "message": "Se esse email estiver cadastrado, você receberá instruções para redefinir sua senha."
        })

    # Verificar limite de trocas por dia (3 trocas máximas)
    hoje = timezone.now().date()
    if usuario.ultima_data_troca != hoje:
        # Novo dia, zerar contador
        usuario.trocas_senha_hoje = 0
        usuario.ultima_data_troca = hoje

    if usuario.trocas_senha_hoje >= 3:
        return JsonResponse({
            "success": False,
            "errors": {"__all__": "Limite máximo de 3 trocas de senha por dia atingido. Tente novamente amanhã."}
        }, status=429)

    # Gerar token de redefinição com expiração de 30 minutos
    token = uuid.uuid4()
    expiracao = timezone.now() + timedelta(minutes=30)
    usuario.token_verificacao = token
    usuario.token_expiracao = expiracao
    usuario.save(update_fields=['token_verificacao', 'token_expiracao'])

    # Monta e envia o email de redefinição de senha
    link = f"{settings.SITE_URL}/redefinir-senha/{token}/"

    try:
        html = f"""
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
        <meta charset="UTF-8">
        <meta http-equiv="X-UA-Compatible" content="IE=edge">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Redefinição de Senha</title>
        </head>

        <body style="margin:0;padding:0;background-color:#f4f6f9;font-family:Arial,Helvetica,sans-serif;">

        <table width="100%" border="0" cellspacing="0" cellpadding="0" bgcolor="#f4f6f9">
        <tr>
        <td align="center" style="padding:30px 15px;">

        <table width="600" border="0" cellspacing="0" cellpadding="0" bgcolor="#ffffff" style="border:1px solid #dddddd;">

            <!-- Cabeçalho -->
            <tr>
                <td align="center" bgcolor="#0b1629" style="padding:35px 20px;">

                    <h1 style="
                        margin:0;
                        color:#ffffff;
                        font-size:30px;
                        font-family:Arial,Helvetica,sans-serif;
                        font-weight:bold;
                    ">
                        Imex Solutions
                    </h1>

                    <p style="
                        margin:12px 0 0;
                        color:#c8d2e3;
                        font-size:15px;
                    ">
                        Redefinição de Senha
                    </p>

                </td>
            </tr>

            <!-- Conteúdo -->
            <tr>
                <td style="padding:40px;">

                    <h2 style="
                        margin-top:0;
                        margin-bottom:20px;
                        color:#0b1629;
                        font-size:24px;
                        font-weight:bold;
                    ">
                        Olá, {usuario.nome}
                    </h2>

                    <p style="
                        color:#555555;
                        font-size:16px;
                        line-height:26px;
                        margin:0 0 18px;
                    ">
                        Recebemos uma solicitação para redefinir a senha da sua conta.
                    </p>

                    <p style="
                        color:#555555;
                        font-size:16px;
                        line-height:26px;
                        margin:0 0 30px;
                    ">
                        Para criar uma nova senha, clique no botão abaixo:
                    </p>

                    <!-- Botão -->
                    <table align="center" border="0" cellspacing="0" cellpadding="0">
                        <tr>
                            <td bgcolor="#2c5aa0"
                                align="center"
                                style="padding:16px 34px;">

                                <a href="{link}"
                                target="_blank"
                                style="
                                        color:#ffffff;
                                        text-decoration:none;
                                        font-size:16px;
                                        font-weight:bold;
                                        font-family:Arial,Helvetica,sans-serif;
                                        display:inline-block;
                                ">
                                    Redefinir minha senha
                                </a>

                            </td>
                        </tr>
                    </table>
                        <br><br>
                    <table width="100%" border="0" cellspacing="0" cellpadding="0" bgcolor="#f8f9fb" style="border-left:4px solid #2c5aa0;margin-top:30px;">
                        <tr>
                            <td style="padding:18px;">

                                <strong style="color:#0b1629;font-size:15px;">
                                    Importante
                                </strong>

                                <p style="
                                    margin:10px 0 0;
                                    color:#666666;
                                    font-size:14px;
                                    line-height:22px;
                                ">
                                    Este link é válido por apenas
                                    <strong>30 minutos</strong> e poderá ser utilizado
                                    somente uma vez.
                                </p>

                            </td>
                        </tr>
                    </table>

                    <table width="100%" border="0" cellspacing="0" cellpadding="0" bgcolor="#f8f9fb" style="border-left:4px solid #d9534f;margin-top:20px;">
                        <tr>
                            <td style="padding:18px;">

                                <strong style="color:#0b1629;font-size:15px;">
                                    Não foi você?
                                </strong>

                                <p style="
                                    margin:10px 0 0;
                                    color:#666666;
                                    font-size:14px;
                                    line-height:22px;
                                ">
                                    Caso você não tenha solicitado esta alteração,
                                    ignore este e-mail. Sua senha permanecerá
                                    inalterada e nenhuma ação adicional será necessária.
                                </p>

                            </td>
                        </tr>
                    </table>

                    <p style="
                        margin-top:35px;
                        color:#555555;
                        font-size:15px;
                        line-height:24px;
                    ">
                        Atenciosamente,<br>
                        <strong>Equipe Imex Solutions</strong>
                    </p>

                </td>
            </tr>

            <!-- Rodapé -->
            <tr>
                <td align="center"
                    bgcolor="#f4f6f9"
                    style="
                        padding:20px;
                        color:#888888;
                        font-size:12px;
                    ">

                    © 2026 Imex Solutions<br>
                    Este é um e-mail automático. Não responda esta mensagem.

                </td>
            </tr>

        </table>

        </td>
        </tr>
        </table>

        </body>
        </html>
        """

        email_msg = EmailMessage(
            subject="Redefinição de senha — Imex Solutions",
            body=html,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email],
        )

        email_msg.content_subtype = "html"
        email_msg.send(fail_silently=False)

    except Exception:
        return JsonResponse({
            "success": False,
            "errors": {"__all__": "Não foi possível enviar o email. Tente novamente."}
        }, status=500)

    return JsonResponse({
        "success": True,
        "message": "Se esse email estiver cadastrado, você receberá instruções para redefinir sua senha."
    })

def redefinir_senha_form(request, token):
    """Renderiza o formulário de redefinição de senha com validação do token."""
    try:
        usuario = Tb_Usuario.objects.get(token_verificacao=token)
    except Tb_Usuario.DoesNotExist:
        return render(request, "core/redefinir_senha_resultado.html", {
            "sucesso": False,
            "mensagem": "Link de redefinição inválido ou expirado."
        })

    # Verificar se o token expirou
    if usuario.token_expiracao and timezone.now() > usuario.token_expiracao:
        return render(request, "core/redefinir_senha_resultado.html", {
            "sucesso": False,
            "mensagem": "Link de redefinição expirado. Solicite uma nova redefinição de senha."
        })

    return render(request, "core/redefinir_senha.html", {
        "token": token
    })


@csrf_exempt
@require_POST
def redefinir_senha_api(request):
    """Redefine a senha do usuário usando o token."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "errors": {"__all__": "Dados inválidos"}}, status=400)

    token = data.get("token")
    nova_senha = data.get("nova_senha")
    nova_senha_confirm = data.get("nova_senha_confirm")

    if not token or not nova_senha or not nova_senha_confirm:
        return JsonResponse({"success": False, "errors": {"__all__": "Todos os campos são obrigatórios"}}, status=400)

    if nova_senha != nova_senha_confirm:
        return JsonResponse({"success": False, "errors": {"nova_senha_confirm": "As senhas não coincidem"}}, status=400)

    if len(nova_senha) < 8:
        return JsonResponse({"success": False, "errors": {"nova_senha": "Senha deve ter no mínimo 8 caracteres"}}, status=400)

    try:
        usuario = Tb_Usuario.objects.get(token_verificacao=token)
    except Tb_Usuario.DoesNotExist:
        return JsonResponse({"success": False, "errors": {"__all__": "Token inválido ou expirado"}}, status=400)

    # Verificar se o token expirou
    if usuario.token_expiracao and timezone.now() > usuario.token_expiracao:
        return JsonResponse({"success": False, "errors": {"__all__": "Token expirado. Solicite uma nova redefinição de senha."}}, status=400)

    # Verificar se a nova senha não é igual à senha atual
    if usuario.check_password(nova_senha):
        return JsonResponse({"success": False, "errors": {"nova_senha": "A nova senha não pode ser igual à senha atual"}}, status=400)

    # Redefinir senha
    usuario.set_password(nova_senha)

    # Atualizar campos de controle
    usuario.senha_alterada_em = timezone.now()
    hoje = timezone.now().date()
    if usuario.ultima_data_troca != hoje:
        usuario.trocas_senha_hoje = 0
        usuario.ultima_data_troca = hoje
    usuario.trocas_senha_hoje += 1

    # Gerar novo token para invalidar o link usado e definir nova expiração (se for usar novamente)
    usuario.token_verificacao = uuid.uuid4()
    # Opcional: definir nova expiração para o token (se for reutilizado para algo mais)
    # Por enquanto, deixamos como está (será sobrescrito na próxima solicitação de esqueci senha)

    # Salvar o usuário
    usuario.save()

    # Invalidar todas as sessões do usuário para logout em todos os dispositivos
    from django.contrib.sessions.models import Session
    for session in Session.objects.all():
        session_data = session.get_decoded()
        if session_data.get('_auth_user_id') == str(usuario.id):
            session.delete()

    return JsonResponse({
        "success": True,
        "message": "Senha redefinida com sucesso! Você já pode fazer login com sua nova senha. Todas as sessões foram encerradas por segurança."
    })
# .
